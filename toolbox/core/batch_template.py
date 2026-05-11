# -*- coding: utf-8 -*-
"""
批量模板生成模块 - 唯品上新工具箱
根据输入表格（款号/类目/品牌）批量生成 QA、配件明细、试穿报告、属性表格

GUI 集成接口:
    gen = TemplateGenerator(input_path, output_dir)
    result = gen.run_all()  # 或 gen.run_qa() / gen.run_accessories() / ...
    result.summary()        # 打印摘要
"""
import os
import logging
import openpyxl
from copy import copy
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Callable

from toolbox.core.utils import find_template_dir

logger = logging.getLogger(__name__)


_TEMPLATE_DIR = find_template_dir()

QA_TEMPLATE = os.path.join(_TEMPLATE_DIR, 'QA.xlsx')
ACCESSORY_TEMPLATE = os.path.join(_TEMPLATE_DIR, '批量导入配件明细信息-模板.xlsx')
CATEGORY_MAPPING = os.path.join(_TEMPLATE_DIR, '唯品类目类型映射.xlsx')

TRYON_DIR = os.path.join(_TEMPLATE_DIR, '试穿报告')
ATTR_MALE_DIR = os.path.join(_TEMPLATE_DIR, '属性', '男装属性')
ATTR_FEMALE_DIR = os.path.join(_TEMPLATE_DIR, '属性', '女装属性')

# 试穿报告模板映射
TRYON_TEMPLATES = {
    ('女装', '女上装'): os.path.join(TRYON_DIR, '女装试穿报告(上装).xlsx'),
    ('女装', '女下装'): os.path.join(TRYON_DIR, '女装试穿报告(下装）.xlsx'),
    ('男装', '男上装'): os.path.join(TRYON_DIR, '男装试穿报告(上装).xlsx'),
    ('男装', '男下装'): os.path.join(TRYON_DIR, '男装试穿报告(下装）.xlsx'),
}


# ============ 数据结构 ============
@dataclass
class ProductInfo:
    """输入产品信息"""
    sku: str          # 唯品款号
    category: str     # 唯品类目
    brand: str        # 品牌
    washing: str = '' # 洗涤说明（可选）


@dataclass
class CategoryMapping:
    """类目映射"""
    l1_category: str    # 一级类目（女装/男装）
    template_type: str  # 模板类型（女上装/女下装/男上装/男下装）


@dataclass
class BrandInfo:
    """品牌信息"""
    supplier_id: int     # 供应商ID
    brand_sn: int        # 品牌SN
    factory: str         # 生产/经销/进口厂家


# ============ 映射加载 ============
def load_category_mapping() -> Dict[str, CategoryMapping]:
    """加载唯品类目→一级类目+模板类型的映射"""
    wb = openpyxl.load_workbook(CATEGORY_MAPPING)
    ws = wb['类目类型映射']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat, l1, tpl = row[0], row[1], row[2]
        if cat:
            cat = cat.strip()
            mapping[cat] = CategoryMapping(
                l1_category=(l1.strip() if l1 else ''),
                template_type=(tpl.strip() if tpl else ''),
            )
    wb.close()
    return mapping


def load_brand_mapping() -> Dict[str, BrandInfo]:
    """加载品牌→供应商ID+品牌SN+生产厂家的映射"""
    wb = openpyxl.load_workbook(CATEGORY_MAPPING)
    ws = wb['品牌信息']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand, sup_id, brand_sn, factory = row[0], row[1], row[2], row[3]
        if brand:
            brand = brand.strip()
            mapping[brand] = BrandInfo(
                supplier_id=sup_id,
                brand_sn=brand_sn,
                factory=(factory.strip() if factory else ''),
            )
    wb.close()
    return mapping


def load_input(input_path: str) -> List[ProductInfo]:
    """读取输入表格，按表头名精准定位列（支持任意列顺序）

    表头匹配规则（精准匹配，strip 后完全相等）：
        - "唯品款号"  — 仅匹配此名称，不匹配"款号"、"淘宝款号"等
        - "唯品类目"  — 精准匹配
        - "品牌"      — 仅匹配"品牌"，不匹配"品牌名称"、"品牌SN"等
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 读取表头，查找列索引
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row:
        wb.close()
        return []
    headers = [str(h).strip() if h else '' for h in header_row[0]]

    idx_sku = _find_exact_column(headers, '唯品款号')
    idx_cat = _find_exact_column(headers, '唯品类目')
    idx_brand = _find_exact_column(headers, '品牌')
    idx_wash = _find_exact_column(headers, '洗涤说明')  # 可选列

    if idx_sku < 0:
        raise ValueError('输入表格中未找到"唯品款号"列（注意：不是款号或淘宝款号）')
    if idx_cat < 0:
        raise ValueError('输入表格中未找到"唯品类目"列')
    if idx_brand < 0:
        raise ValueError('输入表格中未找到"品牌"列（注意：不是品牌名称或品牌SN）')

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[idx_sku] if idx_sku < len(row) else None
        cat = row[idx_cat] if idx_cat < len(row) else None
        brand = row[idx_brand] if idx_brand < len(row) else None
        wash = row[idx_wash] if idx_wash >= 0 and idx_wash < len(row) else None
        if sku:
            products.append(ProductInfo(
                sku=str(sku).strip(),
                category=str(cat).strip() if cat else '',
                brand=str(brand).strip() if brand else '',
                washing=str(wash).strip() if wash else '',
            ))
    wb.close()
    return products


def _find_exact_column(headers: list, target: str) -> int:
    """精准匹配表头列名（完全相等），返回 0-based 索引，-1 表示未找到"""
    for i, h in enumerate(headers):
        if h == target:
            return i
    return -1


# ============ QA 生成 ============
def _load_qa_template_qa_pairs() -> List[Tuple[str, str]]:
    """从 QA 模板读取固定3组问答（按索引取第一个 sheet）"""
    wb = openpyxl.load_workbook(QA_TEMPLATE)
    ws = wb.worksheets[0]  # Sheet1: 问答模板
    qa_pairs = []
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        q = row[1] if row[1] else ''
        a = row[2] if row[2] else ''
        qa_pairs.append((q, a))
    wb.close()
    return qa_pairs


def _copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def generate_qa(products: List[ProductInfo], output_path: str):
    """生成 QA 表格"""
    qa_pairs = _load_qa_template_qa_pairs()
    if not qa_pairs:
        raise ValueError("QA 模板中未找到问答数据（问答模板 sheet 的 R2-R4）")

    # 加载模板获取样式（按索引取第一个 sheet）
    wb_template = openpyxl.load_workbook(QA_TEMPLATE)
    ws_tpl = wb_template.worksheets[0]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '问答模板'

    # 写表头（复制模板 R1 的前3列）
    headers = []
    for c in range(1, 4):
        src = ws_tpl.cell(row=1, column=c)
        headers.append(src.value)
    ws.append(headers)
    ws.freeze_panes = 'A2'

    # 复制表头样式
    for c in range(1, 4):
        _copy_cell_style(ws_tpl.cell(row=1, column=c), ws.cell(row=1, column=c))

    # 每个产品3行问答
    for product in products:
        for q, a in qa_pairs:
            ws.append([product.sku, q, a])

    # 复制数据行样式（从模板 R2）
    for row_idx in range(2, ws.max_row + 1):
        for c in range(1, 4):
            _copy_cell_style(ws_tpl.cell(row=2, column=c), ws.cell(row=row_idx, column=c))

    # 列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 80

    wb_template.close()
    wb.save(output_path)
    wb.close()
    return len(products) * len(qa_pairs)


# ============ 配件明细生成 ============
def generate_accessories(products: List[ProductInfo], output_path: str) -> int:
    """生成批量导入配件明细（每个款号一行，B列填"否"）"""
    wb_template = openpyxl.load_workbook(ACCESSORY_TEMPLATE)
    ws_tpl = wb_template.worksheets[0]  # Sheet1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # 写表头（仅前5列，与模板一致）
    headers = []
    for c in range(1, 6):
        headers.append(ws_tpl.cell(row=1, column=c).value)
    ws.append(headers)
    ws.freeze_panes = 'A2'

    # 表头样式
    for c in range(1, 6):
        _copy_cell_style(ws_tpl.cell(row=1, column=c), ws.cell(row=1, column=c))

    # 每个款号一行，B列填"否"，去重
    seen = set()
    for product in products:
        if product.sku in seen:
            continue
        seen.add(product.sku)
        ws.append([product.sku, '否', None, None, None])

    # 样式
    for row_idx in range(2, ws.max_row + 1):
        for c in range(1, 3):
            _copy_cell_style(ws_tpl.cell(row=2, column=c), ws.cell(row=row_idx, column=c))

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16

    wb_template.close()
    wb.save(output_path)
    wb.close()
    return len(seen)


# ============ 试穿报告生成 ============
def _get_tryon_key(cat_map: Dict[str, CategoryMapping], product: ProductInfo) -> Optional[Tuple[str, str]]:
    """获取试穿报告分组 key: (一级类目, 模板类型)"""
    cm = cat_map.get(product.category)
    if not cm:
        print(f"  [警告] 类目 '{product.category}' 未在映射表中找到，跳过产品 {product.sku}")
        return None
    return (cm.l1_category, cm.template_type)


def generate_tryon_reports(
    products: List[ProductInfo],
    cat_map: Dict[str, CategoryMapping],
    brand_map: Dict[str, BrandInfo],
    output_dir: str,
) -> Dict[Tuple[str, str], int]:
    """
    生成试穿报告（4个文件: 男/女 × 上/下装）
    返回: {(一级类目, 模板类型): 生成行数}
    """
    # 按 (一级类目, 模板类型) 分组
    groups: Dict[Tuple[str, str], List[ProductInfo]] = {}
    skipped = []
    for product in products:
        key = _get_tryon_key(cat_map, product)
        if key is None:
            skipped.append(product.sku)
            continue
        groups.setdefault(key, []).append(product)

    result = {}
    for (l1_cat, tpl_type), group_products in groups.items():
        tpl_path = TRYON_TEMPLATES.get((l1_cat, tpl_type))
        if not tpl_path:
            print(f"  [警告] 未找到模板: ({l1_cat}, {tpl_type})")
            continue

        # 读取模板
        wb_tpl = openpyxl.load_workbook(tpl_path)
        ws_tpl = wb_tpl.worksheets[0]
        max_col = ws_tpl.max_column  # 15 或 16

        # 提取3个试穿人数据行 (R2-R4)
        person_rows = []
        for row_idx in range(2, 5):
            row_data = []
            for c in range(1, max_col + 1):
                row_data.append(ws_tpl.cell(row=row_idx, column=c).value)
            person_rows.append(row_data)

        # 创建输出文件
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = ws_tpl.title

        # 写表头 (R1)
        header = []
        for c in range(1, max_col + 1):
            header.append(ws_tpl.cell(row=1, column=c).value)
        ws_out.append(header)
        ws_out.freeze_panes = 'A2'

        # 表头样式
        for c in range(1, max_col + 1):
            _copy_cell_style(ws_tpl.cell(row=1, column=c), ws_out.cell(row=1, column=c))

        # 每个产品3行试穿数据
        row_count = 0
        for product in group_products:
            brand_info = brand_map.get(product.brand)
            if not brand_info:
                print(f"  [警告] 品牌 '{product.brand}' 未在映射表中找到，跳过产品 {product.sku}")
                continue

            for person_row in person_rows:
                new_row = list(person_row)  # 复制试穿人数据
                # 替换关键列 (0-based index)
                # B(1) = 供应商ID, C(2) = 款号, D(3) = 品牌SN
                new_row[1] = brand_info.supplier_id
                new_row[2] = product.sku
                new_row[3] = brand_info.brand_sn
                # A(0) 商品名称 留空（模板已是空）
                # E(4) 试穿者头像URL 留空（模板已是空）
                ws_out.append(new_row)
                row_count += 1

        # 数据行样式
        for row_idx in range(2, ws_out.max_row + 1):
            for c in range(1, max_col + 1):
                _copy_cell_style(ws_tpl.cell(row=2, column=c), ws_out.cell(row=row_idx, column=c))

        # 列宽
        for c in range(1, max_col + 1):
            tpl_width = ws_tpl.column_dimensions[openpyxl.utils.get_column_letter(c)].width
            if tpl_width:
                ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = tpl_width

        # 保存
        filename = f"试穿报告_{l1_cat}_{tpl_type}_生成.xlsx"
        out_path = os.path.join(output_dir, filename)
        wb_out.save(out_path)
        wb_out.close()
        wb_tpl.close()

        result[(l1_cat, tpl_type)] = row_count

    if skipped:
        print(f"  [跳过] {len(skipped)} 个产品因类目未映射: {skipped}")
    return result


# ============ 属性表格生成 ============
def _find_attr_template(category: str, l1_category: str) -> Optional[str]:
    """根据类目和一级类目找到对应的属性模板文件路径"""
    if l1_category == '女装':
        search_dir = ATTR_FEMALE_DIR
    elif l1_category == '男装':
        search_dir = ATTR_MALE_DIR
    else:
        return None

    # 直接匹配: {category}.xlsx
    direct = os.path.join(search_dir, f'{category}.xlsx')
    if os.path.exists(direct):
        return direct

    # 模糊匹配（处理名称差异）
    for fn in os.listdir(search_dir):
        if fn.startswith('~$'):
            continue
        name = os.path.splitext(fn)[0]
        if name == category or category in name or name in category:
            return os.path.join(search_dir, fn)

    return None


def _find_column_index(header_row: list, *names: str) -> int:
    """在表头行中查找匹配的列索引（0-based），返回 -1 表示未找到"""
    for i, val in enumerate(header_row):
        if val is None:
            continue
        s = str(val).strip()
        for name in names:
            if name in s:
                return i
    return -1


def generate_attributes(
    products: List[ProductInfo],
    cat_map: Dict[str, CategoryMapping],
    brand_map: Dict[str, BrandInfo],
    output_dir: str,
) -> Dict[str, int]:
    """
    生成属性表格（同类目产品合并到一个文件）
    返回: {类目名: 生成产品数}
    """
    # 按类目分组
    groups: Dict[str, List[ProductInfo]] = {}
    for product in products:
        groups.setdefault(product.category, []).append(product)

    result = {}
    skipped = []

    for category, group_products in groups.items():
        cm = cat_map.get(category)
        if not cm:
            skipped.extend([p.sku for p in group_products])
            continue

        tpl_path = _find_attr_template(category, cm.l1_category)
        if not tpl_path:
            print(f"  [警告] 未找到类目 '{category}' 的属性模板文件")
            skipped.extend([p.sku for p in group_products])
            continue

        # 读取模板
        wb_tpl = openpyxl.load_workbook(tpl_path)
        ws_tpl = wb_tpl.worksheets[0]  # 产品信息 sheet
        max_col = ws_tpl.max_column

        # 读表头
        header = []
        for c in range(1, max_col + 1):
            header.append(ws_tpl.cell(row=1, column=c).value)

        # 找到关键列索引
        idx_sku = _find_column_index(header, '款号')
        idx_category = _find_column_index(header, '产品类目', '商品类目')
        idx_factory = _find_column_index(header, '生产/经销/进口厂家')

        # 检查模板是否标准（必须有款号和产品类目列）
        if idx_sku < 0 or idx_category < 0:
            print(f"  [警告] 类目 '{category}' 的属性模板非标准结构（缺少款号/产品类目列），跳过 {len(group_products)} 个产品")
            skipped.extend([p.sku for p in group_products])
            wb_tpl.close()
            continue

        idx_brand_name = _find_column_index(header, '品牌名称')
        idx_product_name = _find_column_index(header, '商品名称')
        idx_washing = _find_column_index(header, '洗涤说明', '洗护说明', '洗涤')

        # 读模板 Row2 作为基础数据
        template_row = []
        for c in range(1, max_col + 1):
            template_row.append(ws_tpl.cell(row=2, column=c).value)

        # 创建输出
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = ws_tpl.title

        # 写表头
        ws_out.append(header)
        ws_out.freeze_panes = 'A2'
        for c in range(1, max_col + 1):
            _copy_cell_style(ws_tpl.cell(row=1, column=c), ws_out.cell(row=1, column=c))

        # 每个产品一行
        count = 0
        for product in group_products:
            brand_info = brand_map.get(product.brand)
            if not brand_info:
                print(f"  [警告] 品牌 '{product.brand}' 未映射，类目 {category} 产品 {product.sku}")
                continue

            new_row = list(template_row)  # 复制模板 Row2

            # 覆盖关键字段
            if idx_brand_name >= 0:
                new_row[idx_brand_name] = None    # 品牌名称留空
            if idx_product_name >= 0:
                new_row[idx_product_name] = None  # 商品名称留空
            if idx_sku >= 0:
                new_row[idx_sku] = product.sku
            if idx_category >= 0:
                new_row[idx_category] = product.category
            if idx_factory >= 0:
                new_row[idx_factory] = brand_info.factory
            if idx_washing >= 0 and product.washing:
                new_row[idx_washing] = product.washing

            ws_out.append(new_row)
            count += 1

        # 数据行样式
        for row_idx in range(2, ws_out.max_row + 1):
            for c in range(1, max_col + 1):
                _copy_cell_style(ws_tpl.cell(row=2, column=c), ws_out.cell(row=row_idx, column=c))

        # 列宽
        for c in range(1, max_col + 1):
            letter = openpyxl.utils.get_column_letter(c)
            tpl_width = ws_tpl.column_dimensions[letter].width
            if tpl_width:
                ws_out.column_dimensions[letter].width = tpl_width

        # 保存
        out_path = os.path.join(output_dir, f'属性_{category}.xlsx')
        wb_out.create_sheet('自定义属性')  # 需求4: 空的自定义属性sheet
        wb_out.save(out_path)
        wb_out.close()
        wb_tpl.close()

        result[category] = count

    if skipped:
        print(f"  [跳过] {len(skipped)} 个产品: {skipped}")
    return result


# ============ 统一测试入口 ============
def run_all_tests(skip_qa: bool = False, skip_accessories: bool = False,
                 skip_tryon: bool = False, skip_attr: bool = False):
    """统一测试入口"""
    test_input = os.path.join(_TEMPLATE_DIR, '测试输入表.xlsx')
    out_dir = os.path.join(_TEMPLATE_DIR, 'output')
    os.makedirs(out_dir, exist_ok=True)

    products = load_input(test_input)
    print(f"加载了 {len(products)} 个产品:")
    for p in products:
        print(f"  {p.sku} | {p.category} | {p.brand}")

    # 预加载映射
    cat_map = load_category_mapping()
    brand_map = load_brand_mapping()

    if not skip_qa:
        print("\n--- 生成 QA ---")
        n = generate_qa(products, os.path.join(out_dir, 'QA_生成测试.xlsx'))
        print(f"  QA 生成完成: {n} 行")

    if not skip_accessories:
        print("\n--- 生成配件明细 ---")
        n = generate_accessories(products, os.path.join(out_dir, '配件明细_生成测试.xlsx'))
        print(f"  配件明细生成完成: {n} 行")

    if not skip_tryon:
        print("\n--- 生成试穿报告 ---")
        result = generate_tryon_reports(products, cat_map, brand_map, out_dir)
        for key, n in result.items():
            print(f"  {key}: {n} 行")

    if not skip_attr:
        print("\n--- 生成属性表格 ---")
        result = generate_attributes(products, cat_map, brand_map, out_dir)
        for cat, n in result.items():
            print(f"  {cat}: {n} 个产品")


if __name__ == '__main__':
    run_all_tests()


# ============ GUI 集成入口 ============
@dataclass
class GenerationResult:
    """单步生成结果"""
    success: bool
    output_file: str
    row_count: int = 0
    warnings: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.success and not self.warnings


@dataclass
class BatchResult:
    """批量生成总结果"""
    input_path: str
    output_dir: str
    product_count: int = 0
    qa: Optional[GenerationResult] = None
    accessories: Optional[GenerationResult] = None
    tryon: Dict[Tuple[str, str], GenerationResult] = field(default_factory=dict)
    attributes: Dict[str, GenerationResult] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    def summary(self) -> str:
        """生成可读摘要"""
        lines = [f"=== 批量模板生成结果 ==="]
        lines.append(f"输入: {self.input_path}")
        lines.append(f"输出: {self.output_dir}")
        lines.append(f"产品数: {self.product_count}")

        if self.qa:
            status = "[OK]" if self.qa.success else "[FAIL]"
            lines.append(f"\nQA: {status} {self.qa.output_file} ({self.qa.row_count}行)")
        if self.accessories:
            status = "[OK]" if self.accessories.success else "[FAIL]"
            lines.append(f"配件明细: {status} {self.accessories.output_file} ({self.accessories.row_count}行)")
        if self.tryon:
            lines.append(f"\n试穿报告:")
            for key, r in self.tryon.items():
                status = "[OK]" if r.success else "[FAIL]"
                lines.append(f"  {key}: {status} {r.row_count}行")
        if self.attributes:
            lines.append(f"\n属性表格:")
            for cat, r in self.attributes.items():
                status = "[OK]" if r.success else "[FAIL]"
                lines.append(f"  {cat}: {status} {r.row_count}个产品")

        if self.warnings:
            lines.append(f"\n警告 ({len(self.warnings)}):")
            for w in self.warnings:
                lines.append(f"  * {w}")
        if self.errors:
            lines.append(f"\n错误 ({len(self.errors)}):")
            for e in self.errors:
                lines.append(f"  ! {e}")

        return '\n'.join(lines)


class TemplateGenerator:
    """模板批量生成器 — GUI 可直接调用

    Usage:
        gen = TemplateGenerator('input.xlsx', 'output/')
        result = gen.run_all()  # 运行全部4种
        # 或单独运行
        result = gen.run_qa()
        gen.run_accessories()
        gen.run_tryon()
        gen.run_attributes()
        print(result.summary())
    """

    def __init__(self, input_path: str, output_dir: str,
                 progress_callback: Optional[Callable[[str, int, int], None]] = None):
        """
        Args:
            input_path: 输入表格路径（款号/类目/品牌 三列）
            output_dir: 输出目录
            progress_callback: 进度回调 (step_name, current, total)
        """
        self.input_path = input_path
        self.output_dir = output_dir
        self.progress_callback = progress_callback

        os.makedirs(output_dir, exist_ok=True)

        self.products: List[ProductInfo] = []
        self.cat_map: Dict[str, CategoryMapping] = {}
        self.brand_map: Dict[str, BrandInfo] = {}
        self._loaded = False

    def _load(self):
        """加载输入数据和映射表"""
        if self._loaded:
            return
        self.products = load_input(self.input_path)
        self.cat_map = load_category_mapping()
        self.brand_map = load_brand_mapping()
        self._loaded = True

    def _progress(self, step: str, current: int, total: int):
        if self.progress_callback:
            try:
                self.progress_callback(step, current, total)
            except Exception:
                pass

    def run_qa(self) -> GenerationResult:
        """生成 QA 表格"""
        self._load()
        out = os.path.join(self.output_dir, 'QA_生成.xlsx')
        try:
            n = generate_qa(self.products, out)
            logger.info(f"QA 生成完成: {n} 行")
            return GenerationResult(success=True, output_file=out, row_count=n)
        except Exception as e:
            logger.error(f"QA 生成失败: {e}")
            return GenerationResult(success=False, output_file=out, warnings=[str(e)])

    def run_accessories(self) -> GenerationResult:
        """生成配件明细"""
        self._load()
        out = os.path.join(self.output_dir, '配件明细_生成.xlsx')
        try:
            n = generate_accessories(self.products, out)
            logger.info(f"配件明细生成完成: {n} 行")
            return GenerationResult(success=True, output_file=out, row_count=n)
        except Exception as e:
            logger.error(f"配件明细生成失败: {e}")
            return GenerationResult(success=False, output_file=out, warnings=[str(e)])

    def run_tryon(self) -> Dict[Tuple[str, str], GenerationResult]:
        """生成试穿报告（4种类型）"""
        self._load()
        result = {}
        try:
            counts = generate_tryon_reports(self.products, self.cat_map, self.brand_map, self.output_dir)
            for key, n in counts.items():
                l1, tpl = key
                fn = f"试穿报告_{l1}_{tpl}_生成.xlsx"
                result[key] = GenerationResult(
                    success=True,
                    output_file=os.path.join(self.output_dir, fn),
                    row_count=n,
                )
            logger.info(f"试穿报告生成完成: {len(counts)} 个文件")
        except Exception as e:
            logger.error(f"试穿报告生成失败: {e}")
        return result

    def run_attributes(self) -> Dict[str, GenerationResult]:
        """生成属性表格（按类目）"""
        self._load()
        result = {}
        try:
            counts = generate_attributes(self.products, self.cat_map, self.brand_map, self.output_dir)
            for cat, n in counts.items():
                result[cat] = GenerationResult(
                    success=True,
                    output_file=os.path.join(self.output_dir, f'属性_{cat}.xlsx'),
                    row_count=n,
                )
            logger.info(f"属性表格生成完成: {len(counts)} 个类目")
        except Exception as e:
            logger.error(f"属性表格生成失败: {e}")
        return result

    def run_all(self) -> BatchResult:
        """运行全部4种生成器"""
        self._load()

        bt = BatchResult(
            input_path=self.input_path,
            output_dir=self.output_dir,
            product_count=len(self.products),
        )

        # 1. QA
        self._progress('QA', 0, 4)
        bt.qa = self.run_qa()
        if not bt.qa.success:
            bt.errors.append(f"QA 失败: {bt.qa.warnings}")

        # 2. 配件明细
        self._progress('配件明细', 1, 4)
        bt.accessories = self.run_accessories()
        if not bt.accessories.success:
            bt.errors.append(f"配件明细失败: {bt.accessories.warnings}")

        # 3. 试穿报告
        self._progress('试穿报告', 2, 4)
        bt.tryon = self.run_tryon()

        # 4. 属性表格
        self._progress('属性表格', 3, 4)
        bt.attributes = self.run_attributes()

        self._progress('完成', 4, 4)

        # 收集警告
        all_warnings = []
        for r in [bt.qa, bt.accessories]:
            if r and r.warnings:
                all_warnings.extend(r.warnings)
        for r in bt.tryon.values():
            all_warnings.extend(r.warnings)
        for r in bt.attributes.values():
            all_warnings.extend(r.warnings)
        bt.warnings = all_warnings

        logger.info(f"全部生成完成: {bt.product_count} 个产品")
        return bt
