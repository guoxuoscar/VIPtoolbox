# -*- coding: utf-8 -*-
"""
ERP商品资料批量导出模块 - 唯品上新工具箱
根据商品信息表+商品库存表+选款资料表批量生成：
  1. ERP条码对照表
  2. 批量新增商品资料模板
  3. 定价导入-扣点模式模板
"""
import os
import logging
from copy import copy
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Callable

import openpyxl
from openpyxl.styles import PatternFill

from toolbox.core.size_mapping import canonicalize_size
from toolbox.core.utils import find_template_dir

logger = logging.getLogger(__name__)

# ============ 常量 ============
_SIZE_ORDER_LIST = ['XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL', '6XL', '7XL', '8XL']
SIZE_ORDER = {s: i for i, s in enumerate(_SIZE_ORDER_LIST)}

# 标准条码尾号→尺码对照: 03-S, 04-M, 05-L, 06-XL, 07-2XL, 08-3XL, 09-4XL, 10-5XL
CODE_TO_SIZE = {
    '03': 'S', '04': 'M', '05': 'L', '06': 'XL',
    '07': '2XL', '08': '3XL', '09': '4XL', '10': '5XL',
}
SIZE_TO_CODE = {v: k for k, v in CODE_TO_SIZE.items()}

# 商品名称最大字符数
MAX_NAME_LEN = 30

# 高亮颜色
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


def _size_sort_key(size: str) -> int:
    """尺码排序键"""
    return SIZE_ORDER.get(size.upper(), 999)


# ============ 模板路径查找 ============
_TEMPLATE_DIR = find_template_dir()
CATEGORY_MAPPING_PATH = os.path.join(_TEMPLATE_DIR, '唯品类目类型映射.xlsx')


# ============ 数据结构 ============
@dataclass
class SelectionRecord:
    """选款资料表一行数据"""
    款式编码: str
    唯品款号: str
    标题: str
    唯品类目: str
    品牌: str
    唯品价: float
    吊牌价: float
    尺码: str


@dataclass
class ErpRow:
    """ERP输出的一行数据"""
    款式编码: str
    商品编码: str
    唯品款号: str
    唯品货号: str
    唯品条码: str
    商品名称: str
    商品类目: str
    库存: int
    色库存: int
    orig_rest: list
    颜色: str = ''
    规格: str = ''
    规格_raw: str = ''
    highlight: bool = False
    highlight_reason: str = ''


@dataclass
class ConflictInfo:
    """冲突信息记录"""
    row_data: dict
    reason: str


# ============ 数据加载 ============
def _find_col_index(headers: list, target: str) -> int:
    for i, h in enumerate(headers):
        if h and str(h).strip() == target:
            return i
    return -1


def load_selection_table(path: str) -> Dict[str, SelectionRecord]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    idx_sku = _find_col_index(headers, '商家编码')
    idx_vsku = _find_col_index(headers, '唯品款号')
    idx_title = _find_col_index(headers, '标题')
    idx_vcat = _find_col_index(headers, '唯品类目')
    idx_brand = _find_col_index(headers, '品牌')
    idx_price = _find_col_index(headers, '唯品价')
    idx_tag = _find_col_index(headers, '吊牌价')
    idx_sizes = _find_col_index(headers, '尺码')

    if idx_sku < 0:
        raise ValueError('选款资料表中未找到"商家编码"列')
    if idx_vsku < 0:
        raise ValueError('选款资料表中未找到"唯品款号"列')

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[idx_sku]).strip() if row[idx_sku] is not None else ''
        if not sku:
            continue
        vsku = str(row[idx_vsku]).strip() if idx_vsku < len(row) and row[idx_vsku] is not None else ''
        title = str(row[idx_title]).strip() if idx_title >= 0 and idx_title < len(row) and row[idx_title] is not None else ''
        vcat = str(row[idx_vcat]).strip() if idx_vcat >= 0 and idx_vcat < len(row) and row[idx_vcat] is not None else ''
        brand = str(row[idx_brand]).strip() if idx_brand >= 0 and idx_brand < len(row) and row[idx_brand] is not None else ''
        sizes = str(row[idx_sizes]).strip() if idx_sizes >= 0 and idx_sizes < len(row) and row[idx_sizes] is not None else ''
        try:
            price = float(row[idx_price]) if idx_price >= 0 and idx_price < len(row) and row[idx_price] is not None else 0
        except (ValueError, TypeError):
            price = 0
        try:
            tag_price = float(row[idx_tag]) if idx_tag >= 0 and idx_tag < len(row) and row[idx_tag] is not None else 0
        except (ValueError, TypeError):
            tag_price = 0
        result[sku] = SelectionRecord(
            款式编码=sku, 唯品款号=vsku, 标题=title, 唯品类目=vcat,
            品牌=brand, 唯品价=price, 吊牌价=tag_price, 尺码=sizes,
        )
    wb.close()
    logger.info(f"加载选款资料: {len(result)} 条")
    return result


def load_inventory_table(path: str) -> Dict[str, Dict[str, int]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    idx_code = _find_col_index(headers, '商品编码')
    idx_public = _find_col_index(headers, '公有可用数')
    idx_purchase = _find_col_index(headers, '采购在途数')
    idx_warehouse = _find_col_index(headers, '进货仓库库存')

    if idx_code < 0:
        raise ValueError('商品库存表中未找到"商品编码"列')

    def _int_or(v, default=0):
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return default

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[idx_code]).strip() if row[idx_code] is not None else ''
        if not code:
            continue
        public = _int_or(row[idx_public]) if idx_public >= 0 and idx_public < len(row) else 0
        purchase = _int_or(row[idx_purchase]) if idx_purchase >= 0 and idx_purchase < len(row) else 0
        warehouse = _int_or(row[idx_warehouse]) if idx_warehouse >= 0 and idx_warehouse < len(row) else 0
        if code in result:
            result[code]['公有可用数'] += public
            result[code]['采购在途数'] += purchase
            result[code]['进货仓库库存'] += warehouse
        else:
            result[code] = {
                '公有可用数': public,
                '采购在途数': purchase,
                '进货仓库库存': warehouse,
            }
    wb.close()
    logger.info(f"加载库存: {len(result)} 条")
    return result


def load_product_info(path: str) -> Tuple[List[dict], dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val if val is not None else ''
        rows.append(row_dict)
    wb.close()
    logger.info(f"加载商品信息: {len(rows)} 行, {len(headers)} 列")
    return rows, header_map


def load_brand_deductions() -> Dict[str, int]:
    wb = openpyxl.load_workbook(CATEGORY_MAPPING_PATH)
    ws = wb['品牌信息']
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    idx_brand = _find_col_index(headers, '品牌')
    idx_deduction = _find_col_index(headers, '扣点')
    if idx_brand < 0:
        wb.close()
        logger.warning("品牌信息 sheet 中未找到'品牌'列，扣点全部使用默认29")
        return {}
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        brand = str(row[idx_brand]).strip() if idx_brand < len(row) and row[idx_brand] is not None else ''
        if not brand:
            continue
        try:
            d = int(float(row[idx_deduction])) if idx_deduction >= 0 and idx_deduction < len(row) and row[idx_deduction] is not None else 29
        except (ValueError, TypeError):
            d = 29
        result[brand] = d
    wb.close()
    logger.info(f"加载品牌扣点: {len(result)} 条")
    return result


# ============ 样式工具 ============
def _copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def _highlight_row(ws, row_idx: int, max_col: int):
    """将整行标记为黄色高亮"""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = YELLOW_FILL


# ============ 条码尾号规则校验 ============
def _get_barcode_ending(barcode: str) -> str:
    """提取唯品条码最后2位"""
    return barcode[-2:] if len(barcode) >= 2 else ''


def _is_standard_ending(barcode: str, size: str) -> bool:
    """检查条码尾号是否符合该尺码的标准对照规则"""
    ending = _get_barcode_ending(barcode)
    expected = SIZE_TO_CODE.get(size, '')
    return ending == expected


# ============ 核心生成逻辑 ============
def _extract_last_segment(code: str) -> str:
    idx = code.rfind('-')
    if idx < 0:
        return '-' + code
    return code[idx:]


def _compute_inventory(code: str, inv_lookup: Dict[str, Dict[str, int]]) -> int:
    inv = inv_lookup.get(code, {})
    return inv.get('公有可用数', 0) + inv.get('采购在途数', 0) + inv.get('进货仓库库存', 0)


def build_erp_rows(
    info_rows: List[dict],
    header_map: dict,
    sel_lookup: Dict[str, SelectionRecord],
    inv_lookup: Dict[str, Dict[str, int]],
    min_stock: int = 10,
) -> Tuple[List[ErpRow], List[str], List[dict]]:
    """
    构建ERP行数据（含过滤、标准化、冲突解决）
    返回: (erp_rows, warnings, conflict_rows_for_report)
    """
    warnings = []
    conflict_report = []

    all_headers = sorted(header_map.keys(), key=lambda h: header_map[h])
    orig_rest_order = [h for h in all_headers if h not in ('款式编码', '商品编码')]

    erp_rows = []

    for row_data in info_rows:
        款式编码 = str(row_data.get('款式编码', '')).strip()
        商品编码 = str(row_data.get('商品编码', '')).strip()
        if not 款式编码 or not 商品编码:
            continue

        sel = sel_lookup.get(款式编码)
        if not sel:
            warnings.append(f"款式编码 '{款式编码}' 未在选款资料中找到匹配，跳过")
            continue

        唯品款号 = sel.唯品款号
        if not 唯品款号:
            warnings.append(f"款式编码 '{款式编码}' 对应的唯品款号为空，跳过")
            continue

        尾段 = _extract_last_segment(商品编码)
        唯品条码 = 唯品款号 + 尾段
        唯品货号 = 唯品条码[:-2] if len(唯品条码) > 2 else 唯品条码
        商品名称 = sel.标题
        商品类目 = sel.唯品类目
        库存 = _compute_inventory(商品编码, inv_lookup)

        颜色 = str(row_data.get('颜色', '')).strip()
        规格_raw = str(row_data.get('规格', '')).strip()
        规格 = canonicalize_size(规格_raw) or 规格_raw

        orig_rest = [row_data.get(h, '') for h in orig_rest_order]

        row = ErpRow(
            款式编码=款式编码, 商品编码=商品编码,
            唯品款号=唯品款号, 唯品货号=唯品货号, 唯品条码=唯品条码,
            商品名称=商品名称, 商品类目=商品类目,
            库存=库存, 色库存=0,
            orig_rest=orig_rest, 颜色=颜色,
            规格=规格, 规格_raw=规格_raw,
        )

        # --- 商品名称长度检查 ---
        if len(商品名称) > MAX_NAME_LEN:
            row.highlight = True
            row.highlight_reason = f"商品名称超过{MAX_NAME_LEN}字符: {len(商品名称)}字"
            warnings.append(f"{row.highlight_reason} — 款号={唯品款号} 条码={唯品条码} 名称={商品名称}")

        # --- 条码尾号规则校验 ---
        if not _is_standard_ending(唯品条码, 规格):
            row.highlight = True
            ending = _get_barcode_ending(唯品条码)
            expected = SIZE_TO_CODE.get(规格, '无')
            reason = f"条码尾号异常: 条码={唯品条码} 尾号={ending} 规格={规格}(应为{expected})"
            if not row.highlight_reason:
                row.highlight_reason = reason
            else:
                row.highlight_reason += "; " + reason
            warnings.append(f"{reason} — 商品编码={商品编码}")

        erp_rows.append(row)

    if not erp_rows:
        return erp_rows, warnings, conflict_report

    # ---- 汇总色库存 ----
    货号库存 = {}
    for row in erp_rows:
        货号库存[row.唯品货号] = 货号库存.get(row.唯品货号, 0) + row.库存
    for row in erp_rows:
        row.色库存 = 货号库存.get(row.唯品货号, 0)

    # ---- 过滤：色库存 < min_stock ----
    before_color_filter = len(erp_rows)
    removed_货号 = {h for h, s in 货号库存.items() if s < min_stock}
    if removed_货号:
        for h in removed_货号:
            logger.info(f"色库存不足: 货号={h}, 色库存={货号库存[h]}, 阈值={min_stock}")
        erp_rows = [r for r in erp_rows if r.唯品货号 not in removed_货号]
        warnings.append(f"色库存<N({min_stock})过滤: 移除 {len(removed_货号)} 个唯品货号, {before_color_filter - len(erp_rows)} 行")

    if not erp_rows:
        return erp_rows, warnings, conflict_report

    # ---- 过滤：首尾零库存尺码 ----
    款号分组: Dict[str, List[ErpRow]] = {}
    for row in erp_rows:
        款号分组.setdefault(row.唯品款号, []).append(row)

    rows_to_remove = set()
    for vsku, group in 款号分组.items():
        sizes = sorted(set(r.规格 for r in group), key=_size_sort_key)
        removed_sizes = []
        while sizes:
            min_sz = sizes[0]
            min_total = sum(r.库存 for r in group if r.规格 == min_sz and id(r) not in rows_to_remove)
            if min_total == 0:
                removed_sizes.append(min_sz)
                sizes = sizes[1:]
                continue
            if len(sizes) > 1:
                max_sz = sizes[-1]
                max_total = sum(r.库存 for r in group if r.规格 == max_sz and id(r) not in rows_to_remove)
                if max_total == 0:
                    removed_sizes.append(max_sz)
                    sizes = sizes[:-1]
                    continue
            break
        if removed_sizes:
            for r in group:
                if r.规格 in removed_sizes:
                    rows_to_remove.add(id(r))
            logger.info(f"款号 {vsku} 首尾零库存: 删除尺码 {removed_sizes}")

    if rows_to_remove:
        erp_rows = [r for r in erp_rows if id(r) not in rows_to_remove]
        warnings.append(f"首尾零库存过滤: 移除 {len(rows_to_remove)} 行")

    if not erp_rows:
        return erp_rows, warnings, conflict_report

    # ---- 规格标准化冲突：自动解决 (优先保留标准尾号条码) ----
    dedup_groups: Dict[Tuple[str, str, str], List[ErpRow]] = {}
    for row in erp_rows:
        key = (row.唯品款号, row.唯品货号, row.规格)
        dedup_groups.setdefault(key, []).append(row)

    conflict_deleted = set()
    for key, group in dedup_groups.items():
        if len(group) <= 1:
            continue
        # 分类：标准尾号 vs 非标准尾号
        standard_rows = [r for r in group if _is_standard_ending(r.唯品条码, r.规格)]
        non_standard_rows = [r for r in group if not _is_standard_ending(r.唯品条码, r.规格)]

        vsku, vhuo, sz = key
        if standard_rows and non_standard_rows:
            # 有标准有非标准 → 删除非标准，保留标准
            for r in non_standard_rows:
                conflict_deleted.add(id(r))
                ending = _get_barcode_ending(r.唯品条码)
                reasons = []
                if r.highlight_reason:
                    reasons.append(r.highlight_reason)
                reasons.append(f"规格冲突已删除: 与标准尾号条码 {standard_rows[0].唯品条码} 规格同为{sz}")
                detail = {
                    '款号': vsku, '货号': vhuo, '规格': sz,
                    '被删除条码': r.唯品条码, '被删除尾号': ending,
                    '保留条码': standard_rows[0].唯品条码, '保留尾号': _get_barcode_ending(standard_rows[0].唯品条码),
                    '商品编码': r.商品编码, '颜色': r.颜色,
                    '原因': '; '.join(reasons),
                }
                conflict_report.append(detail)
                warnings.append(
                    f"规格标准化冲突已解决: 款号={vsku} 货号={vhuo} 规格={sz} "
                    f"删除={r.唯品条码}(尾号{ending}) 保留={standard_rows[0].唯品条码}"
                )
        else:
            # 全是标准或全非标准 → 保留所有，但告警
            barcodes = [r.唯品条码 for r in group]
            warnings.append(
                f"规格标准化冲突(无法自动解决): 款号={vsku} 货号={vhuo} 规格={sz} "
                f"条码={barcodes}"
            )

    if conflict_deleted:
        erp_rows = [r for r in erp_rows if id(r) not in conflict_deleted]
        warnings.append(f"规格冲突自动删除: {len(conflict_deleted)} 行(非标准尾号)")

    logger.info(f"ERP行数: {len(erp_rows)}, 警告数: {len(warnings)}, 冲突报告: {len(conflict_report)}")
    return erp_rows, warnings, conflict_report


def write_erp_sheet(erp_rows: List[ErpRow], header_map: dict, output_path: str,
                    input_path: str, conflict_report: List[dict] = None,
                    error_dir: str = '') -> Tuple[str, str, str]:
    """
    生成ERP条码对照表 + 报错文件
    error_dir: 报错文件输出目录，不传则与 output_path 同级
    返回: (output_path, erp_path, error_path)
    """
    logger.info(f"生成ERP条码对照表 → {output_path}")

    wb_src = openpyxl.load_workbook(input_path)
    ws_src = wb_src.active

    # ---- Sheet1: ERP ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ERP'

    erp_headers = ['款式编码', '商品编码',
                   '唯品款号', '唯品货号', '唯品条码', '商品名称', '商品类目', '库存', '色库存']
    erp_headers += [str(c.value).strip() if c.value else '' for c in ws_src[1]][2:]

    ws.append(erp_headers)
    ws.freeze_panes = 'A2'

    for c_idx in range(1, len(erp_headers) + 1):
        if c_idx <= 2:
            src_col = c_idx
        elif c_idx <= 9:
            src_col = 3
        else:
            src_col = c_idx - 7
        _copy_cell_style(ws_src.cell(row=1, column=src_col), ws.cell(row=1, column=c_idx))

    all_orig_headers = [str(c.value).strip() if c.value else '' for c in ws_src[1]]
    orig_rest_order = [h for h in all_orig_headers if h not in ('款式编码', '商品编码')]

    for row_data in erp_rows:
        row_values = [
            row_data.款式编码, row_data.商品编码,
            row_data.唯品款号, row_data.唯品货号, row_data.唯品条码,
            row_data.商品名称, row_data.商品类目,
            row_data.库存, row_data.色库存,
        ]
        rest_map = dict(zip(orig_rest_order, row_data.orig_rest))
        for h in all_orig_headers:
            if h in ('款式编码', '商品编码'):
                continue
            row_values.append(rest_map.get(h, ''))
        ws.append(row_values)

    # 高亮异常行
    for r_idx in range(2, ws.max_row + 1):
        # 通过唯品条码匹配 (列E=5)
        条码_in_sheet = str(ws.cell(row=r_idx, column=5).value).strip()
        match_row = next((r for r in erp_rows if r.唯品条码 == 条码_in_sheet), None)
        if match_row and match_row.highlight:
            _highlight_row(ws, r_idx, len(erp_headers))

    col_widths = {1: 18, 2: 24, 3: 18, 4: 22, 5: 26, 6: 40, 7: 16, 8: 10, 9: 10}
    for c, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # ---- Sheet2: 条码对照表 ----
    ws2 = wb.create_sheet('条码对照表')
    sorted_rows = sorted(erp_rows, key=lambda r: r.唯品条码)

    barcode_headers = ['唯品款号', '唯品货号', '唯品条码', '款式编码', '商品编码', '颜色', '规格']
    ws2.append(barcode_headers)
    ws2.freeze_panes = 'A2'

    for r_data in sorted_rows:
        ws2.append([
            r_data.唯品款号, r_data.唯品货号, r_data.唯品条码,
            r_data.款式编码, r_data.商品编码, r_data.颜色, r_data.规格,
        ])

    barcode_widths = {1: 18, 2: 22, 3: 26, 4: 18, 5: 24, 6: 10, 7: 10}
    for c, w in barcode_widths.items():
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    wb_src.close()
    wb.save(output_path)
    wb.close()

    # ---- 报错文件 ----
    error_path = ''
    has_conflicts = bool(conflict_report)
    has_highlights = any(r.highlight for r in erp_rows)

    if has_conflicts or has_highlights:
        if error_dir:
            error_path = os.path.join(error_dir, '生成_报错信息.xlsx')
        else:
            error_path = output_path.replace('.xlsx', '_报错信息.xlsx')
        wb_err = openpyxl.Workbook()

        # Sheet1: 已删除的冲突行
        if conflict_report:
            ws_err = wb_err.active
            ws_err.title = '已删除冲突行'
            headers = ['款号', '货号', '规格', '被删除条码', '被删除尾号',
                       '保留条码', '保留尾号', '商品编码', '颜色', '原因']
            ws_err.append(headers)
            ws_err.freeze_panes = 'A2'
            for d in conflict_report:
                ws_err.append([d.get(k, '') for k in headers])
            for c in range(1, len(headers) + 1):
                ws_err.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
            # 高亮提示
            for r_idx in range(2, ws_err.max_row + 1):
                _highlight_row(ws_err, r_idx, len(headers))
        else:
            ws_err = wb_err.active
            ws_err.title = '已删除冲突行'
            ws_err.append(['(无冲突删除记录)'])

        # Sheet2: 当前高亮行(异常但未删除)
        highlight_rows = [r for r in erp_rows if r.highlight]
        ws_hl = wb_err.create_sheet('异常高亮行')
        hl_headers = ['唯品款号', '唯品货号', '唯品条码', '款式编码', '商品编码', '颜色', '规格', '异常原因']
        ws_hl.append(hl_headers)
        ws_hl.freeze_panes = 'A2'
        for r in highlight_rows:
            ws_hl.append([
                r.唯品款号, r.唯品货号, r.唯品条码, r.款式编码,
                r.商品编码, r.颜色, r.规格, r.highlight_reason,
            ])
        for c in range(1, len(hl_headers) + 1):
            ws_hl.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
        for r_idx in range(2, ws_hl.max_row + 1):
            _highlight_row(ws_hl, r_idx, len(hl_headers))

        wb_err.save(error_path)
        wb_err.close()
        logger.info(f"报错文件已保存: {error_path}")

    logger.info(f"ERP条码对照表已保存: {output_path}")
    return output_path, output_path, error_path


def _copy_sheet_data_and_style(src_ws, dst_ws):
    """逐单元格复制内容+样式（用于跨workbook复制sheet）"""
    dst_ws.freeze_panes = src_ws.freeze_panes
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, dst_cell)
    # 列宽
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width


def generate_batch_add_template(erp_rows: List[ErpRow], sel_lookup: Dict[str, SelectionRecord],
                                output_path: str, template_path: str) -> str:
    """
    生成批量新增商品资料模板
    使用 .xlsx 模板，完整保留表头格式、填充色、字体、边框等
    """
    logger.info(f"生成批量新增商品资料 → {output_path}")

    wb_tpl = openpyxl.load_workbook(template_path)
    ws_tpl = wb_tpl.worksheets[0]
    tpl_headers = [str(c.value).strip() if c.value else '' for c in ws_tpl[1]]

    # 找关键列索引
    idx_brand_name = _find_col_index(tpl_headers, '品牌名称')
    idx_sku = _find_col_index(tpl_headers, '款号')
    idx_huohao = _find_col_index(tpl_headers, '货号')
    idx_barcode = _find_col_index(tpl_headers, '条形码')
    idx_product_name = _find_col_index(tpl_headers, '商品名称')
    idx_category = _find_col_index(tpl_headers, '商品类目')
    idx_custom_size = _find_col_index(tpl_headers, '自定义尺码')
    idx_custom_color = _find_col_index(tpl_headers, '自定义颜色')

    # 读模板 Row2 作为默认值
    tpl_row1 = [ws_tpl.cell(row=2, column=c).value for c in range(1, ws_tpl.max_column + 1)]

    wb = openpyxl.Workbook()

    # ---- Sheet1: 商品资料 ----
    ws1 = wb.active
    ws1.title = ws_tpl.title

    # 写表头 (复制模板格式)
    for c in range(1, ws_tpl.max_column + 1):
        src = ws_tpl.cell(row=1, column=c)
        dst = ws1.cell(row=1, column=c, value=src.value)
        _copy_cell_style(src, dst)
    ws1.freeze_panes = ws_tpl.freeze_panes or 'A2'

    # 写数据行
    for row_data in erp_rows:
        sel = sel_lookup.get(row_data.款式编码)
        brand_name = sel.品牌 if sel else ''

        new_row = list(tpl_row1)
        if idx_brand_name >= 0:
            new_row[idx_brand_name] = brand_name or ''
        if idx_sku >= 0:
            new_row[idx_sku] = row_data.唯品款号
        if idx_huohao >= 0:
            new_row[idx_huohao] = row_data.唯品货号
        if idx_barcode >= 0:
            new_row[idx_barcode] = row_data.唯品条码
        if idx_product_name >= 0:
            # 从选款资料表匹配标题作为商品名称，超长自动截断
            name = row_data.商品名称 or ''
            if len(name) > MAX_NAME_LEN:
                name = name[:MAX_NAME_LEN]
            new_row[idx_product_name] = name
        if idx_category >= 0:
            new_row[idx_category] = row_data.商品类目
        if idx_custom_size >= 0:
            new_row[idx_custom_size] = row_data.规格
        if idx_custom_color >= 0:
            new_row[idx_custom_color] = row_data.颜色
        new_row = [v if v is not None else '' for v in new_row]
        ws1.append(new_row)

    # 数据行格式（从模板 Row2 复制）
    for r_idx in range(2, ws1.max_row + 1):
        for c in range(1, ws_tpl.max_column + 1):
            _copy_cell_style(ws_tpl.cell(row=2, column=c), ws1.cell(row=r_idx, column=c))

    # 列宽
    for col_letter, dim in ws_tpl.column_dimensions.items():
        if dim.width:
            ws1.column_dimensions[col_letter].width = dim.width

    # ---- Sheet2: 商品类目表 (从模板复制，保留格式) ----
    ws2 = wb.create_sheet(wb_tpl.sheetnames[1])
    _copy_sheet_data_and_style(wb_tpl.worksheets[1], ws2)

    # ---- Sheet3: 填写说明 (从模板复制，保留格式) ----
    ws3 = wb.create_sheet(wb_tpl.sheetnames[2])
    _copy_sheet_data_and_style(wb_tpl.worksheets[2], ws3)

    wb_tpl.close()
    wb.save(output_path)
    wb.close()
    logger.info(f"批量新增商品资料已保存: {output_path}")
    return output_path


def generate_pricing_template(erp_rows: List[ErpRow], sel_lookup: Dict[str, SelectionRecord],
                              brand_deductions: Dict[str, int],
                              output_path: str, template_path: str) -> str:
    logger.info(f"生成定价导入模板 → {output_path}")
    wb_tpl = openpyxl.load_workbook(template_path)
    ws_tpl = wb_tpl.active
    tpl_headers = [str(c.value).strip() if c.value else '' for c in ws_tpl[1]]

    idx_barcode = _find_col_index(tpl_headers, '条形码')
    idx_sale_price = _find_col_index(tpl_headers, '销售价')
    idx_market_price = _find_col_index(tpl_headers, '市场价')
    idx_deduction = _find_col_index(tpl_headers, '扣点')
    idx_currency = _find_col_index(tpl_headers, '币种')
    if idx_currency < 0:
        idx_currency = _find_col_index(tpl_headers, '货币')

    if idx_barcode < 0:
        wb_tpl.close()
        raise ValueError('定价导入模板中未找到"条形码"列')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_tpl.title
    ws.append(tpl_headers)
    ws.freeze_panes = 'A2'

    for c in range(1, len(tpl_headers) + 1):
        _copy_cell_style(ws_tpl.cell(row=1, column=c), ws.cell(row=1, column=c))

    seen_barcodes = set()
    for row_data in erp_rows:
        barcode = row_data.唯品条码
        if barcode in seen_barcodes:
            continue
        seen_barcodes.add(barcode)
        sel = sel_lookup.get(row_data.款式编码)
        new_row = [''] * len(tpl_headers)
        if idx_barcode >= 0:
            new_row[idx_barcode] = barcode
        if idx_sale_price >= 0 and sel:
            new_row[idx_sale_price] = sel.唯品价
        if idx_market_price >= 0 and sel:
            new_row[idx_market_price] = sel.吊牌价
        if idx_deduction >= 0 and sel:
            new_row[idx_deduction] = brand_deductions.get(sel.品牌, 29)
        if idx_currency >= 0:
            new_row[idx_currency] = 'CNY'
        ws.append(new_row)

    for c in range(1, len(tpl_headers) + 1):
        try:
            w = ws_tpl.column_dimensions[openpyxl.utils.get_column_letter(c)].width
            if w:
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
        except Exception:
            pass

    wb_tpl.close()
    wb.save(output_path)
    wb.close()
    logger.info(f"定价导入模板已保存: {output_path}")
    return output_path


# ============ 结果类型 ============
@dataclass
class ErpGenerationResult:
    success: bool
    output_file: str
    row_count: int = 0
    warnings: List[str] = field(default_factory=list)
    error_file: str = ''


@dataclass
class ErpBatchResult:
    erp_barcode: Optional[ErpGenerationResult] = None
    batch_add: Optional[ErpGenerationResult] = None
    pricing: Optional[ErpGenerationResult] = None
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    error_file: str = ''

    @property
    def has_critical_warnings(self) -> bool:
        """是否有需要高亮弹窗的严重警告"""
        for w in self.warnings:
            if '冲突' in w or '异常' in w or '超过' in w or '已删除' in w:
                return True
        return False

    def summary(self) -> str:
        lines = ["=== ERP商品资料批量导出结果 ==="]
        if self.erp_barcode:
            s = "[OK]" if self.erp_barcode.success else "[FAIL]"
            lines.append(f"ERP条码对照表: {s} {self.erp_barcode.output_file} ({self.erp_barcode.row_count}行)")
        if self.batch_add:
            s = "[OK]" if self.batch_add.success else "[FAIL]"
            lines.append(f"批量新增商品资料: {s} {self.batch_add.output_file} ({self.batch_add.row_count}行)")
        if self.pricing:
            s = "[OK]" if self.pricing.success else "[FAIL]"
            lines.append(f"定价导入模板: {s} {self.pricing.output_file} ({self.pricing.row_count}行)")
        if self.error_file:
            lines.append(f"\n报错文件: {self.error_file}")
        if self.warnings:
            lines.append(f"\n警告 ({len(self.warnings)}):")
            for w in self.warnings[:20]:
                lines.append(f"  * {w}")
            if len(self.warnings) > 20:
                lines.append(f"  ... 等共 {len(self.warnings)} 条")
        if self.errors:
            lines.append(f"\n错误 ({len(self.errors)}):")
            for e in self.errors:
                lines.append(f"  ! {e}")
        return '\n'.join(lines)


# ============ 主生成器类 ============
class ErpProductGenerator:
    def __init__(self, info_path: str, inventory_path: str, selection_path: str,
                 output_dir: str, min_stock: int = 10,
                 progress_callback: Optional[Callable[[str, int, int], None]] = None,
                 error_dir: str = ''):
        self.info_path = info_path
        self.inventory_path = inventory_path
        self.selection_path = selection_path
        self.output_dir = output_dir
        self.min_stock = min_stock
        self.progress_callback = progress_callback
        self.error_dir = error_dir

        os.makedirs(output_dir, exist_ok=True)

        self._loaded = False
        self._info_rows: List[dict] = []
        self._header_map: dict = {}
        self._sel_lookup: Dict[str, SelectionRecord] = {}
        self._inv_lookup: Dict[str, Dict[str, int]] = {}
        self._brand_deductions: Dict[str, int] = {}
        self._erp_rows: List[ErpRow] = []
        self._warnings: List[str] = []
        self._conflict_report: List[dict] = []

    def _load(self):
        if self._loaded:
            return
        self._info_rows, self._header_map = load_product_info(self.info_path)
        self._sel_lookup = load_selection_table(self.selection_path)
        self._inv_lookup = load_inventory_table(self.inventory_path)
        self._brand_deductions = load_brand_deductions()
        self._erp_rows, self._warnings, self._conflict_report = build_erp_rows(
            self._info_rows, self._header_map,
            self._sel_lookup, self._inv_lookup,
            self.min_stock,
        )
        self._loaded = True

    def _progress(self, step: str, current: int, total: int):
        if self.progress_callback:
            try:
                self.progress_callback(step, current, total)
            except Exception:
                pass

    def _find_batch_add_template(self) -> str:
        """查找批量新增商品资料模板，优先 .xlsx 再 .xls"""
        # 优先：表格模板目录中的 .xlsx
        candidates = [
            os.path.join(_TEMPLATE_DIR, '批量新增商品资料.xlsx'),
            os.path.join(os.path.dirname(self.info_path), '批量新增商品资料.xlsx'),
        ]
        for p in candidates:
            if os.path.exists(p):
                return p

        # 回退搜索
        for search_dir in [os.path.dirname(self.info_path), _TEMPLATE_DIR]:
            if not os.path.isdir(search_dir):
                continue
            for fn in os.listdir(search_dir):
                if '批量新增' in fn and fn.lower().endswith('.xlsx'):
                    return os.path.join(search_dir, fn)
        raise FileNotFoundError('未找到 批量新增商品资料.xlsx 模板')

    def _find_pricing_template(self) -> str:
        """查找定价导入模板，优先 表格模板"""
        candidates = [
            os.path.join(_TEMPLATE_DIR, '定价导入—扣点模式.xlsx'),
            os.path.join(os.path.dirname(self.info_path), '定价导入—扣点模式.xlsx'),
        ]
        for p in candidates:
            if os.path.exists(p):
                return p
        for search_dir in [os.path.dirname(self.info_path), _TEMPLATE_DIR]:
            if not os.path.isdir(search_dir):
                continue
            for fn in os.listdir(search_dir):
                if '定价导入' in fn and '扣点' in fn and fn.endswith('.xlsx'):
                    return os.path.join(search_dir, fn)
        raise FileNotFoundError('未找到 定价导入—扣点模式.xlsx 模板')

    def run_erp_barcode(self) -> ErpGenerationResult:
        self._load()
        out = os.path.join(self.output_dir, 'ERP条码对照表_生成.xlsx')
        try:
            _, _, error_path = write_erp_sheet(
                self._erp_rows, self._header_map, out, self.info_path,
                conflict_report=self._conflict_report,
                error_dir=self.error_dir,
            )
            logger.info(f"ERP条码对照表生成完成: {len(self._erp_rows)} 行")
            return ErpGenerationResult(
                success=True, output_file=out,
                row_count=len(self._erp_rows),
                warnings=self._warnings,
                error_file=error_path,
            )
        except Exception as e:
            logger.error(f"ERP条码对照表生成失败: {e}")
            return ErpGenerationResult(
                success=False, output_file=out,
                warnings=[str(e)],
            )

    def run_batch_add(self) -> ErpGenerationResult:
        self._load()
        out = os.path.join(self.output_dir, '批量新增商品资料_生成.xlsx')
        try:
            tpl_path = self._find_batch_add_template()
            generate_batch_add_template(self._erp_rows, self._sel_lookup, out, tpl_path)
            logger.info(f"批量新增商品资料生成完成: {len(self._erp_rows)} 行")
            return ErpGenerationResult(success=True, output_file=out, row_count=len(self._erp_rows))
        except Exception as e:
            logger.error(f"批量新增商品资料生成失败: {e}")
            return ErpGenerationResult(success=False, output_file=out, warnings=[str(e)])

    def run_pricing(self) -> ErpGenerationResult:
        self._load()
        out = os.path.join(self.output_dir, '定价导入_生成.xlsx')
        try:
            tpl_path = self._find_pricing_template()
            unique_barcodes = set(r.唯品条码 for r in self._erp_rows)
            generate_pricing_template(self._erp_rows, self._sel_lookup,
                                      self._brand_deductions, out, tpl_path)
            logger.info(f"定价导入生成完成: {len(unique_barcodes)} 条")
            return ErpGenerationResult(success=True, output_file=out, row_count=len(unique_barcodes))
        except Exception as e:
            logger.error(f"定价导入生成失败: {e}")
            return ErpGenerationResult(success=False, output_file=out, warnings=[str(e)])

    def run_all(self) -> ErpBatchResult:
        self._load()
        bt = ErpBatchResult()

        self._progress('ERP条码对照表', 0, 3)
        bt.erp_barcode = self.run_erp_barcode()
        if not bt.erp_barcode.success:
            bt.errors.append(f"ERP条码对照表失败: {bt.erp_barcode.warnings}")

        self._progress('批量新增商品资料', 1, 3)
        bt.batch_add = self.run_batch_add()
        if not bt.batch_add.success:
            bt.errors.append(f"批量新增商品资料失败: {bt.batch_add.warnings}")

        self._progress('定价导入', 2, 3)
        bt.pricing = self.run_pricing()
        if not bt.pricing.success:
            bt.errors.append(f"定价导入失败: {bt.pricing.warnings}")

        self._progress('完成', 3, 3)

        all_warnings = []
        for r in [bt.batch_add, bt.pricing]:
            if r and r.warnings:
                all_warnings.extend(r.warnings)
        if bt.erp_barcode and bt.erp_barcode.warnings:
            all_warnings = bt.erp_barcode.warnings + all_warnings
        bt.warnings = all_warnings

        if bt.erp_barcode and bt.erp_barcode.error_file:
            bt.error_file = bt.erp_barcode.error_file

        logger.info(f"ERP商品资料导出完成")
        return bt


if __name__ == '__main__':
    import logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

    erp_dir = r'E:\ai 编程测试\唯品上新工具箱4.29\erp商品资料'
    test_info = os.path.join(erp_dir, '商品信息_2026-05-05_14-20-03.xlsx')
    test_inv = os.path.join(erp_dir, '商品库存_2026-05-05_14-21-14.xlsx')
    test_sel = os.path.join(erp_dir, '4.17港仔唯品选款P1.xlsx')
    test_out = os.path.join(erp_dir, 'output_test')

    gen = ErpProductGenerator(test_info, test_inv, test_sel, test_out, min_stock=10)
    result = gen.run_all()
    print(result.summary())
