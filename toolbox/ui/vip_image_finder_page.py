# -*- coding: utf-8 -*-
"""
唯品批量找图 - 从共享盘批量查找并复制商品原图
"""
import copy
import logging
import os
import re
import shutil
import threading
import time
import traceback
from difflib import SequenceMatcher

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtCore import Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from toolbox.core.utils import save_config
from toolbox.ui.path_drop import DirDropLineEdit, ExcelDropLineEdit


LOGGER = logging.getLogger("vip_image_finder")


默认品牌配置表 = {
    "港仔文艺男": {
        "共享盘路径": [r"\\192.168.31.100\运营中心\运营一部\港仔文艺男【勿删】\000   港仔文艺男上新【重要勿删】"],
        "品牌关键词": ["港仔文艺男", "港仔", "GZ", "gangzai", "GANGZAI"],
        "输出名称": "港仔文艺男",
    },
    "BEETLE TOWN": {
        "共享盘路径": [
            r"\\192.168.31.100\运营中心\运营一部\BEETLE TOWN项目组【勿删】\0 得物\得物  BEETLE TOWN【勿删】\3 淘宝上新【勿删】\上新套版",
            r"\\192.168.31.100\运营中心\运营一部\BEETLE TOWN项目组【勿删】\0 得物\得物  BEETLE TOWN【勿删】\5 得物套版【勿删】",
        ],
        "品牌关键词": ["BEETLE TOWN", "BT", "BEETLE", "beetle", "beetle town"],
        "输出名称": "BT",
    },
    "Dream made": {
        "共享盘路径": [r"\\192.168.31.100\运营中心\运营五部\00 新项目 Dream made\00-C店上新"],
        "品牌关键词": ["Dream made", "DM", "DREAM MADE", "dream made", "dream"],
        "输出名称": "DM",
    },
}

默认品牌名映射 = {
    "港仔文艺男": "港仔文艺男",
    "港仔": "港仔文艺男",
    "GZ": "港仔文艺男",
    "gangzai": "港仔文艺男",
    "BEETLE TOWN": "BEETLE TOWN",
    "BT": "BEETLE TOWN",
    "BEETLE": "BEETLE TOWN",
    "beetle": "BEETLE TOWN",
    "Dream made": "Dream made",
    "DM": "Dream made",
    "DREAM": "Dream made",
    "dream made": "Dream made",
}

跳过目录 = {
    "主图", "详情", "SKU", "sku", "视频", "直播", "素材", "模板",
    "临时", "备份", "回收站", "Thumbs.db", ".DS_Store",
    "历史", "归档", "完成", "已上架", "待上架", "未上架",
}


class VipImageFinderPage(QWidget):
    """唯品批量找图页面"""

    log_signal = Signal(str)
    progress_signal = Signal(int)
    done_signal = Signal(tuple)
    error_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.品牌配置表 = self._加载品牌配置()
        self.品牌名映射 = self._构建品牌名映射()
        self.search_thread = None
        self.should_stop = False
        self.索引 = None
        self._index_cache = {}
        self._cache_expiry = 24 * 3600
        self._init_ui()
        self._connect_signals()

    def _init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        self.setLayout(layout)

        title = QLabel("唯品批量找图")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1E88E5;")
        layout.addWidget(title)

        hint = QLabel("从Excel读取款号，在共享盘中查找对应原图，复制到本地")
        hint.setFont(QFont("Microsoft YaHei", 9))
        hint.setStyleSheet("color: #666;")
        layout.addWidget(hint)

        brand_group = QGroupBox("品牌选择（加载Excel后自动识别）")
        brand_layout = QHBoxLayout()
        brand_group.setLayout(brand_layout)
        self.brand_combo = QComboBox()
        self.brand_combo.setFont(QFont("Microsoft YaHei", 10))
        self.brand_combo.addItems(["自动检测"] + list(self.品牌配置表.keys()))
        self.brand_combo.setCurrentText("自动检测")
        self.brand_label = QLabel("加载Excel后自动显示路径")
        self.brand_label.setStyleSheet("color: #666;")
        self.brand_label.setWordWrap(True)
        brand_layout.addWidget(QLabel("选择品牌:"))
        brand_layout.addWidget(self.brand_combo)
        brand_layout.addWidget(self.brand_label)
        brand_layout.addStretch()
        layout.addWidget(brand_group)

        excel_group = QGroupBox("Excel文件")
        excel_layout = QHBoxLayout()
        excel_group.setLayout(excel_layout)
        self.excel_edit = ExcelDropLineEdit()
        self.excel_edit.excel_dropped.connect(self._on_excel_dropped)
        self.browse_btn = QPushButton("📂 浏览")
        self.browse_btn.clicked.connect(self._browse_excel)
        excel_layout.addWidget(self.excel_edit)
        excel_layout.addWidget(self.browse_btn)
        layout.addWidget(excel_group)

        output_group = QGroupBox("输出设置")
        output_layout = QHBoxLayout()
        output_group.setLayout(output_layout)
        self.output_edit = DirDropLineEdit("拖拽目录到此处，或点击浏览...（默认: Excel所在目录下的[日期_品牌_原图]文件夹）")
        self.output_edit.dir_dropped.connect(lambda p: self._log(f"输出目录: {p}"))
        self.output_btn = QPushButton("📂 浏览")
        self.output_btn.clicked.connect(self._browse_output)
        output_layout.addWidget(QLabel("输出目录:"))
        output_layout.addWidget(self.output_edit)
        output_layout.addWidget(self.output_btn)
        layout.addWidget(output_group)

        path_group = QGroupBox("共享盘路径维护（支持手动修改，无需改代码）")
        path_layout = QVBoxLayout()
        path_group.setLayout(path_layout)
        path_tip = QLabel("当前品牌的每一行表示一个共享盘路径；可直接粘贴多行")
        path_tip.setStyleSheet("color: #666;")
        path_layout.addWidget(path_tip)
        self.path_editor = QTextEdit()
        self.path_editor.setPlaceholderText("每行一个路径，例如：\\\\192.168.x.x\\xxx\\xxx")
        self.path_editor.setMinimumHeight(90)
        path_layout.addWidget(self.path_editor)
        path_btn_row = QHBoxLayout()
        self.path_add_btn = QPushButton("➕ 添加路径")
        self.path_save_btn = QPushButton("💾 保存当前品牌路径")
        self.path_add_btn.clicked.connect(self._添加路径到编辑框)
        self.path_save_btn.clicked.connect(self._保存当前品牌路径)
        path_btn_row.addWidget(self.path_add_btn)
        path_btn_row.addWidget(self.path_save_btn)
        path_btn_row.addStretch()
        path_layout.addLayout(path_btn_row)
        layout.addWidget(path_group)

        button_layout = QHBoxLayout()
        self.start_btn = QPushButton("▶️ 开始查找")
        self.start_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.start_btn.setMinimumHeight(45)
        self.start_btn.setStyleSheet("QPushButton{background-color:#1E88E5;color:white;border:none;border-radius:6px;padding:10px 20px;}QPushButton:hover{background-color:#1976D2;}QPushButton:disabled{background-color:#BDBDBD;}")
        self.start_btn.clicked.connect(self._start_search)
        self.stop_btn = QPushButton("⏹️ 停止")
        self.stop_btn.setFont(QFont("Microsoft YaHei", 11))
        self.stop_btn.setMinimumHeight(45)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("QPushButton{background-color:#F44336;color:white;border:none;border-radius:6px;padding:10px 20px;}QPushButton:hover{background-color:#D32F2F;}QPushButton:disabled{background-color:#BDBDBD;}")
        self.stop_btn.clicked.connect(self._stop_search)
        button_layout.addStretch()
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.stop_btn)
        layout.addLayout(button_layout)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)

        log_group = QGroupBox("处理日志")
        log_layout = QVBoxLayout()
        log_group.setLayout(log_layout)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        log_layout.addWidget(self.log_text)
        layout.addWidget(log_group)

        self._on_brand_changed()
        self.brand_combo.currentIndexChanged.connect(self._on_brand_changed)

    def _connect_signals(self):
        self.log_signal.connect(self._on_log_signal)
        self.progress_signal.connect(self._on_progress_signal)
        self.done_signal.connect(self._on_done_signal)
        self.error_signal.connect(self._on_error_signal)

    def _on_log_signal(self, msg):
        self.log_text.append(msg)
        QApplication.processEvents()

    def _on_progress_signal(self, value):
        self.progress.setValue(value)

    def _on_done_signal(self, data):
        找到数, 总数, output_dir = data
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress.setValue(100)
        reply = QMessageBox.information(
            self,
            "查找完成",
            f"查找完成！\n已找到: {找到数} 款\n未找到: {总数 - 找到数} 款\n\n是否打开输出目录？",
            QMessageBox.Ok | QMessageBox.Open,
        )
        if reply == QMessageBox.Open and os.path.exists(output_dir):
            os.startfile(output_dir)

    def _on_error_signal(self, msg):
        self.log_text.append(f"错误: {msg}")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress.setValue(0)

    def _log(self, msg):
        self.log_text.append(msg)
        QApplication.processEvents()

    def _get_cached_index(self, base_path):
        if base_path not in self._index_cache:
            return None
        cache = self._index_cache[base_path]
        if time.time() - cache.get("time", 0) > self._cache_expiry:
            del self._index_cache[base_path]
            return None
        return cache.get("index")

    def _save_index_cache(self, base_path, index):
        self._index_cache[base_path] = {"index": index, "time": time.time()}

    def _on_excel_dropped(self, path):
        self._log(f"已加载Excel: {path}")
        self._auto_detect_brand()

    def _on_brand_changed(self):
        brand = self.brand_combo.currentText()
        self._更新品牌路径显示(brand)
        self._刷新路径编辑框(brand)

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)
            self._auto_detect_brand()

    def _browse_output(self):
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if path:
            self.output_edit.setText(path)

    def _auto_detect_brand(self):
        excel_path = self.excel_edit.text().strip()
        if not excel_path or not os.path.exists(excel_path):
            self._log("请先选择Excel文件")
            return
        try:
            self._log("正在自动检测品牌...")
            原始数据, 列映射 = self._读取Excel(excel_path)
            if not 原始数据:
                self._log("Excel文件为空")
                return
            品牌 = self._检测品牌(原始数据, 列映射)
            self.brand_combo.blockSignals(True)
            self.brand_combo.setCurrentText(品牌)
            self.brand_combo.blockSignals(False)
            self._更新品牌路径显示(品牌)
            self._log(f"自动检测到品牌: {品牌}")
        except Exception as e:
            self._log(f"品牌检测失败: {e}")

    def _更新品牌路径显示(self, brand):
        if brand == "自动检测":
            self.brand_label.setText("加载Excel后自动显示路径")
            return
        cfg = self.品牌配置表.get(brand, {})
        paths = cfg.get("共享盘路径", [])
        if isinstance(paths, list) and paths:
            rows = []
            for p in paths:
                ok = "✓" if os.path.exists(p) else "✗"
                short = p[:40] + "..." if len(p) > 40 else p
                rows.append(f"[{ok}] {short}")
            self.brand_label.setText("\n".join(rows))
        else:
            self.brand_label.setText("未配置共享盘路径")

    def _start_search(self):
        excel_path = self.excel_edit.text().strip()
        if not excel_path:
            QMessageBox.warning(self, "提示", "请先选择Excel文件")
            return
        if not os.path.exists(excel_path):
            QMessageBox.warning(self, "提示", "Excel文件不存在")
            return

        output_dir = self.output_edit.text().strip()
        if not output_dir:
            当前日期 = time.strftime("%Y%m%d")
            品牌 = self.brand_combo.currentText()
            if 品牌 == "自动检测":
                品牌 = "唯品"
            elif 品牌 in self.品牌配置表:
                品牌 = self.品牌配置表[品牌].get("输出名称", 品牌)
            output_dir = os.path.join(os.path.dirname(excel_path), f"{当前日期}_{品牌}_原图")
        os.makedirs(output_dir, exist_ok=True)

        brand = self.brand_combo.currentText()
        路径列表 = self._获取品牌路径列表(brand, excel_path)
        if not 路径列表:
            QMessageBox.warning(self, "提示", f"未找到品牌【{brand}】的共享盘路径配置或路径不存在")
            return
        for i, p in enumerate(路径列表):
            self._log(f"搜索路径{i + 1}: {p[:80]}")

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.should_stop = False
        self.progress.setValue(0)
        self.log_text.clear()

        self.search_thread = threading.Thread(
            target=self._search_worker,
            args=(excel_path, output_dir, 路径列表, brand),
            daemon=True,
        )
        self.search_thread.start()

    def _获取品牌路径列表(self, brand, excel_path):
        路径列表 = []
        if brand == "自动检测":
            try:
                原始数据, 列映射 = self._读取Excel(excel_path)
                if 原始数据:
                    brand = self._检测品牌(原始数据, 列映射)
            except Exception:
                pass
        cfg = self.品牌配置表.get(brand, {})
        paths = cfg.get("共享盘路径", [])
        if isinstance(paths, list):
            路径列表 = [p for p in paths if os.path.exists(p)]
        elif paths and os.path.exists(paths):
            路径列表 = [paths]
        return 路径列表

    def _stop_search(self):
        self.should_stop = True
        self._log("正在停止...")

    def _search_worker(self, excel_path, output_dir, 共享盘路径列表, brand):
        try:
            self.log_signal.emit(f"Excel文件: {excel_path}")
            self.log_signal.emit(f"输出目录: {output_dir}")
            self.log_signal.emit("-" * 50)
            self.log_signal.emit("读取Excel...")
            原始数据, 列映射 = self._读取Excel(excel_path)
            if not 原始数据:
                self.error_signal.emit("Excel为空或无可用数据")
                return
            self.log_signal.emit(f"共 {len(原始数据)} 条记录")

            self.log_signal.emit("构建文件夹索引（带缓存）...")
            self.索引 = {"文件夹列表": [], "按款号索引": {}}
            for 共享盘路径 in 共享盘路径列表:
                if self.should_stop:
                    break
                索引 = self._get_cached_index(共享盘路径)
                if not 索引:
                    索引 = self._构建文件夹索引(共享盘路径)
                    if 索引:
                        self._save_index_cache(共享盘路径, 索引)
                if 索引:
                    self.索引["文件夹列表"].extend(索引.get("文件夹列表", []))
                    for 代码, 路径s in 索引.get("按款号索引", {}).items():
                        self.索引["按款号索引"].setdefault(代码, []).extend(路径s)

            if not self.索引.get("文件夹列表"):
                self.error_signal.emit("无法访问任何共享盘路径")
                return
            self.log_signal.emit(f"索引构建完成：共 {len(self.索引.get('文件夹列表', []))} 个产品文件夹")

            结果列表 = []
            找到计数 = 0
            total = len(原始数据)
            列_商家编码 = 列映射.get("商家编码")
            列_供应商款号 = 列映射.get("供应商款号")
            列_唯品款号 = 列映射.get("唯品款号")
            列_货号 = 列映射.get("货号")
            列_品牌 = 列映射.get("品牌")

            for 序号, (行, _) in enumerate(原始数据, 1):
                if self.should_stop:
                    self.log_signal.emit("已停止")
                    break
                唯品款号 = 行[列_唯品款号] if 列_唯品款号 is not None and len(行) > 列_唯品款号 else ""
                供应商款号 = 行[列_供应商款号] if 列_供应商款号 is not None and len(行) > 列_供应商款号 else ""
                商家编码 = 行[列_商家编码] if 列_商家编码 is not None and len(行) > 列_商家编码 else ""
                货号 = 行[列_货号] if 列_货号 is not None and len(行) > 列_货号 else ""
                品牌名 = 行[列_品牌] if 列_品牌 is not None and len(行) > 列_品牌 else ""
                唯品款号 = 唯品款号 or 供应商款号 or f"商品{序号}"

                if 品牌名:
                    品牌全称 = self.品牌名映射.get(str(品牌名).strip())
                    if 品牌全称:
                        self.log_signal.emit(f"[{序号}/{total}] 品牌: {品牌名} -> {品牌全称}")

                self.log_signal.emit(f"[{序号}/{total}] 供应商: {供应商款号}")
                文件夹路径, 分数, 匹配类型, 所有匹配 = self._快速搜索文件夹(供应商款号, 商家编码)
                结果 = {"唯品款号": 唯品款号, "供应商款号": 供应商款号, "商家编码": 商家编码, "货号": 货号, "找到": False}

                if 文件夹路径 and 分数 >= 0.8:
                    if len(所有匹配) > 1:
                        匹配文件夹列表 = [m[0] for m in 所有匹配[:5]]
                        图片数 = self._复制多个文件夹(匹配文件夹列表, output_dir, 唯品款号)
                        结果.update({
                            "找到": True,
                            "匹配文件夹名": "; ".join([os.path.basename(f) for f in 匹配文件夹列表]),
                            "原图路径": "; ".join(匹配文件夹列表),
                            "匹配类型": "多文件夹",
                            "匹配分数": 分数,
                            "图片数量": 图片数,
                            "本地路径": os.path.join(output_dir, str(唯品款号)),
                        })
                    else:
                        图片数 = self._复制文件夹(文件夹路径, output_dir, 唯品款号)
                        结果.update({
                            "找到": True,
                            "匹配文件夹名": os.path.basename(文件夹路径),
                            "原图路径": 文件夹路径,
                            "匹配类型": 匹配类型,
                            "匹配分数": 分数,
                            "图片数量": 图片数,
                            "本地路径": os.path.join(output_dir, str(唯品款号)),
                        })
                    找到计数 += 1
                else:
                    self.log_signal.emit("  [未找到]")
                结果列表.append(结果)
                self.progress_signal.emit(int(序号 * 100 / total))

            品牌输出名 = self.品牌配置表.get(brand, {}).get("输出名称", brand)
            结果文件 = os.path.join(output_dir, f"查找结果_{品牌输出名}.xlsx")
            self._写入结果Excel(结果列表, 结果文件)
            self.log_signal.emit(f"处理完成！总计: {len(结果列表)} 款，找到: {找到计数} 款")
            self.log_signal.emit(f"结果已保存到: {结果文件}")
            self.done_signal.emit((找到计数, len(结果列表), output_dir))
        except Exception as e:
            LOGGER.exception("找图异常")
            self.log_signal.emit(traceback.format_exc())
            self.error_signal.emit(str(e))
        finally:
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    def _检测品牌(self, 原始数据, 列映射):
        列_商家编码 = 列映射.get("商家编码")
        for 行, _ in 原始数据[:10]:
            if 列_商家编码 is not None and len(行) > 列_商家编码:
                商家编码 = str(行[列_商家编码]).strip().upper()
                for 品牌, 配置 in self.品牌配置表.items():
                    for 关键词 in 配置.get("品牌关键词", []):
                        if 关键词.upper() in 商家编码:
                            return 品牌
                if 商家编码.startswith("DM-"):
                    return "Dream made"
                if 商家编码.startswith("BT-"):
                    return "BEETLE TOWN"
        return "港仔文艺男"

    def _读取Excel(self, 文件路径):
        try:
            工作簿 = load_workbook(文件路径, data_only=False)
        except Exception:
            工作簿 = load_workbook(文件路径, data_only=True)
        工作表 = 工作簿.active
        表头 = [str(c.value).strip() if c.value else "" for c in 工作表[1]]
        列映射 = {"商家编码": None, "供应商款号": None, "唯品款号": None, "货号": None, "品牌": None}
        for i, 列名 in enumerate(表头):
            if "品牌" in 列名 and 列映射["品牌"] is None:
                列映射["品牌"] = i
            elif ("商家编码" in 列名 or "编码" in 列名) and 列映射["商家编码"] is None:
                列映射["商家编码"] = i
            elif ("供应商款号" in 列名 or "供应商" in 列名) and 列映射["供应商款号"] is None:
                列映射["供应商款号"] = i
            elif "唯品款号" in 列名 and 列映射["唯品款号"] is None:
                列映射["唯品款号"] = i
            elif ("货号" in 列名 or "图片地址" in 列名) and 列映射["货号"] is None:
                列映射["货号"] = i

        数据 = []
        for 行 in 工作表.iter_rows(min_row=2, values_only=True):
            if not 行:
                continue
            清理行 = []
            for v in 行:
                if v is None:
                    清理行.append("")
                    continue
                v_str = str(v)
                if v_str.startswith("=") and "&" in v_str:
                    parts = re.findall(r'"([^"]*)"', v_str)
                    清理行.append("".join(parts) if parts else re.sub(r'&[A-Z]+[0-9]+$', '', v_str.replace('="', "")))
                else:
                    清理行.append(v_str)
            if any(清理行[:3]):
                数据.append((清理行, 列映射))
        return 数据, 列映射

    def _提取产品代码(self, 商家编码):
        if not 商家编码:
            return ""
        parts = str(商家编码).strip().split("-")
        return parts[-1] if len(parts) >= 2 else str(商家编码)

    def _提取中间代码(self, 商家编码):
        if not 商家编码:
            return ""
        parts = str(商家编码).strip().split("-")
        return parts[1] if len(parts) >= 3 else ""

    def _带边界匹配(self, 搜索码, 文件夹名):
        if not 搜索码 or not 文件夹名:
            return False
        搜索大写 = str(搜索码).upper()
        文件夹大写 = str(文件夹名).upper()
        if 搜索大写 == 文件夹大写:
            return True
        return bool(re.search(r'(?<![A-Za-z0-9])' + re.escape(搜索大写) + r'(?![A-Za-z0-9])', 文件夹大写))

    def _计算匹配分数(self, 文件夹名, 供应商款号, 商家编码):
        文件夹大写 = str(文件夹名).upper()
        if 供应商款号 and self._带边界匹配(供应商款号, 文件夹名):
            s = str(供应商款号).upper()
            if s == 文件夹大写:
                return 1.0, "精确匹配"
            if 文件夹大写.startswith(s):
                return 0.98, "前缀匹配"
            return 0.95, "供应商款号匹配"
        if 商家编码 and len(str(商家编码)) >= 6 and self._带边界匹配(商家编码, 文件夹名):
            return 0.92, "商家编码匹配"
        产品代码 = self._提取产品代码(商家编码)
        中间代码 = self._提取中间代码(商家编码)
        if 中间代码 and self._带边界匹配(中间代码, 文件夹名):
            return 0.88, "中间代码匹配"
        if 产品代码 and len(产品代码) >= 4 and self._带边界匹配(产品代码, 文件夹名):
            return 0.8, "产品代码匹配"
        if 供应商款号 and len(str(供应商款号)) >= 4:
            ratio = SequenceMatcher(None, str(供应商款号).upper(), 文件夹大写).ratio()
            if ratio >= 0.8:
                return ratio * 0.6, "模糊匹配"
        return 0.0, "无匹配"

    def _构建文件夹索引(self, 基础路径):
        索引 = {"文件夹列表": [], "按款号索引": {}}
        for 根目录, 子目录列表, _ in os.walk(基础路径):
            if self.should_stop:
                break
            深度 = 根目录.replace(基础路径, "").count(os.sep)
            if 深度 > 6:
                子目录列表.clear()
                continue
            子目录列表[:] = [d for d in 子目录列表 if d not in 跳过目录]
            for 子目录名 in 子目录列表:
                完整路径 = os.path.join(根目录, 子目录名)
                if 完整路径.replace(基础路径, "").count(os.sep) < 3:
                    continue
                索引["文件夹列表"].append({"名称": 子目录名, "路径": 完整路径})
                codes = re.findall(r"[A-Z0-9]{3,}", 子目录名.upper())
                for code in codes:
                    if len(code) >= 4:
                        索引["按款号索引"].setdefault(code, []).append(完整路径)
        return 索引

    def _快速搜索文件夹(self, 供应商款号, 商家编码):
        if not self.索引:
            return None, 0, "无索引", []
        候选 = set()
        if 供应商款号:
            code = str(供应商款号).upper()
            候选.update(self.索引.get("按款号索引", {}).get(code, []))
        产品代码 = self._提取产品代码(商家编码).upper() if 商家编码 else ""
        if 产品代码:
            候选.update(self.索引.get("按款号索引", {}).get(产品代码, []))
        if not 候选:
            候选 = {f["路径"] for f in self.索引.get("文件夹列表", [])}

        最佳路径, 最高分, 最佳类型 = None, 0, "无匹配"
        所有匹配 = []
        for path in 候选:
            score, mtype = self._计算匹配分数(os.path.basename(path), 供应商款号, 商家编码)
            if score >= 0.85:
                所有匹配.append((path, score, mtype))
            if score > 最高分:
                最佳路径, 最高分, 最佳类型 = path, score, mtype
        所有匹配.sort(key=lambda x: x[1], reverse=True)
        return 最佳路径, 最高分, 最佳类型, 所有匹配

    def _清理文件名(self, text):
        bad = '<>:"/\\|?*&'
        text = str(text)
        for ch in bad:
            text = text.replace(ch, "_")
        return text

    def _复制多个文件夹(self, 源文件夹列表, 目标目录, 唯品款号):
        if not 源文件夹列表:
            return 0
        唯品目录 = os.path.join(目标目录, self._清理文件名(唯品款号))
        os.makedirs(唯品目录, exist_ok=True)
        图片扩展名 = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff"}
        count = 0
        for src_dir in 源文件夹列表:
            folder = os.path.basename(src_dir)
            dst_dir = os.path.join(唯品目录, folder)
            for 根, _, files in os.walk(src_dir):
                rel = os.path.relpath(根, src_dir)
                dst_root = dst_dir if rel == "." else os.path.join(dst_dir, rel)
                os.makedirs(dst_root, exist_ok=True)
                for f in files:
                    src = os.path.join(根, f)
                    dst = os.path.join(dst_root, f)
                    try:
                        shutil.copy2(src, dst)
                        if os.path.splitext(f)[1].lower() in 图片扩展名:
                            count += 1
                    except Exception:
                        LOGGER.warning("复制失败: %s", src)
        return count

    def _复制文件夹(self, 源文件夹, 目标目录, 唯品款号):
        return self._复制多个文件夹([源文件夹], 目标目录, 唯品款号)

    def _写入结果Excel(self, 结果列表, 输出路径):
        wb = Workbook()
        ws = wb.active
        ws.title = "查找结果"
        表头 = ["序号", "唯品款号", "供应商款号", "商家编码", "货号", "是否找到", "匹配文件夹名", "原图完整路径", "匹配类型", "匹配分数", "图片数量", "本地路径"]
        ws.append(表头)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, len(表头) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_alignment

        miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        low_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for i, result in enumerate(结果列表, 1):
            row = [
                i, result.get("唯品款号", ""), result.get("供应商款号", ""), result.get("商家编码", ""), result.get("货号", ""),
                "[找到]" if result.get("找到") else "[未找到]", result.get("匹配文件夹名", ""), result.get("原图路径", ""),
                result.get("匹配类型", ""), f"{result.get('匹配分数', 0):.2f}", result.get("图片数量", 0), result.get("本地路径", ""),
            ]
            ws.append(row)
            if not result.get("找到"):
                fill = miss_fill
            elif result.get("匹配分数", 0) < 0.8:
                fill = low_fill
            else:
                fill = ok_fill
            for col in range(1, len(row) + 1):
                ws.cell(row=i + 1, column=col).fill = fill

        widths = [6, 18, 15, 20, 18, 12, 40, 80, 12, 10, 10, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        wb.save(输出路径)

    def _加载品牌配置(self):
        """读取配置中的路径覆盖，和默认配置合并。"""
        brand_cfg = copy.deepcopy(默认品牌配置表)
        saved_paths = self.config.get("vip_finder_brand_paths", {})
        if isinstance(saved_paths, dict):
            for brand, paths in saved_paths.items():
                if brand in brand_cfg and isinstance(paths, list):
                    clean_paths = [str(p).strip() for p in paths if str(p).strip()]
                    if clean_paths:
                        brand_cfg[brand]["共享盘路径"] = clean_paths
        return brand_cfg

    def _构建品牌名映射(self):
        mapping = dict(默认品牌名映射)
        for brand, cfg in self.品牌配置表.items():
            mapping[brand] = brand
            for key in cfg.get("品牌关键词", []):
                if str(key).strip():
                    mapping[str(key).strip()] = brand
        return mapping

    def _刷新路径编辑框(self, brand):
        if brand == "自动检测":
            self.path_editor.setPlainText("请先选择具体品牌后再编辑路径")
            self.path_editor.setEnabled(False)
            return
        self.path_editor.setEnabled(True)
        paths = self.品牌配置表.get(brand, {}).get("共享盘路径", [])
        self.path_editor.setPlainText("\n".join(paths))

    def _添加路径到编辑框(self):
        brand = self.brand_combo.currentText()
        if brand == "自动检测":
            QMessageBox.warning(self, "提示", "请先选择具体品牌")
            return
        path = QFileDialog.getExistingDirectory(self, "选择共享盘目录")
        if not path:
            return
        old_text = self.path_editor.toPlainText().strip()
        lines = [x.strip() for x in old_text.splitlines() if x.strip()]
        if path not in lines:
            lines.append(path)
        self.path_editor.setPlainText("\n".join(lines))

    def _保存当前品牌路径(self):
        brand = self.brand_combo.currentText()
        if brand == "自动检测":
            QMessageBox.warning(self, "提示", "请先选择具体品牌")
            return
        lines = [x.strip() for x in self.path_editor.toPlainText().splitlines() if x.strip()]
        if not lines:
            QMessageBox.warning(self, "提示", "路径不能为空，至少保留一条")
            return

        self.品牌配置表.setdefault(brand, {})["共享盘路径"] = lines
        saved_paths = self.config.get("vip_finder_brand_paths", {})
        if not isinstance(saved_paths, dict):
            saved_paths = {}
        saved_paths[brand] = lines
        self.config["vip_finder_brand_paths"] = saved_paths
        save_config(self.config)
        self._更新品牌路径显示(brand)
        self._log(f"已保存【{brand}】路径，共 {len(lines)} 条")
        QMessageBox.information(self, "完成", f"已保存【{brand}】路径设置")
