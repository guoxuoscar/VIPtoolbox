# -*- coding: utf-8 -*-
import os
import re
import shutil
import threading

import openpyxl

from PySide6.QtCore import Signal, Qt
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from toolbox.ui.path_drop import ExcelDropLineEdit, DirDropLineEdit, enable_path_drop


# ── Tab 1: 批量新建文件夹 ──────────────────────────────────────────
class FolderCreatorTab(QWidget):
    log_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "根据Excel表格中的列批量创建文件夹。选一列 = 单层文件夹，选多列 = 多层嵌套文件夹（按选择顺序作为层级）。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        # Excel 选择
        excel_row = QHBoxLayout()
        excel_row.addWidget(QLabel("Excel表格:"))
        self.excel_input = ExcelDropLineEdit("拖拽或浏览选择Excel文件...")
        self.excel_input.setText(self.config.get("folder_creator_excel", ""))
        excel_row.addWidget(self.excel_input, 1)
        excel_btn = QPushButton("浏览")
        excel_btn.clicked.connect(self._browse_excel)
        excel_row.addWidget(excel_btn)
        layout.addLayout(excel_row)

        # 列选择区域
        col_group = QGroupBox("📋 选择列（多选=多层嵌套，顺序=层级顺序）")
        col_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        col_layout = QVBoxLayout()
        col_layout.setSpacing(6)

        col_top = QHBoxLayout()
        self.cb_vip_style = QCheckBox("唯品款号（默认）")
        self.cb_vip_style.setChecked(True)
        self.cb_vip_style.toggled.connect(self._on_col_check_changed)
        self.cb_vip_goods = QCheckBox("唯品货号")
        self.cb_vip_goods.setChecked(False)
        self.cb_vip_goods.toggled.connect(self._on_col_check_changed)
        self.cb_custom = QCheckBox("自定义列名:")
        self.cb_custom.toggled.connect(self._on_col_check_changed)
        self.custom_col = QLineEdit()
        self.custom_col.setPlaceholderText("输入列名（如: 款号、分类名称）")
        self.custom_col.setEnabled(False)
        col_top.addWidget(self.cb_vip_style)
        col_top.addWidget(self.cb_vip_goods)
        col_top.addWidget(self.cb_custom)
        col_top.addWidget(self.custom_col, 1)
        col_layout.addLayout(col_top)

        col_layout.addWidget(QLabel("💡 提示: 将按勾选顺序从上到下作为文件夹层级。例如: 唯品货号 → 唯品款号 = 货号文件夹/款号子文件夹"))
        col_group.setLayout(col_layout)
        layout.addWidget(col_group)

        # 输出目录
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("输出目录:"))
        self.output_dir = DirDropLineEdit("默认为表格所在目录，可自选")
        self.output_dir.setText(self.config.get("folder_creator_output", ""))
        out_row.addWidget(self.output_dir, 1)
        out_btn = QPushButton("浏览")
        out_btn.clicked.connect(self._browse_output)
        out_row.addWidget(out_btn)
        layout.addLayout(out_row)

        # 运行按钮
        self.run_btn = QPushButton("▶️ 开始批量建文件夹")
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #1976D2; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        # 日志
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)

        layout.addStretch()
        self.setLayout(layout)

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.excel_input.setText(path)

    def _browse_output(self):
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if path:
            self.output_dir.setText(path)

    def _on_col_check_changed(self):
        self.custom_col.setEnabled(self.cb_custom.isChecked())

    def _get_column_names(self):
        """按勾选顺序返回列名列表"""
        cols = []
        if self.cb_vip_style.isChecked():
            cols.append("唯品款号")
        if self.cb_vip_goods.isChecked():
            cols.append("唯品货号")
        if self.cb_custom.isChecked():
            txt = self.custom_col.text().strip()
            if txt:
                cols.append(txt)
        return cols

    def _start(self):
        excel_path = self.excel_input.text().strip()
        if not excel_path or not os.path.isfile(excel_path):
            QMessageBox.warning(self, "提示", "请先选择Excel文件")
            return

        cols = self._get_column_names()
        if not cols:
            QMessageBox.warning(self, "提示", "请至少选择一列")
            return

        output_dir = self.output_dir.text().strip()
        if not output_dir:
            output_dir = os.path.dirname(excel_path)
            self.output_dir.setText(output_dir)

        # 保存配置
        self.config["folder_creator_excel"] = excel_path
        self.config["folder_creator_output"] = output_dir

        self.run_btn.setEnabled(False)
        self.log_area.clear()
        threading.Thread(target=self._do_create, args=(excel_path, cols, output_dir), daemon=True).start()

    def _do_create(self, excel_path, cols, output_dir):
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            ws = wb.active
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header_map = {}
            for idx, h in enumerate(headers):
                if h:
                    header_map[str(h).strip()] = idx

            # 找到各列对应的列索引
            col_indices = []
            for col_name in cols:
                found = None
                for h, idx in header_map.items():
                    if col_name == h or col_name in h:
                        found = idx
                        break
                if found is None:
                    self.log_signal.emit(f"[错误] 未找到列: {col_name}")
                    return
                col_indices.append((col_name, found))

            self.log_signal.emit(f"找到列: {[c[0] for c in col_indices]}")
            self.log_signal.emit(f"输出目录: {output_dir}")
            self.log_signal.emit("")

            created = 0
            skipped = 0
            seen = set()

            for row_data in ws.iter_rows(min_row=2, values_only=True):
                parts = []
                all_valid = True
                for col_name, col_idx in col_indices:
                    val = row_data[col_idx] if col_idx < len(row_data) else None
                    if val is None or str(val).strip() == "":
                        all_valid = False
                        break
                    # 清理文件夹名非法字符
                    folder_name = str(val).strip()
                    folder_name = re.sub(r'[<>:"/\\|?*]', '_', folder_name)
                    parts.append(folder_name)

                if not all_valid or not parts:
                    skipped += 1
                    continue

                target = os.path.join(output_dir, *parts)
                if target in seen:
                    skipped += 1
                    continue
                seen.add(target)

                if not os.path.exists(target):
                    os.makedirs(target, exist_ok=True)
                    self.log_signal.emit(f"[创建] {os.path.relpath(target, output_dir)}")
                    created += 1
                else:
                    skipped += 1

            wb.close()
            self.log_signal.emit("")
            self.log_signal.emit(f"创建 {created} 个文件夹，跳过 {skipped} 个（已存在或无数据）")
            self.log_signal.emit("=" * 50)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.run_btn.setEnabled(True)


# ── Tab 2: 提取文件汇总（按文件夹名匹配） ──────────────────────
# 自定义模板常用预设
TEMPLATE_PRESETS = [
    ("款号_原文件名", "{parent}_{name}{ext}"),
    ("款号_序号", "{parent}_{num:04d}{ext}"),
    ("匹配文件夹_款号_原文件", "{matched}_{parent}_{name}{ext}"),
    ("父文件/原文件", "{parent}/{name}{ext}"),
    ("匹配/款号_原文件", "{matched}/{parent}_{name}{ext}"),
]


class FileCollectExtractTab(QWidget):
    """按文件夹名匹配目标文件夹，提取其中文件，支持多种输出结构和命名规则。"""
    log_signal = Signal(str)

    FOLDER_PRESETS = ["SKU", "详情页", "详情图片", "主图", "白底图", "透明图", "视频", "素材图"]
    OUTPUT_MODES = [
        ("款号文件夹归类 → 父文件夹名/原文件", "parent",
         "如: GZ001/front.jpg"),
        ("匹配文件夹归类 → 匹配文件夹名/原文件", "matched",
         "如: SKU/front.jpg"),
        ("匹配文件夹/款号 双层 → 匹配文件夹/父文件夹/文件", "matched_parent",
         "如: SKU/GZ001/front.jpg"),
        ("款号/匹配文件夹 双层 → 父文件夹/匹配文件夹/文件", "both",
         "如: GZ001/SKU/front.jpg"),
        ("扁平汇总 → 全部放同一目录", "flat",
         "如: GZ001_front.jpg 全部混放"),
    ]
    NAMING_MODES = [
        ("保持原文件名", "original",
         "{name}{ext}  → 如 front.jpg"),
        ("父文件夹_原文件名", "parent_name",
         "{parent}_{name}{ext}  → 如 GZ001_front.jpg"),
        ("父文件夹_序号", "parent_num",
         "{parent}_{num:04d}{ext}  → 如 GZ001_0001.jpg"),
        ("匹配文件夹_父文件夹_原文件名", "matched_parent_name",
         "{matched}_{parent}_{name}{ext}  → 如 SKU_GZ001_front.jpg"),
        ("自定义模板", "custom", ""),
    ]

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def _build_preview_text(self) -> str:
        """根据当前下拉选项，生成结构预览文本"""
        struct = self.struct_combo.currentData()
        naming = self.name_combo.currentData()
        template = self.custom_template.text().strip() if naming == "custom" else ""

        # 示例数据
        parents = ["GZ001", "GZ002"]
        matched = "SKU"
        files = ["front.jpg", "back.jpg"]

        def _make_name(p, m, f):
            name_no_ext = os.path.splitext(f)[0]
            ext = os.path.splitext(f)[1]
            if naming == "original":
                return f
            elif naming == "parent_name":
                return f"{p}_{f}"
            elif naming == "parent_num":
                return f"{p}_0001{ext}"
            elif naming == "matched_parent_name":
                return f"{m}_{p}_{f}"
            elif naming == "custom" and template:
                return template.replace("{parent}", p).replace("{matched}", m) \
                    .replace("{name}", name_no_ext).replace("{ext}", ext) \
                    .replace("{num}", "1").replace("{num:04d}", "0001")
            return f

        lines = ["📂 output/"]
        for i_p, p in enumerate(parents):
            # 确定子目录结构
            if struct == "parent":
                sub = p
            elif struct == "matched":
                sub = matched
            elif struct == "matched_parent":
                sub = f"{matched}/{p}"
            elif struct == "both":
                sub = f"{p}/{matched}"
            else:  # flat
                sub = ""

            is_last_parent = i_p == len(parents) - 1

            if sub:
                parts = sub.split("/")
                for d, part in enumerate(parts):
                    is_last_part = d == len(parts) - 1
                    prefix = "   " * d
                    branch = "   └─ " if (is_last_parent and is_last_part) else "   ├─ "
                    lines.append(f"{prefix}{branch}📂 {part}/")
                for j, f in enumerate(files):
                    base_prefix = "   " * len(parts)
                    branch = "   └─ " if (is_last_parent and j == len(files) - 1) else "   ├─ "
                    lines.append(f"{base_prefix}{branch}{_make_name(p, matched, f)}")
            else:
                for j, f in enumerate(files):
                    branch = "   └─ " if (is_last_parent and j == len(files) - 1) else "   ├─ "
                    lines.append(f"{branch}{_make_name(p, matched, f)}")

        lines.append("")
        lines.append(f"💡 示例: 勾选了 {len(parents)} 个款号, 匹配 {matched} 文件夹")
        lines.append("   实际结果取决于源文件夹结构")
        return "\n".join(lines)

    def _update_preview(self):
        """刷新右侧结构预览"""
        text = self._build_preview_text()
        self.preview_label.setText(text)

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "按文件夹名匹配目标文件夹（如SKU、详情图片），提取其中文件，支持多种输出结构和命名规则。\n"
            "例如: 勾选 SKU、详情页 → 自动找到所有名为SKU和详情页的文件夹 → 提取内部文件 → 按规则命名汇总。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        # 源文件夹
        src_row = QHBoxLayout()
        src_row.addWidget(QLabel("源文件夹:"))
        self.src_dir = DirDropLineEdit("拖拽或浏览选择根目录...")
        self.src_dir.setText(self.config.get("file_collect_src", ""))
        src_row.addWidget(self.src_dir, 1)
        src_btn = QPushButton("浏览")
        src_btn.clicked.connect(self._browse_src)
        src_row.addWidget(src_btn)
        layout.addLayout(src_row)

        # 匹配文件夹名
        folder_group = QGroupBox("📁 匹配文件夹名 (勾选需要提取的文件夹)")
        folder_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        folder_layout = QVBoxLayout()
        folder_layout.setSpacing(4)

        self.folder_cbs = {}
        row = QHBoxLayout()
        row.setSpacing(12)
        for i, name in enumerate(self.FOLDER_PRESETS):
            cb = QCheckBox(name)
            cb.setChecked(name in str(self.config.get("file_collect_folders", "SKU")).split(","))
            cb.toggled.connect(self._on_folder_toggled)
            self.folder_cbs[name] = cb
            row.addWidget(cb)
            if (i + 1) % 4 == 0 and i < len(self.FOLDER_PRESETS) - 1:
                folder_layout.addLayout(row)
                row = QHBoxLayout()
                row.setSpacing(12)
        folder_layout.addLayout(row)

        custom_row = QHBoxLayout()
        custom_row.addWidget(QLabel("自定义:"))
        self.custom_folder = QLineEdit()
        self.custom_folder.setPlaceholderText("空格分隔，如: 缩略图 角度图")
        self.custom_folder.setText(self.config.get("file_collect_custom_folders", ""))
        custom_row.addWidget(self.custom_folder, 1)
        folder_layout.addLayout(custom_row)

        folder_group.setLayout(folder_layout)
        layout.addWidget(folder_group)

        # 文件筛选
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("文件类型:"))
        self.ext_input = QLineEdit()
        self.ext_input.setPlaceholderText("空格分隔，如: .jpg .png（留空=所有文件）")
        self.ext_input.setText(self.config.get("file_collect_ext", ""))
        filter_row.addWidget(self.ext_input, 1)
        filter_row.addWidget(QLabel("  文件名含:"))
        self.kw_input = QLineEdit()
        self.kw_input.setPlaceholderText("如: front（留空=不筛选）")
        self.kw_input.setText(self.config.get("file_collect_kw", ""))
        filter_row.addWidget(self.kw_input, 1)
        layout.addLayout(filter_row)

        # ===== 输出结构 & 命名规则（左右布局） =====
        out_group = QGroupBox("📂 输出结构 & ✏️ 命名规则")
        out_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        out_horiz = QHBoxLayout()
        out_horiz.setSpacing(16)

        # ── 左侧: 控件 ──
        left_col = QVBoxLayout()
        left_col.setSpacing(8)

        # 输出结构
        struct_row = QHBoxLayout()
        struct_row.addWidget(QLabel("输出结构:"))
        self.struct_combo = QComboBox()
        for label, val, hint in self.OUTPUT_MODES:
            self.struct_combo.addItem(f"{label}", val)
        prev_struct = self.config.get("file_collect_struct", "parent")
        for k in range(self.struct_combo.count()):
            if self.struct_combo.itemData(k) == prev_struct:
                self.struct_combo.setCurrentIndex(k)
                break
        self.struct_combo.currentIndexChanged.connect(self._update_preview)
        struct_row.addWidget(self.struct_combo, 1)

        # 结构模式提示
        self.struct_hint = QLabel("")
        self.struct_hint.setStyleSheet("color: #999; font-size: 9px;")
        struct_row.addWidget(self.struct_hint, 1)

        def _show_struct_hint(idx):
            _, _, hint = self.OUTPUT_MODES[idx]
            self.struct_hint.setText(hint)
        self.struct_combo.currentIndexChanged.connect(_show_struct_hint)
        left_col.addLayout(struct_row)

        # 命名规则
        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("命名规则:"))
        self.name_combo = QComboBox()
        for label, val, hint in self.NAMING_MODES:
            self.name_combo.addItem(label, val)
        prev_name = self.config.get("file_collect_naming", "original")
        for k in range(self.name_combo.count()):
            if self.name_combo.itemData(k) == prev_name:
                self.name_combo.setCurrentIndex(k)
                break
        self.name_combo.currentIndexChanged.connect(self._on_naming_changed)
        self.name_combo.currentIndexChanged.connect(self._update_preview)
        name_row.addWidget(self.name_combo, 1)

        # 命名规则提示
        self.naming_hint = QLabel("")
        self.naming_hint.setStyleSheet("color: #999; font-size: 9px;")
        name_row.addWidget(self.naming_hint, 1)

        def _show_naming_hint(idx):
            _, _, hint = self.NAMING_MODES[idx]
            self.naming_hint.setText(hint)
        self.name_combo.currentIndexChanged.connect(_show_naming_hint)
        left_col.addLayout(name_row)

        # 自定义模板 + 预设按钮
        templ_row = QHBoxLayout()
        templ_row.addWidget(QLabel("自定义模板:"))
        self.custom_template = QLineEdit()
        self.custom_template.setPlaceholderText("变量: {parent} {matched} {name} {ext} {num}")
        self.custom_template.setText(self.config.get("file_collect_template", "{parent}_{matched}_{name}{ext}"))
        self.custom_template.textChanged.connect(self._update_preview)
        templ_row.addWidget(self.custom_template, 1)
        left_col.addLayout(templ_row)

        # 模板预设按钮行
        preset_row = QHBoxLayout()
        preset_row.setSpacing(4)
        preset_row.addWidget(QLabel("常用模板:"))
        for label, tpl in TEMPLATE_PRESETS:
            btn = QPushButton(label)
            btn.setFixedHeight(24)
            btn.setStyleSheet("""
                QPushButton {
                    background: #E8F5E9; color: #2E7D32; border: 1px solid #A5D6A7;
                    border-radius: 4px; padding: 2px 8px; font-size: 9px; font-weight: normal;
                }
                QPushButton:hover { background: #C8E6C9; }
            """)
            btn.clicked.connect(lambda checked, t=tpl: self._apply_template(t))
            preset_row.addWidget(btn)
        preset_row.addStretch()
        left_col.addLayout(preset_row)

        # 变量说明
        hint = QLabel(
            "💡 {parent}=父文件夹名(款号)  {matched}=匹配文件夹名(SKU)  "
            "{name}=原文件名(无后缀)  {ext}=扩展名  {num}=序号"
        )
        hint.setStyleSheet("color: #888; font-size: 9px;")
        hint.setWordWrap(True)
        left_col.addWidget(hint)

        out_horiz.addLayout(left_col, 3)

        # ── 右侧: 结构预览 ──
        preview_frame = QGroupBox("📐 结构预览")
        preview_frame.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        preview_frame.setStyleSheet("""
            QGroupBox {
                border: 1px solid #C8E6C9; border-radius: 8px; margin-top: 10px;
                background: #F1F8E9; font-family: Microsoft YaHei; font-size: 10px;
                font-weight: bold; color: #2E7D32; padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 10px; padding: 0 6px;
            }
        """)
        preview_vl = QVBoxLayout(preview_frame)
        preview_vl.setContentsMargins(8, 8, 8, 8)
        self.preview_label = QLabel()
        self.preview_label.setFont(QFont("Consolas", 9))
        self.preview_label.setStyleSheet("color: #333; background: transparent;")
        self.preview_label.setWordWrap(False)
        preview_vl.addWidget(self.preview_label)
        out_horiz.addWidget(preview_frame, 2)

        out_group.setLayout(out_horiz)
        layout.addWidget(out_group)

        # ===== 输出目录 & 运行 =====
        outdir_row = QHBoxLayout()
        outdir_row.addWidget(QLabel("输出目录:"))
        self.output_dir = DirDropLineEdit("汇总输出位置...")
        self.output_dir.setText(self.config.get("file_collect_output", ""))
        outdir_row.addWidget(self.output_dir, 1)
        out_btn = QPushButton("浏览")
        out_btn.clicked.connect(self._browse_output)
        outdir_row.addWidget(out_btn)
        layout.addLayout(outdir_row)

        self.run_btn = QPushButton("▶️ 开始提取汇总")
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #00796B; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #00897B; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

        # 初始状态
        self._on_naming_changed()
        _show_struct_hint(self.struct_combo.currentIndex())
        _show_naming_hint(self.name_combo.currentIndex())
        self._update_preview()

    def _apply_template(self, tpl: str):
        """点击模板预设按钮时填入自定义模板"""
        self.custom_template.setText(tpl)
        # 自动切换到"自定义模板"命名模式
        for k in range(self.name_combo.count()):
            if self.name_combo.itemData(k) == "custom":
                self.name_combo.setCurrentIndex(k)
                break

    def _browse_src(self):
        path = QFileDialog.getExistingDirectory(self, "选择源文件夹")
        if path:
            self.src_dir.setText(path)

    def _browse_output(self):
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if path:
            self.output_dir.setText(path)

    def _on_folder_toggled(self):
        pass

    def _on_naming_changed(self):
        is_custom = self.name_combo.currentData() == "custom"
        self.custom_template.setEnabled(is_custom)

    def _get_folder_names(self):
        names = []
        for name, cb in self.folder_cbs.items():
            if cb.isChecked():
                names.append(name)
        custom = self.custom_folder.text().strip()
        if custom:
            for n in custom.split():
                n = n.strip()
                if n and n not in names:
                    names.append(n)
        return names

    def _start(self):
        src = self.src_dir.text().strip()
        if not src or not os.path.isdir(src):
            QMessageBox.warning(self, "提示", "请先选择源文件夹")
            return

        folder_names = self._get_folder_names()
        if not folder_names:
            QMessageBox.warning(self, "提示", "请至少选择一个目标文件夹名")
            return

        output = self.output_dir.text().strip()
        if not output:
            QMessageBox.warning(self, "提示", "请选择输出目录")
            return

        exts = [e.strip().lower() for e in self.ext_input.text().strip().split() if e.strip()]
        kw = self.kw_input.text().strip()
        struct = self.struct_combo.currentData()
        naming = self.name_combo.currentData()
        template = self.custom_template.text().strip() if naming == "custom" else ""

        # 保存配置
        self.config["file_collect_src"] = src
        self.config["file_collect_folders"] = ",".join(folder_names)
        self.config["file_collect_custom_folders"] = self.custom_folder.text().strip()
        self.config["file_collect_ext"] = self.ext_input.text().strip()
        self.config["file_collect_kw"] = kw
        self.config["file_collect_struct"] = struct
        self.config["file_collect_naming"] = naming
        self.config["file_collect_template"] = template
        self.config["file_collect_output"] = output

        self.run_btn.setEnabled(False)
        self.log_area.clear()
        threading.Thread(
            target=self._do_collect,
            args=(src, folder_names, exts, kw, struct, naming, template, output),
            daemon=True,
        ).start()

    def _do_collect(self, src, folder_names, exts, kw, struct, naming, template, output):
        try:
            src = os.path.abspath(src)
            output = os.path.abspath(output)
            folder_set = set(folder_names)
            found = 0
            copied = 0
            counters = {}  # per-dest-dir counters for {num}

            for root, dirs, files in os.walk(src):
                current_name = os.path.basename(root)
                if current_name not in folder_set:
                    continue

                # 找到了匹配的文件夹
                rel = os.path.relpath(root, src)
                parts = rel.split(os.sep) if rel != "." else []
                parent = parts[-2] if len(parts) >= 2 else os.path.basename(src)
                matched = current_name

                for f in files:
                    # 文件筛选
                    ext = os.path.splitext(f)[1].lower()
                    if exts and ext not in exts:
                        continue
                    if kw and kw.lower() not in f.lower():
                        continue

                    found += 1
                    src_path = os.path.join(root, f)
                    name_no_ext = os.path.splitext(f)[0]

                    # 确定目标子目录
                    if struct == "parent":
                        dest_sub = parent
                    elif struct == "matched":
                        dest_sub = matched
                    elif struct == "matched_parent":
                        dest_sub = os.path.join(matched, parent)
                    elif struct == "both":
                        dest_sub = os.path.join(parent, matched)
                    else:  # flat
                        dest_sub = ""

                    # 确定新文件名
                    if naming == "original":
                        new_fname = f
                    elif naming == "parent_name":
                        new_fname = f"{parent}_{f}"
                    elif naming == "parent_num":
                        dest_key = dest_sub or "_flat"
                        idx = counters.get(dest_key, 0) + 1
                        counters[dest_key] = idx
                        new_fname = f"{parent}_{idx:04d}{ext}"
                    elif naming == "matched_parent_name":
                        new_fname = f"{matched}_{parent}_{f}"
                    elif naming == "custom":
                        dest_key = dest_sub or "_flat"
                        idx = counters.get(dest_key, 0) + 1
                        counters[dest_key] = idx
                        new_fname = template.replace("{parent}", parent).replace("{matched}", matched).replace("{name}", name_no_ext).replace("{ext}", ext).replace("{num}", str(idx)).replace("{num:04d}", f"{idx:04d}")
                    else:
                        new_fname = f

                    # 构建目标路径
                    if dest_sub:
                        dest_dir = os.path.join(output, dest_sub)
                    else:
                        dest_dir = output
                    os.makedirs(dest_dir, exist_ok=True)
                    dest_path = os.path.join(dest_dir, new_fname)

                    # 防重名
                    base, ext_f = os.path.splitext(new_fname)
                    counter = 2
                    while os.path.exists(dest_path):
                        dest_path = os.path.join(dest_dir, f"{base}_{counter}{ext_f}")
                        counter += 1

                    shutil.copy2(src_path, dest_path)
                    copied += 1
                    self.log_signal.emit(f"[提取] {os.path.relpath(src_path, src)} → {os.path.relpath(dest_path, output)}")

            self.log_signal.emit("")
            self.log_signal.emit(f"扫描到 {found} 个匹配文件，成功复制 {copied} 个")
            self.log_signal.emit(f"匹配文件夹: {', '.join(folder_names)}")
            self.log_signal.emit("=" * 50)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
            import traceback
            self.log_signal.emit(traceback.format_exc())
        finally:
            self.run_btn.setEnabled(True)


# ── Tab 3: 图片归入同名文件夹 ────────────────────────────────────
class ImageToFolderTab(QWidget):
    log_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "将目标目录下的图片文件，自动移入与之同名的文件夹中。\n"
            "例如: A001.jpg 移入 A001 文件夹；支持多种图片格式。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel("目标目录:"))
        self.target_dir = DirDropLineEdit("拖拽或浏览选择包含图片和文件夹的目录...")
        self.target_dir.setText(self.config.get("img2folder_dir", ""))
        dir_row.addWidget(self.target_dir, 1)
        dir_btn = QPushButton("浏览")
        dir_btn.clicked.connect(self._browse_dir)
        dir_row.addWidget(dir_btn)
        layout.addLayout(dir_row)

        ext_row = QHBoxLayout()
        ext_row.addWidget(QLabel("图片格式:"))
        self.ext_input = QLineEdit()
        self.ext_input.setText(self.config.get("img2folder_ext", ".jpg .jpeg .png .bmp .webp .gif"))
        self.ext_input.setPlaceholderText("空格分隔，如: .jpg .png .bmp")
        ext_row.addWidget(self.ext_input, 1)
        layout.addLayout(ext_row)

        self.run_btn = QPushButton("▶️ 开始归入")
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #E65100; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #F57C00; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

    def _browse_dir(self):
        path = QFileDialog.getExistingDirectory(self, "选择目录")
        if path:
            self.target_dir.setText(path)

    def _start(self):
        target = self.target_dir.text().strip()
        if not target or not os.path.isdir(target):
            QMessageBox.warning(self, "提示", "请选择目标目录")
            return
        exts = [e.strip().lower() for e in self.ext_input.text().strip().split() if e.strip()]
        if not exts:
            QMessageBox.warning(self, "提示", "请至少输入一个图片格式")
            return

        self.config["img2folder_dir"] = target
        self.config["img2folder_ext"] = " ".join(exts)

        self.run_btn.setEnabled(False)
        self.log_area.clear()
        threading.Thread(target=self._do_move, args=(target, exts), daemon=True).start()

    def _do_move(self, target, exts):
        try:
            moved = 0
            skipped = 0
            for f in os.listdir(target):
                fp = os.path.join(target, f)
                if not os.path.isfile(fp):
                    continue
                ext = os.path.splitext(f)[1].lower()
                if ext not in exts:
                    continue
                name = os.path.splitext(f)[0]
                dest_folder = os.path.join(target, name)
                os.makedirs(dest_folder, exist_ok=True)
                dest = os.path.join(dest_folder, f)
                if os.path.exists(dest):
                    self.log_signal.emit(f"[跳过] {f} (目标已存在)")
                    skipped += 1
                    continue
                shutil.move(fp, dest)
                self.log_signal.emit(f"[移入] {f} → {name}/")
                moved += 1
            self.log_signal.emit("")
            self.log_signal.emit(f"移入 {moved} 个文件，跳过 {skipped} 个")
            self.log_signal.emit("=" * 50)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.run_btn.setEnabled(True)


# ── Tab 4: 导出文件清单 ──────────────────────────────────────────
class FileListExportTab(QWidget):
    log_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel("将目录下的所有文件/文件夹名称导出为文本文件，支持递归扫描。")
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel("扫描目录:"))
        self.target_dir = DirDropLineEdit("拖拽或浏览选择目录...")
        self.target_dir.setText(self.config.get("filelist_dir", ""))
        dir_row.addWidget(self.target_dir, 1)
        dir_btn = QPushButton("浏览")
        dir_btn.clicked.connect(self._browse_dir)
        dir_row.addWidget(dir_btn)
        layout.addLayout(dir_row)

        opt_row = QHBoxLayout()
        self.recursive_cb = QCheckBox("递归子文件夹")
        self.recursive_cb.setChecked(True)
        opt_row.addWidget(self.recursive_cb)
        self.only_files_cb = QCheckBox("仅文件")
        opt_row.addWidget(self.only_files_cb)
        self.only_dirs_cb = QCheckBox("仅文件夹")
        opt_row.addWidget(self.only_dirs_cb)
        opt_row.addStretch()
        layout.addLayout(opt_row)

        self.export_btn = QPushButton("📋 导出文件清单")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #5E35B1; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #7E57C2; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.export_btn.clicked.connect(self._start)
        layout.addWidget(self.export_btn)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

    def _browse_dir(self):
        path = QFileDialog.getExistingDirectory(self, "选择目录")
        if path:
            self.target_dir.setText(path)

    def _start(self):
        target = self.target_dir.text().strip()
        if not target or not os.path.isdir(target):
            QMessageBox.warning(self, "提示", "请选择目录")
            return
        self.config["filelist_dir"] = target
        self.export_btn.setEnabled(False)
        self.log_area.clear()
        threading.Thread(target=self._do_export, args=(target,), daemon=True).start()

    def _do_export(self, target):
        try:
            recursive = self.recursive_cb.isChecked()
            only_files = self.only_files_cb.isChecked()
            only_dirs = self.only_dirs_cb.isChecked()
            out_file = os.path.join(target, "文件清单.txt")
            lines = []
            lines.append(f"目录: {target}")
            lines.append(f"导出时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            lines.append("=" * 60)
            lines.append("")

            if recursive:
                for root, dirs, files in os.walk(target):
                    rel = os.path.relpath(root, target)
                    if not only_files:
                        for d in dirs:
                            lines.append(f"[D] {os.path.join(rel, d) if rel != '.' else d}")
                    if not only_dirs:
                        for f in files:
                            lines.append(f"[F] {os.path.join(rel, f) if rel != '.' else f}")
            else:
                for item in sorted(os.listdir(target)):
                    full = os.path.join(target, item)
                    is_dir = os.path.isdir(full)
                    if only_files and is_dir:
                        continue
                    if only_dirs and not is_dir:
                        continue
                    prefix = "[D]" if is_dir else "[F]"
                    lines.append(f"{prefix} {item}")

            with open(out_file, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            self.log_signal.emit(f"清单已导出: {out_file}")
            self.log_signal.emit(f"共 {len(lines) - 4} 条记录")
            self.log_signal.emit("=" * 50)
            try:
                os.startfile(out_file)
            except Exception:
                pass
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.export_btn.setEnabled(True)


# ── Tab 5: 批量分发文件 ──────────────────────────────────────────
class FileDistributeTab(QWidget):
    log_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "将指定的文件批量复制到目标目录下的每一个子文件夹中。\n"
            "常用于: 将封面图、默认图等素材分发到各商品文件夹。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        # 源文件
        src_row = QHBoxLayout()
        src_row.addWidget(QLabel("源文件:"))
        self.src_files = QLineEdit()
        self.src_files.setPlaceholderText("支持拖拽多个文件到此处")
        self.src_files.setText(self.config.get("distribute_files", ""))
        enable_path_drop(self.src_files, mode="file", multi=True)
        src_row.addWidget(self.src_files, 1)
        src_btn = QPushButton("浏览")
        src_btn.clicked.connect(self._browse_files)
        src_row.addWidget(src_btn)
        layout.addLayout(src_row)

        # 目标父目录
        dest_row = QHBoxLayout()
        dest_row.addWidget(QLabel("目标父目录:"))
        self.dest_dir = DirDropLineEdit("拖拽或浏览选择含子文件夹的父目录...")
        self.dest_dir.setText(self.config.get("distribute_dest", ""))
        dest_row.addWidget(self.dest_dir, 1)
        dest_btn = QPushButton("浏览")
        dest_btn.clicked.connect(self._browse_dest)
        dest_row.addWidget(dest_btn)
        layout.addLayout(dest_row)

        self.run_btn = QPushButton("▶️ 开始批量分发")
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #388E3C; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

    def _browse_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择要分发的文件", "", "所有文件 (*.*)")
        if paths:
            self.src_files.setText(";".join(paths))

    def _browse_dest(self):
        path = QFileDialog.getExistingDirectory(self, "选择目标父目录")
        if path:
            self.dest_dir.setText(path)

    def _start(self):
        src_text = self.src_files.text().strip()
        if not src_text:
            QMessageBox.warning(self, "提示", "请选择要分发的源文件")
            return
        dest = self.dest_dir.text().strip()
        if not dest or not os.path.isdir(dest):
            QMessageBox.warning(self, "提示", "请选择目标父目录")
            return

        self.config["distribute_files"] = src_text
        self.config["distribute_dest"] = dest

        self.run_btn.setEnabled(False)
        self.log_area.clear()

        src_paths = [p.strip() for p in src_text.split(";") if p.strip()]
        threading.Thread(target=self._do_distribute, args=(src_paths, dest), daemon=True).start()

    def _do_distribute(self, src_paths, dest):
        try:
            subdirs = [os.path.join(dest, d) for d in os.listdir(dest)
                       if os.path.isdir(os.path.join(dest, d))]
            if not subdirs:
                self.log_signal.emit("[错误] 目标目录下没有子文件夹")
                return

            total = 0
            for sub in subdirs:
                for src in src_paths:
                    if not os.path.isfile(src):
                        continue
                    fname = os.path.basename(src)
                    dest_path = os.path.join(sub, fname)
                    shutil.copy2(src, dest_path)
                    total += 1
                    self.log_signal.emit(f"[复制] {fname} → {os.path.basename(sub)}/")

            self.log_signal.emit("")
            self.log_signal.emit(f"分发完成: {len(src_paths)} 个文件 × {len(subdirs)} 个文件夹 = {total} 次复制")
            self.log_signal.emit("=" * 50)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.run_btn.setEnabled(True)


# ── (已合并到 Tab 2: FileCollectExtractTab) ─────────────────
# ── Tab 7: 供应商款号提取 ────────────────────────────────────────
class SupplierCodeTab(QWidget):
    log_signal = Signal(str)

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self._pywin32_ok = False
        self._check_pywin32()
        self.init_ui()

    def _check_pywin32(self):
        try:
            import pythoncom
            import win32com.client
            self._pywin32_ok = True
        except ImportError:
            self._pywin32_ok = False

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "从Excel的商家编码/商品编码列自动提取供应商款号，在原文件旁生成带前缀的新文件。\n"
            "编码规则: 去尾部后缀(DD/DK/DZ/单数字/EKxx等)取最后剩余段为款号。\n"
            "使用 win32com 操作以保留 DISPIMG 嵌入式图片。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        # 输入模式
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("输入方式:"))
        self.mode_combo = QComboBox()
        self.mode_combo.addItem("选择单个Excel文件", "file")
        self.mode_combo.addItem("选择文件夹(批量处理)", "folder")
        self.mode_combo.currentIndexChanged.connect(self._on_mode_changed)
        mode_row.addWidget(self.mode_combo)
        mode_row.addStretch()
        layout.addLayout(mode_row)

        # 输入路径
        src_row = QHBoxLayout()
        src_row.addWidget(QLabel("输入路径:"))
        self.src_input = QLineEdit()
        self.src_input.setPlaceholderText("拖拽Excel或浏览选择...")
        self.src_input.setText(self.config.get("supplier_code_input", ""))
        enable_path_drop(self.src_input, mode="file_or_dir")
        src_row.addWidget(self.src_input, 1)
        self.src_btn = QPushButton("浏览")
        self.src_btn.clicked.connect(self._browse_src)
        src_row.addWidget(self.src_btn)
        layout.addLayout(src_row)

        # 目标列选择
        col_row = QHBoxLayout()
        col_row.addWidget(QLabel("目标列名:"))
        self.col_input = QLineEdit()
        self.col_input.setPlaceholderText("自动识别，也可手动输入(如: 商家编码)")
        self.col_input.setText(self.config.get("supplier_code_col", ""))
        col_row.addWidget(self.col_input, 1)
        layout.addLayout(col_row)
        hint = QLabel("留空则自动匹配: 商家编码 > 商品编码 > 货号")
        hint.setStyleSheet("color: #999; font-size: 9px;")
        layout.addWidget(hint)

        # 输出前缀
        prefix_row = QHBoxLayout()
        prefix_row.addWidget(QLabel("输出前缀:"))
        self.prefix_input = QLineEdit()
        self.prefix_input.setText(self.config.get("supplier_code_prefix", "供应商款号-"))
        prefix_row.addWidget(self.prefix_input)
        prefix_row.addWidget(QLabel("  新文件名 = 前缀 + 原文件名"))
        prefix_row.addStretch()
        layout.addLayout(prefix_row)

        self.run_btn = QPushButton("▶️ 开始提取供应商款号")
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #6A1B9A; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #8E24AA; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

    def _on_mode_changed(self):
        mode = self.mode_combo.currentData()
        if mode == "file":
            self.src_input.setPlaceholderText("拖拽Excel文件或浏览选择...")
        else:
            self.src_input.setPlaceholderText("拖拽文件夹或浏览选择...")

    def _browse_src(self):
        mode = self.mode_combo.currentData()
        if mode == "file":
            path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx *.xls)")
            if path:
                self.src_input.setText(path)
        else:
            path = QFileDialog.getExistingDirectory(self, "选择文件夹")
            if path:
                self.src_input.setText(path)

    def _start(self):
        if not self._pywin32_ok:
            ret = QMessageBox.question(
                self, "缺少依赖",
                "供应商款号提取需要 pywin32 库，是否现在安装？\n\n"
                "安装后需要重启程序。",
                QMessageBox.Yes | QMessageBox.No
            )
            if ret == QMessageBox.Yes:
                import subprocess, sys
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "pywin32", "-q"])
                    QMessageBox.information(self, "完成", "pywin32 安装完成，请重启程序。")
                    self._pywin32_ok = True
                except Exception as e:
                    QMessageBox.critical(self, "失败", f"安装失败: {e}")
            return

        src = self.src_input.text().strip()
        if not src or not os.path.exists(src):
            QMessageBox.warning(self, "提示", "请选择有效的输入路径")
            return

        self.config["supplier_code_input"] = src
        self.config["supplier_code_col"] = self.col_input.text().strip()
        self.config["supplier_code_prefix"] = self.prefix_input.text().strip() or "供应商款号-"

        self.run_btn.setEnabled(False)
        self.log_area.clear()
        threading.Thread(target=self._do_extract, daemon=True).start()

    def _find_files(self, src):
        mode = self.mode_combo.currentData()
        if mode == "file":
            if src.endswith((".xlsx", ".xls")) and not os.path.basename(src).startswith("~$"):
                return [src]
            return []
        else:
            import glob
            files = glob.glob(os.path.join(src, "*.xlsx"))
            files = [f for f in files
                     if not os.path.basename(f).startswith("供应商款号-")
                     and not os.path.basename(f).startswith("~$")]
            return sorted(files)

    def _find_column(self, ws, max_col, target_col_name):
        for col in range(1, max_col + 1):
            val = ws.Cells(1, col).Value
            if val and target_col_name in str(val):
                return col
        return None

    def _extract_code(self, code):
        if not code:
            return ""
        parts = str(code).split("-")
        if len(parts) <= 1:
            return parts[0]

        def is_suffix(seg):
            if re.match(r'^[0-9]$', seg):
                return True
            if seg in ("01", "02"):
                return True
            if seg in ("DD", "DK", "DZ"):
                return True
            if re.match(r'^EK.{2}$', seg):
                return True
            return False

        suffix_count = 0
        for i in range(len(parts) - 1, -1, -1):
            if is_suffix(parts[i]):
                suffix_count += 1
            else:
                break

        if suffix_count > 0:
            start = len(parts) - suffix_count - 1
            return "-".join(parts[start:])

        if len(parts[-1]) < 4 and len(parts) >= 2:
            return "-".join(parts[-2:])
        return parts[-1]

    def _do_extract(self):
        import pythoncom
        pythoncom.CoInitialize()
        excel = None
        try:
            excel = __import__("win32com.client", fromlist=["win32com"]).Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            src = self.src_input.text().strip()
            files = self._find_files(src)
            if not files:
                self.log_signal.emit("[错误] 未找到可处理的Excel文件")
                return

            col_name = self.col_input.text().strip() or ""
            prefix = self.prefix_input.text().strip() or "供应商款号-"
            XL_FORMAT = 51

            self.log_signal.emit(f"找到 {len(files)} 个文件待处理")
            self.log_signal.emit(f"输出前缀: {prefix}")
            self.log_signal.emit("")

            success = 0
            for i, fpath in enumerate(files, 1):
                fname = os.path.basename(fpath)
                self.log_signal.emit(f"--- 处理 ({i}/{len(files)}): {fname} ---")
                try:
                    wb = excel.Workbooks.Open(fpath)
                    ws = wb.ActiveSheet
                    used = ws.UsedRange
                    max_row = used.Rows.Count
                    max_col = used.Columns.Count

                    # 找列
                    col_idx = None
                    if col_name:
                        col_idx = self._find_column(ws, max_col, col_name)
                    if col_idx is None:
                        for cn in ("商家编码", "商品编码", "货号"):
                            col_idx = self._find_column(ws, max_col, cn)
                            if col_idx is not None:
                                break

                    if col_idx is None:
                        self.log_signal.emit(f"  [跳过] 未找到目标列，表头: {[ws.Cells(1,c).Value for c in range(1,min(6,max_col+1))]}")
                        wb.Close(SaveChanges=False)
                        continue

                    header = ws.Cells(1, col_idx).Value
                    self.log_signal.emit(f"  定位列: \"{header}\" (第{col_idx}列)")

                    new_col = col_idx + 1
                    try:
                        ws.Columns(new_col).Insert()
                    except Exception:
                        pass
                    ws.Cells(1, new_col).Value = "供应商款号"

                    changed = 0
                    for row in range(2, max_row + 1):
                        val = ws.Cells(row, col_idx).Value
                        if val is None:
                            continue
                        new_val = self._extract_code(str(val))
                        ws.Cells(row, new_col).Value = new_val
                        changed += 1
                        if row <= 5:
                            self.log_signal.emit(f"  Row {row}: {str(val)[:40]} → {new_val}")

                    folder = os.path.dirname(fpath) or "."
                    name = os.path.basename(fpath)
                    out_path = os.path.join(folder, f"{prefix}{name}")
                    if os.path.exists(out_path):
                        os.remove(out_path)
                    wb.SaveAs(out_path, XL_FORMAT)
                    self.log_signal.emit(f"  [完成] 处理 {changed} 行 → {prefix}{name}")
                    success += 1
                    wb.Close(SaveChanges=False)
                except Exception as e:
                    self.log_signal.emit(f"  [错误] {e}")
                    try:
                        wb.Close(SaveChanges=False)
                    except Exception:
                        pass

            self.log_signal.emit("")
            self.log_signal.emit(f"全部完成: 成功 {success}/{len(files)}")
            self.log_signal.emit("=" * 50)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
            import traceback
            self.log_signal.emit(traceback.format_exc())
        finally:
            if excel:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
            self.run_btn.setEnabled(True)


# ── Tab 8: 清理重复文件 ──────────────────────────────────────────
class DedupCleanerTab(QWidget):
    log_signal = Signal(str)

    PATTERNS = [
        re.compile(r'^(.*)\((\d+)\)(\.[^.]+)$'),
        re.compile(r'^(.*) - 副本(?: \((\d+)\))?(\.[^.]+)?$'),
        re.compile(r'^副本 (.*)(\.[^.]+)$'),
        re.compile(r'^(.*) - Copy(?: \((\d+)\))?(\.[^.]+)?$'),
        re.compile(r'^Copy of (.*)(\.[^.]+)$'),
        re.compile(r'^(.*) \((\d+)\)(\.[^.]+)$'),
        re.compile(r"^(.*) \([\w-]+'s conflicted copy \d{4}-\d{2}-\d{2}\)(\.[^.]+)$"),
        re.compile(r'^(.*)_conflicted copy_\d{4}-\d{2}-\d{2}(\.[^.]+)$'),
        re.compile(r'^(.*)-([0-9a-f]{8,40})(\.[^.]+)$'),
        re.compile(r'^(.*)_([0-9a-f]{8,40})(\.[^.]+)$'),
        re.compile(r'^(.*)(\.backup_\d{8})(\.[^.]+)$'),
    ]
    SKIP_DIRS = {".opencode", "__pycache__", "node_modules", ".git", ".svn"}

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self._duplicates = []
        self.init_ui()

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        desc = QLabel(
            "扫描目录中的重复文件(Windows自动重命名、中文/英文副本、OneDrive冲突/哈希备份、手动备份等11种模式)。\n"
            "删除前会列出全部匹配项并二次确认，原始文件不受影响。"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel("目标目录:"))
        self.target_dir = DirDropLineEdit("拖拽或浏览选择要清理的目录...")
        self.target_dir.setText(self.config.get("dedup_dir", ""))
        dir_row.addWidget(self.target_dir, 1)
        dir_btn = QPushButton("浏览")
        dir_btn.clicked.connect(self._browse_dir)
        dir_row.addWidget(dir_btn)
        layout.addLayout(dir_row)

        btn_row = QHBoxLayout()
        self.scan_btn = QPushButton("🔍 扫描重复文件")
        self.scan_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #1976D2; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.scan_btn.clicked.connect(self._scan)
        btn_row.addWidget(self.scan_btn)

        self.delete_btn = QPushButton("🗑️ 确认删除重复文件")
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #C62828; color: white; border: none;
                border-radius: 6px; padding: 9px 16px; font-size: 11px; font-weight: bold;
            }
            QPushButton:hover { background-color: #D32F2F; }
            QPushButton:disabled { background-color: #B0BEC5; color: #ECEFF1; }
        """)
        self.delete_btn.setEnabled(False)
        self.delete_btn.clicked.connect(self._confirm_delete)
        btn_row.addWidget(self.delete_btn)
        layout.addLayout(btn_row)

        self.stat_label = QLabel("")
        self.stat_label.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(self.stat_label)

        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(300)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

    def _browse_dir(self):
        path = QFileDialog.getExistingDirectory(self, "选择目录")
        if path:
            self.target_dir.setText(path)

    def _format_size(self, b):
        if b < 1024:
            return f"{b}B"
        elif b < 1024 ** 2:
            return f"{b / 1024:.1f}KB"
        elif b < 1024 ** 3:
            return f"{b / 1024 ** 2:.1f}MB"
        else:
            return f"{b / 1024 ** 3:.2f}GB"

    def _match(self, name, parent_path, is_dir=False):
        for pat in self.PATTERNS:
            m = pat.match(name)
            if m:
                stem = m.group(1)
                if not is_dir and m.lastindex >= 3:
                    suffix = m.group(3) or ""
                elif not is_dir:
                    suffix = m.group(2) or ""
                else:
                    suffix = ""
                orig = os.path.join(parent_path, stem + suffix)
                backup = os.path.join(parent_path, name)
                if os.path.exists(orig):
                    return (orig, backup)
        return None

    def _scan(self):
        target = self.target_dir.text().strip()
        if not target or not os.path.isdir(target):
            QMessageBox.warning(self, "提示", "请选择有效的目录")
            return
        self.config["dedup_dir"] = target

        self.scan_btn.setEnabled(False)
        self.delete_btn.setEnabled(False)
        self._duplicates = []
        self.log_area.clear()
        self.stat_label.setText("正在扫描...")
        threading.Thread(target=self._do_scan, args=(target,), daemon=True).start()

    def _do_scan(self, target):
        try:
            scanned = 0
            for dirpath, dirnames, filenames in os.walk(target):
                rel = os.path.relpath(dirpath, target)
                parts = set(p.lower() for p in (rel.split(os.sep) if rel != "." else []))
                if parts & self.SKIP_DIRS:
                    continue
                for fname in filenames:
                    scanned += 1
                    r = self._match(fname, dirpath)
                    if r:
                        self._duplicates.append(r)
                for dname in dirnames[:]:
                    r = self._match(dname, dirpath, is_dir=True)
                    if r and os.path.isdir(r[0]):
                        self._duplicates.append(r)

            self._duplicates.sort(key=lambda x: x[1])
            total_bytes = sum(
                os.path.getsize(f) for _, f in self._duplicates if os.path.isfile(f)
            )

            target_abs = os.path.abspath(target)
            self.log_signal.emit(f"扫描目录: {target}")
            self.log_signal.emit(f"扫描文件数: {scanned}")
            self.log_signal.emit(f"发现重复: {len(self._duplicates)} 个")
            self.log_signal.emit(f"可释放空间: {self._format_size(total_bytes)}")
            self.log_signal.emit("-" * 60)

            MAX_SHOW = 100
            for orig, backup in self._duplicates[:MAX_SHOW]:
                sz = os.path.getsize(backup) if os.path.isfile(backup) else 0
                rel_bak = os.path.relpath(backup, target_abs)
                rel_org = os.path.relpath(orig, target_abs)
                self.log_signal.emit(f"[删] {self._format_size(sz):>8}  {rel_bak}")
                self.log_signal.emit(f"[留]           {rel_org}")

            if len(self._duplicates) > MAX_SHOW:
                self.log_signal.emit(f"  ... 还有 {len(self._duplicates) - MAX_SHOW} 个未显示")

            self.log_signal.emit("=" * 60)
            self.stat_label.setText(
                f"扫描完成: 发现 {len(self._duplicates)} 个重复文件, "
                f"可释放 {self._format_size(total_bytes)}"
            )
            if self._duplicates:
                self.delete_btn.setEnabled(True)
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.scan_btn.setEnabled(True)

    def _confirm_delete(self):
        if not self._duplicates:
            return
        total_bytes = sum(
            os.path.getsize(f) for _, f in self._duplicates if os.path.isfile(f)
        )
        ret = QMessageBox.warning(
            self, "⚠️ 确认删除",
            f"即将删除 {len(self._duplicates)} 个重复文件\n"
            f"可释放空间: {self._format_size(total_bytes)}\n\n"
            f"原始文件不受影响，仅删除重复/备份文件。\n\n"
            f"确定要继续吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return
        self.delete_btn.setEnabled(False)
        self.scan_btn.setEnabled(False)
        self.log_signal.emit("")
        self.log_signal.emit("正在删除...")
        threading.Thread(target=self._do_delete, daemon=True).start()

    def _do_delete(self):
        try:
            ok = fail = total = 0
            target = self.target_dir.text().strip()
            log_dir = os.path.join(target, ".opencode")
            os.makedirs(log_dir, exist_ok=True)
            from datetime import datetime
            log_path = os.path.join(log_dir, f"清理重复文件_{datetime.now():%Y%m%d_%H%M%S}.log")
            log_lines = []

            for orig, backup in self._duplicates:
                try:
                    sz = os.path.getsize(backup) if os.path.isfile(backup) else 0
                    total += sz
                    if os.path.isdir(backup):
                        shutil.rmtree(backup)
                    else:
                        os.remove(backup)
                    ok += 1
                    msg = f"[已删] ({self._format_size(sz)}) {os.path.relpath(backup, target)}"
                    self.log_signal.emit(msg)
                    log_lines.append(msg)
                except Exception as e:
                    fail += 1
                    msg = f"[失败] {os.path.relpath(backup, target)} -> {e}"
                    self.log_signal.emit(msg)
                    log_lines.append(msg)

            with open(log_path, "w", encoding="utf-8") as f:
                f.write("\n".join(log_lines))

            self.log_signal.emit("")
            self.log_signal.emit(f"清理完成! 成功: {ok} | 失败: {fail} | 释放: {self._format_size(total)}")
            self.log_signal.emit(f"日志已保存: {log_path}")
            self.log_signal.emit("=" * 50)
            self.stat_label.setText(f"清理完成: 删除 {ok} 个, 失败 {fail} 个")
        except Exception as e:
            self.log_signal.emit(f"[异常] {e}")
        finally:
            self.scan_btn.setEnabled(True)
            self.delete_btn.setEnabled(False)
            self._duplicates = []


# ── Tab 8: 批量合并表格 ──────────────────────────────────────────────
"""New MergeTablesTab class - to be spliced into file_tools_page.py"""
# ── Tab 8: 批量合并表格 ──────────────────────────────────────────────
class MergeTablesTab(QWidget):
    """批量合并多个 Excel / CSV 表格，支持纵向、横向、多表合一、高级四种合并模式。"""
    log_signal = Signal(str)
    progress_signal = Signal(int)

    # 合并模式
    MODE_VERTICAL_POS = "vertical_pos"      # 纵向-按位置匹配
    MODE_VERTICAL_NAME = "vertical_name"    # 纵向-按列名匹配
    MODE_HORIZONTAL = "horizontal"          # 横向-VLOOKUP
    MODE_MULTI_SHEET = "multi_sheet"        # 多表合一
    MODE_ADVANCED = "advanced"              # 高级合并

    ADV_SAME_SHEET = "same_sheet"
    ADV_ALL_ONE = "all_one"

    DEDUP_NONE = "none"
    DEDUP_FULL_ROW = "full_row"

    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.log_signal.connect(self._append_log)
        self._file_paths = []
        self._selected_sheets = {}  # {fpath: [sheet_name, ...]}
        self.init_ui()
        self.progress_signal.connect(self.progress_bar.setValue)

    def _append_log(self, msg):
        self.log_area.append(msg)
        bar = self.log_area.verticalScrollBar()
        if bar:
            bar.setValue(bar.maximum())

    # ── 内嵌 QListWidget 子类：支持拖拽 .xlsx/.xls/.csv ──
    class _DropFileList(QListWidget):
        def __init__(self, parent_tab):
            super().__init__()
            self._pt = parent_tab
            self.setAcceptDrops(True)

        def dragEnterEvent(self, event):
            if event.mimeData().hasUrls():
                for url in event.mimeData().urls():
                    p = url.toLocalFile()
                    if p.lower().endswith(('.xlsx', '.xls', '.csv')):
                        event.acceptProposedAction()
                        return
            event.ignore()

        def dragMoveEvent(self, event):
            event.acceptProposedAction()

        def dropEvent(self, event):
            for url in event.mimeData().urls():
                p = url.toLocalFile()
                if p.lower().endswith(('.xlsx', '.xls', '.csv')):
                    self._pt._add_file(p)
            event.acceptProposedAction()

    # ========== UI ==========
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        font = QFont("Microsoft YaHei", 10)

        desc = QLabel(
            "将多个 Excel / CSV 表格合并为一个文件，支持 5 种合并模式：\n"
            "  • 纵向合并（按位置/按列名）：多个表格纵向拼接\n"
            "  • 横向合并（VLOOKUP）：按关键列拼接多个表格的列\n"
            "  • 多表合一：每个文件单独一个 Sheet\n"
            "  • 高级合并：处理多 Sheet 文件，同名合并或全部汇总"
        )
        desc.setWordWrap(True)
        desc.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(desc)

        # ── 文件列表 ──
        file_group = QGroupBox("📂 待合并文件列表")
        file_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        fg_layout = QVBoxLayout()
        fg_layout.setSpacing(6)

        self.file_list = self._DropFileList(self)
        self.file_list.setAlternatingRowColors(True)
        self.file_list.setMinimumHeight(120)
        self.file_list.setMaximumHeight(200)
        self.file_list.setStyleSheet("""
            QListWidget { border: 1px solid #CFD8DC; border-radius: 4px;
                          background: #FAFAFA; font-size: 10px; }
            QListWidget::item { padding: 4px 8px; }
            QListWidget::item:alternate { background: #F5F5F5; }
        """)
        saved = self.config.get("merge_tables_files", "")
        if saved:
            for p in saved.split(";"):
                p = p.strip()
                if p and os.path.exists(p):
                    self._add_file(p)
        fg_layout.addWidget(self.file_list)

        btn_row = QHBoxLayout()
        for text, style, cb in [
            ("➕ 添加文件", "background:#E3F2FD;color:#1565C0;border:1px solid #90CAF9;border-radius:4px;padding:6px 14px;", self._browse_files),
            ("🗑️ 移除选中", "background:#FFF3E0;color:#E65100;border:1px solid #FFCC80;border-radius:4px;padding:6px 14px;", self._remove_selected),
            ("🧹 清空列表", "background:#FBE9E7;color:#BF360C;border:1px solid #FFAB91;border-radius:4px;padding:6px 14px;", self._clear_files),
        ]:
            btn = QPushButton(text)
            btn.setFont(font)
            btn.setStyleSheet(style)
            btn.clicked.connect(cb)
            btn_row.addWidget(btn)
        btn_row.addStretch()
        fg_layout.addLayout(btn_row)
        file_group.setLayout(fg_layout)
        layout.addWidget(file_group)

        # ── 合并模式 ──
        mode_group = QGroupBox("🔀 合并模式")
        mode_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        mg_layout = QVBoxLayout()
        mg_layout.setSpacing(4)
        self.mode_buttons = {}
        saved_mode = self.config.get("merge_tables_mode", self.MODE_VERTICAL_POS)

        mode_defs = [
            (self.MODE_VERTICAL_POS, "纵向合并（按位置匹配）— 表头列顺序必须一致，直接追加数据行"),
            (self.MODE_VERTICAL_NAME, "纵向合并（按列名匹配）— 按列名自动对应，列顺序可不同"),
            (self.MODE_HORIZONTAL, "横向合并（VLOOKUP）— 按关键列（如款号）将多个表的列拼接到一起"),
            (self.MODE_MULTI_SHEET, "多表合一 — 每个文件单独放在输出文件的一个 Sheet 中"),
            (self.MODE_ADVANCED, "高级合并 — 处理多 Sheet 文件，同名合并或全部汇总到一个表"),
        ]
        for val, text in mode_defs:
            rb = QRadioButton(text)
            rb.setFont(font)
            rb.setChecked(val == saved_mode)
            rb._mode_value = val
            mg_layout.addWidget(rb)
            self.mode_buttons[val] = rb

        self._mode_group = QButtonGroup(self)
        for rb in self.mode_buttons.values():
            self._mode_group.addButton(rb)
        self._mode_group.buttonClicked.connect(self._on_mode_changed)
        mode_group.setLayout(mg_layout)
        layout.addWidget(mode_group)

        # ── 选项面板（根据模式动态切换） ──
        self.option_stack = QGroupBox("⚙️ 选项")
        self.option_stack.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.option_stack.setVisible(True)

        from PySide6.QtWidgets import QStackedWidget
        self.opt_stack = QStackedWidget()
        self._build_opt_vertical()       # index 0
        self._build_opt_vertical_name()  # index 1
        self._build_opt_horizontal()     # index 2
        self._build_opt_simple()         # index 3
        self._build_opt_advanced_sub()   # index 4

        opt_out_layout = QVBoxLayout(self.option_stack)
        opt_out_layout.setContentsMargins(8, 8, 8, 8)
        opt_out_layout.addWidget(self.opt_stack)
        layout.addWidget(self.option_stack)

        # ── Sheet 范围 ──
        sheet_group = QGroupBox("📄 Sheet 范围")
        sheet_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        sg_layout = QHBoxLayout()
        sg_layout.setSpacing(16)
        self.sheet_mode_group = QButtonGroup(self)
        saved_sheet_mode = self.config.get("merge_tables_sheet_mode", "all")
        for val, text in [
            ("all", "所有 Sheet"),
            ("first", "仅第一个 Sheet"),
            ("select", "手动选择"),
        ]:
            rb = QRadioButton(text)
            rb.setFont(font)
            rb.setChecked(val == saved_sheet_mode)
            rb._sm_value = val
            self.sheet_mode_group.addButton(rb)
            sg_layout.addWidget(rb)

        self.sheet_select_btn = QPushButton("选择 Sheet...")
        self.sheet_select_btn.setFont(font)
        self.sheet_select_btn.setStyleSheet("background:#E8F5E9;color:#2E7D32;border:1px solid #A5D6A7;border-radius:4px;padding:4px 12px;")
        self.sheet_select_btn.clicked.connect(self._open_sheet_select)
        sg_layout.addWidget(self.sheet_select_btn)
        sg_layout.addStretch()
        sheet_group.setLayout(sg_layout)
        layout.addWidget(sheet_group)

        # ── 输出文件路径 ──
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("输出文件:"))
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("选择保存合并结果的 .xlsx 文件路径...")
        self.output_path.setText(self.config.get("merge_tables_output", ""))
        out_row.addWidget(self.output_path, 1)
        out_btn = QPushButton("浏览")
        out_btn.setFont(font)
        out_btn.setFixedWidth(80)
        out_btn.clicked.connect(self._browse_output)
        out_row.addWidget(out_btn)
        layout.addLayout(out_row)

        # ── 进度条 ──
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFixedHeight(28)
        self.progress_bar.setFormat("%v / %m")
        self.progress_bar.setStyleSheet("""
            QProgressBar { border: 1px solid #B2DFDB; border-radius: 4px;
                           text-align: center; background: #E0F2F1; }
            QProgressBar::chunk { background-color: #00897B; border-radius: 3px; }
        """)
        layout.addWidget(self.progress_bar)

        # ── 开始按钮 ──
        self.run_btn = QPushButton("▶️ 开始合并")
        self.run_btn.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        self.run_btn.setFixedHeight(38)
        self.run_btn.setStyleSheet("""
            QPushButton {
                background-color: #00838F; color: white; border: none;
                border-radius: 6px; padding: 8px 20px;
            }
            QPushButton:hover { background-color: #00695C; }
            QPushButton:disabled { background-color: #B2DFDB; color: #E0F2F1; }
        """)
        self.run_btn.clicked.connect(self._start)
        layout.addWidget(self.run_btn)

        # ── 日志 ──
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet("font-family: Consolas; font-size: 9px; background: #FAFAFA;")
        layout.addWidget(self.log_area)
        layout.addStretch()
        self.setLayout(layout)

        self._on_mode_changed()

    # ========== 选项面板构建 ==========

    def _build_opt_vertical(self):
        w = QWidget()
        lo = QHBoxLayout(w)
        lo.setContentsMargins(0, 0, 0, 0)
        self.chk_vertical_source = QCheckBox("添加来源文件列（标记每行来自哪个文件）")
        self.chk_vertical_source.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.chk_vertical_source)
        self.chk_vertical_dedup = QCheckBox("合并后去重（整行去重）")
        self.chk_vertical_dedup.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.chk_vertical_dedup)
        lo.addStretch()
        self.opt_stack.addWidget(w)

    def _build_opt_vertical_name(self):
        w = QWidget()
        lo = QHBoxLayout(w)
        lo.setContentsMargins(0, 0, 0, 0)
        self.chk_vname_source = QCheckBox("添加来源文件列")
        self.chk_vname_source.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.chk_vname_source)
        self.chk_vname_dedup = QCheckBox("合并后去重（整行去重）")
        self.chk_vname_dedup.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.chk_vname_dedup)
        lo.addStretch()
        self.opt_stack.addWidget(w)

    def _build_opt_horizontal(self):
        w = QWidget()
        lo = QHBoxLayout(w)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.addWidget(QLabel("关键列:"))
        self.h_key_col = QComboBox()
        self.h_key_col.setMinimumWidth(120)
        self.h_key_col.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.h_key_col)
        lo.addWidget(QLabel("  匹配方式:"))
        self.h_join_mode = QComboBox()
        self.h_join_mode.addItem("左连接（保留全部行）", "left")
        self.h_join_mode.addItem("内连接（仅匹配的行）", "inner")
        self.h_join_mode.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.h_join_mode)
        self.chk_h_source = QCheckBox("来源标记")
        self.chk_h_source.setFont(QFont("Microsoft YaHei", 10))
        lo.addWidget(self.chk_h_source)
        lo.addStretch()
        self.opt_stack.addWidget(w)

    def _build_opt_simple(self):
        w = QWidget()
        lo = QHBoxLayout(w)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.addWidget(QLabel("（多表合一模式无额外选项）"))
        lo.addStretch()
        self.opt_stack.addWidget(w)

    def _build_opt_advanced_sub(self):
        w = QWidget()
        lo = QHBoxLayout(w)
        lo.setContentsMargins(0, 0, 0, 0)
        self.adv_mode_group = QButtonGroup(self)
        saved_adv = self.config.get("merge_tables_advanced_sub", self.ADV_SAME_SHEET)
        for val, text in [
            (self.ADV_SAME_SHEET, "同名 Sheet 合并到一起（同名合并）"),
            (self.ADV_ALL_ONE, "所有 Sheet 全部合并为一个表（全部汇总）"),
        ]:
            rb = QRadioButton(text)
            rb.setFont(QFont("Microsoft YaHei", 10))
            rb.setChecked(val == saved_adv)
            rb._adv_value = val
            self.adv_mode_group.addButton(rb)
            lo.addWidget(rb)
        lo.addStretch()
        self.opt_stack.addWidget(w)

    # ========== 模式切换 ==========

    def _get_mode(self):
        for val, rb in self.mode_buttons.items():
            if rb.isChecked():
                return val
        return self.MODE_VERTICAL_POS

    def _get_adv_sub(self):
        for btn in self.adv_mode_group.buttons():
            if btn.isChecked():
                return btn._adv_value
        return self.ADV_SAME_SHEET

    def _get_sheet_mode(self):
        for btn in self.sheet_mode_group.buttons():
            if btn.isChecked():
                return btn._sm_value
        return "all"

    def _on_mode_changed(self):
        mode = self._get_mode()
        idx_map = {
            self.MODE_VERTICAL_POS: 0,
            self.MODE_VERTICAL_NAME: 1,
            self.MODE_HORIZONTAL: 2,
            self.MODE_MULTI_SHEET: 3,
            self.MODE_ADVANCED: 4,
        }
        self.opt_stack.setCurrentIndex(idx_map.get(mode, 0))
        self.option_stack.setVisible(True)
        if mode == self.MODE_HORIZONTAL:
            self._refresh_key_cols()

    def _refresh_key_cols(self):
        self.h_key_col.clear()
        paths = self._get_file_paths()
        if not paths:
            return
        try:
            headers, _ = self._read_sheet_data(paths[0], sheet_idx=0)
            for h in headers:
                if h:
                    self.h_key_col.addItem(h)
        except Exception as e:
            self.log_signal.emit(f"[提示] 无法读取第一个文件的表头：{e}")

    # ========== 文件列表管理 ==========

    def _add_file(self, path):
        path = os.path.abspath(path)
        for i in range(self.file_list.count()):
            if self.file_list.item(i).data(Qt.UserRole) == path:
                return
        item = QListWidgetItem(os.path.basename(path))
        item.setData(Qt.UserRole, path)
        item.setToolTip(path)
        self.file_list.addItem(item)

    def _browse_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择 Excel / CSV 文件", "",
            "表格文件 (*.xlsx *.xls *.csv);;Excel (*.xlsx *.xls);;CSV (*.csv);;所有文件 (*)")
        for p in paths:
            self._add_file(p)

    def _remove_selected(self):
        for item in self.file_list.selectedItems():
            self.file_list.takeItem(self.file_list.row(item))

    def _clear_files(self):
        self.file_list.clear()

    def _browse_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存合并结果", "合并结果.xlsx", "Excel (*.xlsx)")
        if path:
            self.output_path.setText(path)

    def _get_file_paths(self):
        paths = []
        for i in range(self.file_list.count()):
            p = self.file_list.item(i).data(Qt.UserRole)
            if p:
                paths.append(p)
        return paths

    # ========== Sheet 选择 ==========

    def _open_sheet_select(self):
        files = self._get_file_paths()
        if not files:
            QMessageBox.warning(self, "提示", "请先添加文件")
            return

        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTreeWidget, QTreeWidgetItem, QDialogButtonBox

        dlg = QDialog(self)
        dlg.setWindowTitle("选择 Sheet")
        dlg.resize(500, 400)

        layout = QVBoxLayout(dlg)
        tree = QTreeWidget()
        tree.setHeaderLabels(["文件 / Sheet", ""])
        tree.setRootIsDecorated(True)

        for fpath in files:
            fname = os.path.basename(fpath)
            file_item = QTreeWidgetItem(tree, [fname, ""])
            file_item.setFlags(file_item.flags() & ~Qt.ItemIsUserCheckable)
            file_item.setExpanded(True)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
                for sname in wb.sheetnames:
                    prev_sel = self._selected_sheets.get(fpath, wb.sheetnames)
                    checked = sname in prev_sel if isinstance(prev_sel, list) else True
                    sheet_item = QTreeWidgetItem(file_item, [sname, ""])
                    sheet_item.setFlags(sheet_item.flags() | Qt.ItemIsUserCheckable)
                    sheet_item.setCheckState(0, Qt.Checked if checked else Qt.Unchecked)
                    sheet_item.setData(0, Qt.UserRole, (fpath, sname))
                wb.close()
            except Exception as e:
                QTreeWidgetItem(file_item, [f"读取失败: {e}", ""])

        layout.addWidget(tree)

        def _apply():
            self._selected_sheets = {}
            for i in range(tree.topLevelItemCount()):
                file_item = tree.topLevelItem(i)
                sheets = []
                for j in range(file_item.childCount()):
                    child = file_item.child(j)
                    if child.checkState(0) == Qt.Checked:
                        fpath, sname = child.data(0, Qt.UserRole)
                        sheets.append(sname)
                if sheets and file_item.childCount() > 0:
                    fpath, _ = file_item.child(0).data(0, Qt.UserRole)
                    self._selected_sheets[fpath] = sheets
            dlg.accept()

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(_apply)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)
        dlg.exec()

    # ========== 数据读取工具 ==========

    @staticmethod
    def _read_sheet_data(fpath, sheet_idx=0):
        """读取 Excel(.xlsx/.xls) 或 CSV 文件。
        返回: (headers: list[str], rows: list[list])"""
        ext = os.path.splitext(fpath)[1].lower()
        if ext == '.csv':
            import csv
            encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin-1']
            for enc in encodings:
                try:
                    with open(fpath, 'r', encoding=enc) as f:
                        sample = f.read(4096)
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\\t|')
                    with open(fpath, 'r', encoding=enc) as f:
                        reader = csv.reader(f, dialect)
                        rows = list(reader)
                    if rows:
                        headers = [str(h).strip() if h else '' for h in rows[0]]
                        data = [[str(c).strip() if c else '' for c in row] for row in rows[1:]]
                        return headers, data
                except Exception:
                    continue
            raise ValueError(f"无法读取 CSV 文件: {fpath}")

        wb = openpyxl.load_workbook(fpath, read_only=True)
        if sheet_idx >= len(wb.sheetnames):
            wb.close()
            raise ValueError(f"文件没有索引 {sheet_idx} 的 Sheet")
        ws = wb.worksheets[sheet_idx]
        headers = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
        data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        return headers, data

    @staticmethod
    def _read_all_sheets(fpath):
        """读取一个文件的所有 Sheet，返回 [(sheet_name, headers, rows), ...]"""
        ext = os.path.splitext(fpath)[1].lower()
        if ext == '.csv':
            headers, rows = MergeTablesTab._read_sheet_data(fpath, 0)
            return [(os.path.basename(fpath), headers, rows)]
        wb = openpyxl.load_workbook(fpath, read_only=True)
        result = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            hrow = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
            data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
            result.append((sname, hrow, data))
        wb.close()
        return result

    def _get_active_sheets(self, fpath):
        """根据 Sheet 范围设置返回要读取的 sheet 索引列表"""
        sheet_mode = self._get_sheet_mode()
        ext = os.path.splitext(fpath)[1].lower()
        if ext == '.csv':
            return [0]
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            names = wb.sheetnames
            wb.close()
        except Exception:
            return [0]
        if sheet_mode == "first":
            return [0]
        elif sheet_mode == "select":
            selected = self._selected_sheets.get(fpath, names)
            return [names.index(s) for s in selected if s in names]
        else:
            return list(range(len(names)))

    # ========== 开始合并 ==========

    def _start(self):
        files = self._get_file_paths()
        if len(files) < 2:
            QMessageBox.warning(self, "提示", "请至少添加 2 个表格文件")
            return

        output = self.output_path.text().strip()
        if not output:
            QMessageBox.warning(self, "提示", "请选择输出文件路径")
            return
        if not output.lower().endswith('.xlsx'):
            output += '.xlsx'
            self.output_path.setText(output)

        mode = self._get_mode()
        adv_sub = self._get_adv_sub() if mode == self.MODE_ADVANCED else ""
        sheet_mode = self._get_sheet_mode()

        self.config["merge_tables_files"] = ";".join(files)
        self.config["merge_tables_mode"] = mode
        self.config["merge_tables_advanced_sub"] = adv_sub
        self.config["merge_tables_sheet_mode"] = sheet_mode
        self.config["merge_tables_output"] = output
        from toolbox.core.utils import save_config
        save_config(self.config)

        vertical_opts = {}
        if mode == self.MODE_VERTICAL_POS:
            vertical_opts = dict(add_source=self.chk_vertical_source.isChecked(), dedup=self.chk_vertical_dedup.isChecked())
        elif mode == self.MODE_VERTICAL_NAME:
            vertical_opts = dict(add_source=self.chk_vname_source.isChecked(), dedup=self.chk_vname_dedup.isChecked())

        horizontal_opts = None
        if mode == self.MODE_HORIZONTAL:
            key_col = self.h_key_col.currentText()
            if not key_col:
                QMessageBox.warning(self, "提示", "请选择关键列（横向合并需要指定匹配列）")
                return
            horizontal_opts = dict(key_col=key_col, join_mode=self.h_join_mode.currentData(), add_source=self.chk_h_source.isChecked())

        self.run_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(files))
        self.progress_bar.setValue(0)
        self.log_area.clear()
        self.log_signal.emit(f"合并模式: {mode}")
        self.log_signal.emit(f"文件数: {len(files)}")
        self.log_signal.emit(f"输出: {output}\n")

        threading.Thread(
            target=self._do_merge,
            args=(files, output, mode, adv_sub, vertical_opts, horizontal_opts),
            daemon=True,
        ).start()

    def _do_merge(self, files, output, mode, adv_sub, vertical_opts, horizontal_opts):
        try:
            if mode in (self.MODE_VERTICAL_POS, self.MODE_VERTICAL_NAME):
                by_name = mode == self.MODE_VERTICAL_NAME
                self._merge_vertical(files, output, by_name, **vertical_opts)
            elif mode == self.MODE_HORIZONTAL:
                self._merge_horizontal(files, output, **horizontal_opts)
            elif mode == self.MODE_MULTI_SHEET:
                self._merge_multi_sheet(files, output)
            elif mode == self.MODE_ADVANCED:
                if adv_sub == self.ADV_SAME_SHEET:
                    self._merge_advanced_same_sheet(files, output)
                else:
                    self._merge_advanced_all_one(files, output)
        except Exception as e:
            self.log_signal.emit(f"\n[异常] {e}")
            import traceback
            self.log_signal.emit(traceback.format_exc())
        finally:
            self.run_btn.setEnabled(True)
            self.progress_bar.setVisible(False)

    # ========== 辅助：去重 ==========

    def _dedup_rows(self, headers, rows, full_row=True):
        if not rows:
            return rows
        seen = set()
        result = []
        for row in rows:
            key = tuple(row) if full_row else str(row[0]) if row else ''
            if key not in seen:
                seen.add(key)
                result.append(row)
        return result

    # ========== 模式1: 纵向合并 ==========

    def _merge_vertical(self, files, output, by_name=False, add_source=False, dedup=False):
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "合并结果"
        canonical_headers = None
        current_row = 1
        total_rows = 0
        all_rows = []

        for idx, fpath in enumerate(files):
            fname = os.path.basename(fpath)
            active_sheets = self._get_active_sheets(fpath)
            sheet_rows = 0
            for si in active_sheets:
                try:
                    headers, data = self._read_sheet_data(fpath, si)
                    if not headers or all(h == '' for h in headers):
                        self.log_signal.emit(f"  [跳过] {fname} sheet[{si}]: 空表头")
                        continue
                    if canonical_headers is None:
                        canonical_headers = headers
                        out_headers = list(canonical_headers)
                        if add_source:
                            out_headers.append("来源文件")
                        for ci, h in enumerate(out_headers, 1):
                            ws_out.cell(row=1, column=ci, value=h)
                        current_row = 2

                    if by_name:
                        col_map = {}
                        for ci, h in enumerate(headers):
                            if h in canonical_headers:
                                col_map[ci] = canonical_headers.index(h)
                        for row_data in data:
                            out_row = [None] * len(canonical_headers)
                            for sci, v in enumerate(row_data):
                                if sci in col_map:
                                    out_row[col_map[sci]] = v
                            if add_source:
                                out_row.append(fname)
                            all_rows.append(out_row)
                            sheet_rows += 1
                    else:
                        mismatch = len(headers) != len(canonical_headers)
                        if not mismatch:
                            for a, b in zip(headers, canonical_headers):
                                if a != b:
                                    mismatch = True
                                    break
                        if mismatch:
                            self.log_signal.emit(f"  [警告] {fname} sheet[{si}]: 表头不一致，按位置映射")
                        for row_data in data:
                            out_row = [row_data[ci] if ci < len(row_data) else None for ci in range(len(canonical_headers))]
                            if add_source:
                                out_row.append(fname)
                            all_rows.append(out_row)
                            sheet_rows += 1
                except Exception as e:
                    self.log_signal.emit(f"  [错误] {fname} sheet[{si}]: {e}")
            self.log_signal.emit(f"  [读取] {fname}: {sheet_rows} 行 ({len(active_sheets)} Sheet)")
            self.progress_signal.emit(idx + 1)

        if dedup and all_rows:
            before = len(all_rows)
            all_rows = self._dedup_rows(canonical_headers or [], all_rows, full_row=True)
            removed = before - len(all_rows)
            if removed:
                self.log_signal.emit(f"  [去重] 移除 {removed} 行重复数据")

        for row_data in all_rows:
            for ci, v in enumerate(row_data):
                ws_out.cell(row=current_row, column=ci + 1, value=v)
            current_row += 1
            total_rows += 1

        wb_out.save(output)
        wb_out.close()
        self.log_signal.emit(f"\n✓ 合并完成! {len(files)} 个文件, {total_rows} 行 → {os.path.basename(output)}")

    # ========== 模式2: 横向合并(VLOOKUP) ==========

    def _merge_horizontal(self, files, output, key_col="", join_mode="left", add_source=False):
        base_headers = None
        base_data = {}
        base_key_order = []
        all_headers = []
        col_source = {}

        for idx, fpath in enumerate(files):
            fname = os.path.basename(fpath)
            active_sheets = self._get_active_sheets(fpath)
            for si in active_sheets:
                try:
                    headers, data = self._read_sheet_data(fpath, si)
                    if not headers or key_col not in headers:
                        self.log_signal.emit(f"  [跳过] {fname} sheet[{si}]: 无关键列 '{key_col}'")
                        continue
                    key_idx = headers.index(key_col)

                    if idx == 0 and si == active_sheets[0] and base_headers is None:
                        base_headers = headers
                        all_headers = list(headers)
                        col_source.update({h: fname for h in headers})
                        for row_data in data:
                            k = str(row_data[key_idx]).strip() if row_data[key_idx] is not None else ''
                            if k:
                                base_data[k] = {h: row_data[ci] for ci, h in enumerate(headers)}
                                base_key_order.append(k)
                        self.log_signal.emit(f"  [基准] {fname}: {len(base_data)} 行, {len(all_headers)} 列")
                    else:
                        new_cols = [h for h in headers if h != key_col and h not in all_headers]
                        if not new_cols:
                            continue
                        col_source.update({h: fname for h in new_cols})
                        all_headers.extend(new_cols)
                        matched = 0
                        for row_data in data:
                            k = str(row_data[key_idx]).strip() if row_data[key_idx] is not None else ''
                            if k in base_data:
                                for ci, h in enumerate(headers):
                                    if h != key_col:
                                        base_data[k][h] = row_data[ci]
                                matched += 1
                        self.log_signal.emit(f"  [匹配] {fname}: 匹配 {matched} 行, 新增 {len(new_cols)} 列")
                except Exception as e:
                    self.log_signal.emit(f"  [错误] {fname} sheet[{si}]: {e}")
            self.progress_signal.emit(idx + 1)

        if not base_data:
            self.log_signal.emit("[错误] 无有效数据")
            return

        if add_source:
            all_headers.append("来源文件")
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "横向合并"
        for ci, h in enumerate(all_headers, 1):
            ws_out.cell(row=1, column=ci, value=h)

        current_row = 2
        total_rows = 0
        for k in base_key_order:
            row = base_data[k]
            out_vals = [row.get(h, None) for h in (all_headers[:-1] if add_source else all_headers)]
            if add_source:
                out_vals.append(col_source.get(all_headers[-2], '') if len(all_headers) > 1 else '')
            for ci, v in enumerate(out_vals):
                ws_out.cell(row=current_row, column=ci + 1, value=v)
            current_row += 1
            total_rows += 1

        wb_out.save(output)
        wb_out.close()
        self.log_signal.emit(f"\n✓ 横向合并完成! {total_rows} 行, {len(all_headers)} 列 → {os.path.basename(output)}")

    # ========== 模式3: 多表合一 ==========

    def _sanitize_sheet_name(self, name, used_names):
        raw = re.sub(r'[\[\]:*?/\\\\]', '_', name)
        raw = raw[:31] if raw else "Sheet"
        if not raw:
            raw = "Sheet"
        base = raw
        n = 2
        while raw in used_names:
            raw = f"{base[:27]}_{n}"
            n += 1
        used_names.add(raw)
        return raw

    def _merge_multi_sheet(self, files, output):
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        used_names = set()
        total_sheets = 0

        for idx, fpath in enumerate(files):
            fname = os.path.basename(fpath)
            active_sheets = self._get_active_sheets(fpath)
            for si in active_sheets:
                try:
                    headers, data = self._read_sheet_data(fpath, si)
                    if len(active_sheets) == 1:
                        raw_name = os.path.splitext(fname)[0]
                    else:
                        raw_name = f"{os.path.splitext(fname)[0]}_{headers[0] if headers else 'Sheet'}"
                    sheet_name = self._sanitize_sheet_name(raw_name, used_names)
                    ws_out = wb_out.create_sheet(sheet_name)
                    for ci, h in enumerate(headers, 1):
                        ws_out.cell(row=1, column=ci, value=h)
                    for ri, row_data in enumerate(data, 2):
                        for ci, v in enumerate(row_data, 1):
                            ws_out.cell(row=ri, column=ci, value=v)
                    self.log_signal.emit(f"  [多表] {fname} → Sheet「{sheet_name}」({len(data)} 行)")
                    total_sheets += 1
                except Exception as e:
                    self.log_signal.emit(f"  [错误] {fname} sheet[{si}]: {e}")
            self.progress_signal.emit(idx + 1)

        wb_out.save(output)
        wb_out.close()
        self.log_signal.emit(f"\n✓ 多表合一完成! 共 {total_sheets} 个 Sheet → {os.path.basename(output)}")

    # ========== 模式4a: 高级-同名 Sheet 合并 ==========

    def _merge_advanced_same_sheet(self, files, output):
        sheet_map = {}
        for fpath in files:
            try:
                sheets_info = self._read_all_sheets(fpath)
                if os.path.splitext(fpath)[1].lower() == '.csv':
                    for sname, h, d in sheets_info:
                        sheet_map.setdefault(sname, []).append((fpath, h, d))
                else:
                    active = self._get_active_sheets(fpath)
                    wb = openpyxl.load_workbook(fpath, read_only=True)
                    for si in active:
                        sname = wb.sheetnames[si]
                        ws = wb.worksheets[si]
                        hrow = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
                        data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
                        sheet_map.setdefault(sname, []).append((fpath, hrow, data))
                    wb.close()
            except Exception as e:
                self.log_signal.emit(f"  [跳过] {os.path.basename(fpath)}: {e}")

        if not sheet_map:
            self.log_signal.emit("[错误] 未读取到任何数据")
            return

        for sname in sorted(sheet_map.keys()):
            self.log_signal.emit(f"  [扫描] Sheet「{sname}」在 {len(sheet_map[sname])} 个文件中")

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        used_sheets = set()
        total_all = 0

        for sname, file_list in sorted(sheet_map.items()):
            out_sname = self._sanitize_sheet_name(sname, used_sheets)
            ws_out = wb_out.create_sheet(out_sname)
            canon_headers = None
            current_row = 1
            sheet_total = 0

            for fpath, hrow, data in file_list:
                fname = os.path.basename(fpath)
                if not hrow or all(h == '' for h in hrow):
                    continue
                if canon_headers is None:
                    canon_headers = hrow
                    for ci, h in enumerate(canon_headers, 1):
                        ws_out.cell(row=1, column=ci, value=h)
                    current_row = 2
                mismatch = len(hrow) != len(canon_headers)
                if not mismatch:
                    for a, b in zip(hrow, canon_headers):
                        if a != b:
                            mismatch = True
                            break
                for row_data in data:
                    vals = [row_data[ci] if ci < len(row_data) else None for ci in range(len(canon_headers))] if mismatch else list(row_data)
                    for ci, v in enumerate(vals):
                        ws_out.cell(row=current_row, column=ci + 1, value=v)
                    current_row += 1
                    sheet_total += 1
                self.log_signal.emit(f"    [{fname}] 「{sname}」: {len(data)} 行")

            self.log_signal.emit(f"  → Sheet「{out_sname}」合并: {sheet_total} 行")
            total_all += sheet_total

        wb_out.save(output)
        wb_out.close()
        self.log_signal.emit(f"\n✓ 同名 Sheet 合并完成! {len(sheet_map)} 组, {total_all} 行 → {os.path.basename(output)}")

    # ========== 模式4b: 高级-全部合并 ==========

    def _merge_advanced_all_one(self, files, output):
        all_columns = ["来源文件", "来源Sheet"]
        col_order = []
        col_set = set(all_columns)
        sheet_info = []

        for fpath in files:
            try:
                sheets = self._read_all_sheets(fpath)
                if os.path.splitext(fpath)[1].lower() == '.csv':
                    for sname, h, d in sheets:
                        for hh in h:
                            if hh and hh not in col_set:
                                col_set.add(hh)
                                col_order.append(hh)
                        sheet_info.append((fpath, sname, h, d))
                else:
                    active = self._get_active_sheets(fpath)
                    wb = openpyxl.load_workbook(fpath, read_only=True)
                    for si in active:
                        sname = wb.sheetnames[si]
                        ws = wb.worksheets[si]
                        hrow = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(max_row=1, values_only=True))]
                        data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
                        for hh in hrow:
                            if hh and hh not in col_set:
                                col_set.add(hh)
                                col_order.append(hh)
                        sheet_info.append((fpath, sname, hrow, data))
                    wb.close()
            except Exception as e:
                self.log_signal.emit(f"  [跳过] {os.path.basename(fpath)}: {e}")

        if not sheet_info:
            self.log_signal.emit("[错误] 未读取到任何数据")
            return

        full_headers = all_columns + col_order
        hdr_index = {h: i for i, h in enumerate(full_headers)}
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "全部合并"
        for ci, h in enumerate(full_headers, 1):
            ws_out.cell(row=1, column=ci, value=h)

        current_row = 2
        total_rows = 0
        for fpath, sname, hrow, data in sheet_info:
            fname = os.path.basename(fpath)
            col_map = {}
            for ci, h in enumerate(hrow):
                if h and h in hdr_index:
                    col_map[ci] = hdr_index[h]
            count = 0
            for row_data in data:
                out_vals = [None] * len(full_headers)
                out_vals[hdr_index["来源文件"]] = fname
                out_vals[hdr_index["来源Sheet"]] = sname
                for sci, v in enumerate(row_data):
                    if sci in col_map:
                        out_vals[col_map[sci]] = v
                for ci, v in enumerate(out_vals):
                    ws_out.cell(row=current_row, column=ci + 1, value=v)
                current_row += 1
                count += 1
            self.log_signal.emit(f"  [{fname}] 「{sname}」: {count} 行")
            total_rows += count

        wb_out.save(output)
        wb_out.close()
        self.log_signal.emit(f"\n✓ 全部合并完成! {len(sheet_info)} 个 Sheet, {total_rows} 行, {len(full_headers)} 列 → {os.path.basename(output)}")
# ── 主页面：文件工具 ──────────────────────────────────────────────
class FileToolsPage(QWidget):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(6)
        self.setLayout(layout)

        # 全局样式
        self.setStyleSheet("""
            QGroupBox {
                border: 1px solid #D6E4F0;
                border-radius: 8px;
                margin-top: 10px;
                background: #FFFFFF;
                font-family: Microsoft YaHei;
                font-size: 10px;
                font-weight: bold;
                color: #1565C0;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 6px;
            }
            QLabel {
                color: #333333;
                font-family: Microsoft YaHei;
                font-size: 10px;
            }
            QLineEdit, QComboBox {
                border: 1px solid #CFD8DC;
                border-radius: 6px;
                padding: 6px 8px;
                background: #FAFAFA;
                min-height: 26px;
                font-family: Microsoft YaHei;
                font-size: 10px;
            }
            QLineEdit:focus, QComboBox:focus {
                border: 1px solid #1565C0;
                background: #FFFFFF;
            }
            QCheckBox {
                color: #37474F;
                font-family: Microsoft YaHei;
                font-size: 10px;
            }
            QTextEdit {
                border: 1px solid #CFD8DC;
                border-radius: 6px;
                background: #FAFAFA;
                font-family: Consolas, Microsoft YaHei;
                font-size: 9px;
            }
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #90CAF9;
                border-radius: 6px;
                padding: 7px 12px;
                font-family: Microsoft YaHei;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
            QTabWidget::pane {
                border: 1px solid #D6E4F0;
                border-radius: 8px;
                background: #FFFFFF;
            }
            QTabBar::tab {
                background: #E3F2FD;
                color: #1565C0;
                border: 1px solid #BBDEFB;
                border-bottom: none;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                padding: 8px 16px;
                font-family: Microsoft YaHei;
                font-size: 10px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #FFFFFF;
                color: #0D47A1;
                font-weight: bold;
                border-bottom: 2px solid #1565C0;
            }
            QTabBar::tab:hover {
                background: #BBDEFB;
            }
        """)

        title = QLabel("📁 文件工具")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1565C0;")
        layout.addWidget(title)

        desc = QLabel("文件夹批量操作 & 表格驱动工具集，涵盖建文件夹、提取汇总、表格合并、分发、清单导出等日常办公高频操作")
        desc.setWordWrap(True)
        desc.setFont(QFont("Microsoft YaHei", 10))
        desc.setStyleSheet("color: #666;")
        layout.addWidget(desc)

        # 滚动区域包裹标签页
        self.tabs = QTabWidget()

        self.tab1 = FolderCreatorTab(self.config)
        self.tab2 = FileCollectExtractTab(self.config)
        self.tab3 = ImageToFolderTab(self.config)
        self.tab4 = FileDistributeTab(self.config)
        self.tab5 = SupplierCodeTab(self.config)
        self.tab6 = MergeTablesTab(self.config)
        self.tab7 = FileListExportTab(self.config)
        self.tab8 = DedupCleanerTab(self.config)

        # 包裹到 QScrollArea
        for tab, name in [
            (self.tab1, "📁 批量建文件夹"),
            (self.tab2, "📦 提取文件汇总"),
            (self.tab3, "🖼️ 图片归入同名文件夹"),
            (self.tab4, "📨 批量分发文件"),
            (self.tab5, "🔖 供应商款号提取"),
            (self.tab6, "📊 批量合并表格"),
            (self.tab7, "📋 导出文件清单"),
            (self.tab8, "🧹 清理重复文件"),
        ]:
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(tab)
            scroll.setFrameShape(QScrollArea.NoFrame)
            self.tabs.addTab(scroll, name)

        layout.addWidget(self.tabs)
