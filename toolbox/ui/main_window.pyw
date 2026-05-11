import sys
import os
import time
import threading
import queue
import random
import subprocess
import json
import io

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
    QHBoxLayout, QStackedWidget, QListWidget, QListWidgetItem, QLabel, 
    QPushButton, QLineEdit, QTextEdit, QProgressBar, QFileDialog, QMessageBox,
    QCheckBox, QRadioButton, QButtonGroup, QSpinBox, QTableWidget, QTableWidgetItem,
    QFrame, QScrollArea, QGroupBox, QDialog, QDialogButtonBox, QFormLayout,
    QComboBox, QTabWidget, QToolButton, QSplitter, QScrollBar, QSizePolicy, QGridLayout,
    QTableWidget, QHeaderView, QAbstractItemView, QProgressDialog)
from PySide6.QtCore import Qt, QSize, QTimer, QThread, Signal, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QFont, QColor, QPalette, QIcon, QAction, QCursor, QPixmap, QImage

from toolbox.core.utils import (APP_ROOT, BASE_DIR, CONFIG_FILE, load_config, save_config, 
    clean, load_anti_ban_config, load_proxies, get_chrome_ua,
    compress_image_to_size_v2,
    REFERENCE_DIR, get_ocr_feature_tier)
from toolbox.core.size_mapping import (
    canonicalize_size,
    ensure_default_mapping_files,
    load_external_field_aliases,
    load_external_size_aliases,
    MAPPINGS_DIR,
)
from toolbox.core.browser import PW, DownloadHistory, batch_cutout_skus_under_root

from toolbox.ui.vip_image_finder_page import VipImageFinderPage
from toolbox.ui.history_page import HistoryPage
from toolbox.ui.rename_page import RenamePage
from toolbox.ui.compress_page import CompressPage as SplitCompressPage
from toolbox.ui.path_drop import DirDropLineEdit, ExcelDropLineEdit, enable_path_drop
from toolbox.ui.anti_ban_dialog import AntiBanSettingsDialog
from toolbox.ui import pdf_helpers
from toolbox.core import pdf_edit_core
from toolbox.ui.pages.batch_erp_page import BatchErpPage
from toolbox.ui.file_tools_page import FileToolsPage


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = load_config()
        self.init_ui()
    
    def _get_adaptive_sizes(self):
        """根据屏幕分辨率计算自适应尺寸（支持8档）"""
        screen = QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry()
        else:
            from PySide6.QtWidgets import QDesktopWidget
            available = QDesktopWidget().availableGeometry()
        
        screen_w, screen_h = available.width(), available.height()
        
        # T1: 紧凑型 (1920×1080 及以下)
        if screen_w <= 1920:
            return {
                'min_width': 1100, 'min_height': 640,
                'width': 1280, 'height': 720,
                'max_width': 1536, 'max_height': 864,
                'sidebar': 180
            }
        
        # T2: 标准型 (1920-2200)
        elif screen_w <= 2200:
            return {
                'min_width': 1120, 'min_height': 660,
                'width': 1400, 'height': 840,
                'max_width': 1680, 'max_height': 990,
                'sidebar': 190
            }
        
        # T3: 中小型 (2200-2700) - 13.3寸2K笔记本, 27寸2K显示器
        elif screen_w <= 2700:
            return {
                'min_width': 1280, 'min_height': 720,
                'width': 1600, 'height': 900,
                'max_width': 1920, 'max_height': 1080,
                'sidebar': 200
            }
        
        # T4: 中型 (2700-3000) - 14寸2K笔记本
        elif screen_w <= 3000:
            return {
                'min_width': 1280, 'min_height': 800,
                'width': 1600, 'height': 1000,
                'max_width': 1920, 'max_height': 1200,
                'sidebar': 210
            }
        
        # T5: 大型 (3000-3200) - 14寸高分屏, 16寸笔记本
        elif screen_w <= 3200:
            return {
                'min_width': 1440, 'min_height': 900,
                'width': 1800, 'height': 1125,
                'max_width': 2160, 'max_height': 1350,
                'sidebar': 220
            }
        
        # T6: 超大型 (3200-4000) - 15.6寸3K, 17寸工作站
        elif screen_w <= 4000:
            return {
                'min_width': 1536, 'min_height': 960,
                'width': 1920, 'height': 1200,
                'max_width': 2304, 'max_height': 1440,
                'sidebar': 230
            }
        
        # T7: 巨型 (4000-4500) - 27寸4K显示器
        elif screen_w <= 4500:
            return {
                'min_width': 1600, 'min_height': 900,
                'width': 2000, 'height': 1125,
                'max_width': 2400, 'max_height': 1350,
                'sidebar': 250
            }
        
        # T8: 超巨型 (>4500) - 5K显示器
        else:
            return {
                'min_width': 1800, 'min_height': 1012,
                'width': 2200, 'height': 1237,
                'max_width': 2800, 'max_height': 1575,
                'sidebar': 280
            }
    
    def init_ui(self):
        # 版本号与项目根目录 VERSION 文件一致
        try:
            _vf = os.path.join(APP_ROOT, "VERSION")
            with open(_vf, "r", encoding="utf-8") as _f:
                _ver = (_f.read() or "").strip() or "4.18"
        except Exception:
            _ver = "4.18"
        self.setWindowTitle(f"有想法唯品上新工具箱 v{_ver} - PySide6版")
        
        # 自适应屏幕尺寸
        sizes = self._get_adaptive_sizes()
        self.setMinimumSize(sizes['min_width'], sizes['min_height'])
        self.resize(sizes['width'], sizes['height'])
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        central_widget.setLayout(main_layout)
        
        self.create_sidebar()
        self.create_content_area()
        
        main_layout.addWidget(self.sidebar)
        main_layout.addWidget(self.content_stack)
        self.content_stack.setMinimumSize(680, 520)
        
        self.set_default_page()
    
    def create_sidebar(self):
        self.sidebar = QFrame()
        # 自适应侧边栏宽度
        sizes = self._get_adaptive_sizes()
        self.sidebar.setFixedWidth(sizes['sidebar'])
        self.sidebar.setStyleSheet("""
            QFrame {
                background-color: #1E88E5;
            }
        """)
        
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)
        self.sidebar.setLayout(sidebar_layout)
        
        title_frame = QFrame()
        title_frame.setFixedHeight(84)
        title_frame.setStyleSheet("background-color: #1565C0;")
        title_layout = QVBoxLayout(title_frame)
        title_layout.setContentsMargins(12, 8, 12, 8)
        title_layout.setSpacing(2)
        
        title_label = QLabel("有想法唯品上新工具箱")
        title_label.setFont(QFont("Microsoft YaHei", 13, QFont.Bold))
        title_label.setStyleSheet("color: white;")
        title_label.setWordWrap(True)
        
        subtitle_label = QLabel("开发者：郭大旭")
        subtitle_label.setFont(QFont("Microsoft YaHei", 8))
        subtitle_label.setStyleSheet("color: #BBDEFB;")
        
        title_layout.addWidget(title_label)
        title_layout.addWidget(subtitle_label)
        
        sidebar_layout.addWidget(title_frame)
        
        self.menu_list = QListWidget()
        self.menu_list.setStyleSheet("""
            QListWidget {
                background-color: transparent;
                border: none;
                padding-top: 10px;
            }
            QListWidget::item {
                color: white;
                padding: 12px 15px;
                font-size: 13px;
                font-family: Microsoft YaHei;
            }
            QListWidget::item:selected {
                background-color: #1565C0;
                border-left: 4px solid #FFC107;
            }
            QListWidget::item:hover {
                background-color: #1976D2;
            }
        """)
        self.menu_list.setFocusPolicy(Qt.NoFocus)
        self.menu_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.menu_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        menu_items = [
            ("🔎", "唯品批量找图"),
            ("🛒", "淘宝图片下载"),
            ("🖼️", "图片压缩"),
            ("📄", "PDF工具"),
            ("✂️", "批量抠图"),
            ("📏", "批量尺码表录入"),
            ("📑", "批量模板 & ERP"),
            ("📝", "批量重命名"),
            ("📁", "文件工具"),
            ("📋", "操作历史"),
        ]
        
        for icon, text in menu_items:
            item = QListWidgetItem(f"{icon}  {text}")
            item.setSizeHint(QSize(200, 44))
            self.menu_list.addItem(item)
        
        self.menu_list.currentRowChanged.connect(self.on_menu_changed)
        sidebar_layout.addWidget(self.menu_list)
        sidebar_layout.setStretch(1, 1)
        self._refresh_sidebar_item_sizes()
    
    def create_content_area(self):
        self.content_stack = QStackedWidget()
        self.content_stack.setStyleSheet("background-color: #F5F5F5;")

        # 页面按需创建：先放占位页，用户切换到对应菜单时再真正初始化，减少启动卡顿
        self._page_factories = [
            lambda: VipImageFinderPage(self.config),
            lambda: DownloadPage(self.config),
            lambda: SplitCompressPage(self.config),
            lambda: PDFPage(self.config),
            lambda: CutoutPage(self.config),
            lambda: SizeTablePage(self.config),
            lambda: BatchErpPage(self.config),
            lambda: RenamePage(self.config),
            lambda: FileToolsPage(self.config),
            lambda: HistoryPage(self.config),
        ]
        self._pages = [None] * len(self._page_factories)

        for _ in self._page_factories:
            holder = QWidget()
            holder_layout = QVBoxLayout(holder)
            holder_layout.setContentsMargins(20, 20, 20, 20)
            loading = QLabel("页面加载中，请稍候...")
            loading.setAlignment(Qt.AlignCenter)
            loading.setStyleSheet("color: #666; font-family: Microsoft YaHei; font-size: 14px;")
            holder_layout.addWidget(loading)
            self.content_stack.addWidget(holder)

        # 兼容旧逻辑：这些属性在首次创建对应页面后会被赋值
        self.vip_finder_page = None
        self.download_page = None
        self.compress_page = None
        self.pdf_page = None
        self.cutout_page = None
        self.sizetable_page = None
        self.batch_erp_page = None
        self.rename_page = None
        self.file_tools_page = None
        self.history_page = None

    def _ensure_page(self, index):
        if index < 0 or index >= len(self._pages):
            return
        if self._pages[index] is not None:
            return
        page = self._page_factories[index]()
        self._pages[index] = page
        self.content_stack.removeWidget(self.content_stack.widget(index))
        self.content_stack.insertWidget(index, page)

        if index == 0:
            self.vip_finder_page = page
        elif index == 1:
            self.download_page = page
        elif index == 2:
            self.compress_page = page
        elif index == 3:
            self.pdf_page = page
        elif index == 4:
            self.cutout_page = page
        elif index == 5:
            self.sizetable_page = page
        elif index == 6:
            self.batch_erp_page = page
        elif index == 7:
            self.rename_page = page
        elif index == 8:
            self.file_tools_page = page
        elif index == 9:
            self.history_page = page
    
    def set_default_page(self):
        self.menu_list.setCurrentRow(0)
    
    def on_menu_changed(self, index):
        self._ensure_page(index)
        self.content_stack.setCurrentIndex(index)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._refresh_sidebar_item_sizes()

    def _refresh_sidebar_item_sizes(self):
        if not hasattr(self, "menu_list") or self.menu_list.count() == 0:
            return
        # 保证侧栏按钮始终全部可见，不依赖滚动条
        avail = max(360, self.sidebar.height() - 80)
        per_h = max(34, min(52, int(avail / self.menu_list.count())))
        for i in range(self.menu_list.count()):
            self.menu_list.item(i).setSizeHint(QSize(200, per_h))
    
    def closeEvent(self, event):
        self.save_all_settings()
        event.accept()
    
    def save_all_settings(self):
        try:
            if hasattr(self.compress_page, 'save_settings'):
                self.compress_page.save_settings()
            if hasattr(self.cutout_page, 'save_settings'):
                self.cutout_page.save_settings()
            if hasattr(self.sizetable_page, 'save_settings'):
                self.sizetable_page.save_settings()
            if hasattr(self.rename_page, 'save_settings'):
                self.rename_page.save_settings()
            save_config(self.config)
        except Exception as e:
            print(f"保存配置失败: {e}")


class DownloadPage(QWidget):
    progress_signal = Signal(tuple)
    log_signal = Signal(str)
    done_signal = Signal()
    captcha_need_signal = Signal()
    status_signal = Signal(str, str)
    # 登录子线程结束（仅当启动失败时需恢复「登录」按钮，避免卡死点不了）
    login_launch_done = Signal(bool)
    
    def __init__(self, config, parent=None):
        self.dl_pw = None
        self.dl_running = False
        self.dl_q = queue.Queue()
        self.config = config
        super().__init__(parent)
        
        self.progress_signal.connect(self.on_progress)
        self.log_signal.connect(self.on_log)
        self.done_signal.connect(self.on_done)
        self.captcha_need_signal.connect(self._on_captcha_need_slot)
        self.status_signal.connect(self.on_status_update)
        self.login_launch_done.connect(self._on_login_launch_done)
        
        self.init_ui()
    
    def init_ui(self):
        root_layout = QVBoxLayout()
        root_layout.setContentsMargins(20, 20, 20, 20)
        root_layout.setSpacing(12)
        self.setLayout(root_layout)
        
        # 标题
        title = QLabel("淘宝图片批量下载")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1565C0;")
        root_layout.addWidget(title)
        
        desc = QLabel("从淘宝商品链接批量下载主图、SKU图、详情图，支持断点续传")
        desc.setFont(QFont("Microsoft YaHei", 10))
        desc.setStyleSheet("color: #666;")
        root_layout.addWidget(desc)
        
        # 使用说明框
        info_frame = QFrame()
        info_frame.setStyleSheet("background-color: #E3F2FD; padding: 10px; border-radius: 4px; border-left: 4px solid #1565C0;")
        info_label = QLabel("📌 使用说明：①选择Excel和保存目录 → ②登录淘宝 → ③开始下载 | 登录只需一次，自动保存")
        info_label.setFont(QFont("Microsoft YaHei", 9))
        info_label.setStyleSheet("color: #1565C0;")
        info_frame.setLayout(QVBoxLayout())
        info_frame.layout().addWidget(info_label)
        root_layout.addWidget(info_frame)

        body_scroll = QScrollArea()
        body_scroll.setWidgetResizable(True)
        body_scroll.setFrameShape(QFrame.NoFrame)
        body_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        body_scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
        """)
        body_widget = QWidget()
        layout = QVBoxLayout(body_widget)
        layout.setContentsMargins(0, 0, 8, 0)
        layout.setSpacing(15)
        
        # 主内容区域
        content_frame = QFrame()
        content_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        content_layout = QVBoxLayout()
        content_layout.setSpacing(12)
        content_frame.setLayout(content_layout)
        
        # Excel文件行
        excel_row = QHBoxLayout()
        excel_label = QLabel("Excel文件:")
        excel_label.setFont(QFont("Microsoft YaHei", 10))
        excel_label.setFixedWidth(90)
        excel_row.addWidget(excel_label)
        
        self.excel_path = ExcelDropLineEdit()
        self.excel_path.setText(self.config.get("last_excel", ""))
        self.excel_path.setFont(QFont("Microsoft YaHei", 10))
        self.excel_path.setPlaceholderText("点击浏览或直接拖拽 .xlsx/.xls 到此处...")
        self.excel_path.excel_dropped.connect(self.on_excel_dropped)
        self.excel_path.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus {
                border: 2px solid #1565C0;
                background-color: white;
            }
        """)
        excel_row.addWidget(self.excel_path)
        
        btn_excel = QPushButton("浏览")
        btn_excel.setFont(QFont("Microsoft YaHei", 9))
        btn_excel.setFixedWidth(80)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
        """)
        btn_excel.clicked.connect(self.select_excel)
        excel_row.addWidget(btn_excel)
        content_layout.addLayout(excel_row)
        
        # Excel字段说明
        excel_tip = QLabel("📌 Excel表头必须有：①「商品链接」列（包含taobao.com） ②「唯品款号」列（必须包含这两个关键词，精确匹配）")
        excel_tip.setFont(QFont("Microsoft YaHei", 8))
        excel_tip.setStyleSheet("color: #D32F2F; padding-left: 90px;")
        content_layout.addWidget(excel_tip)
        
        # 保存目录行
        dir_row = QHBoxLayout()
        dir_label = QLabel("保存目录:")
        dir_label.setFont(QFont("Microsoft YaHei", 10))
        dir_label.setFixedWidth(90)
        dir_row.addWidget(dir_label)
        
        self.save_dir = QLineEdit()
        enable_path_drop(self.save_dir, mode="dir")
        self.save_dir.setText(self.config.get("last_dir", os.path.join(BASE_DIR, "下载的商品图片")))
        self.save_dir.setFont(QFont("Microsoft YaHei", 10))
        self.save_dir.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus {
                border: 2px solid #1565C0;
                background-color: white;
            }
        """)
        dir_row.addWidget(self.save_dir)
        
        btn_dir = QPushButton("浏览")
        btn_dir.setFont(QFont("Microsoft YaHei", 9))
        btn_dir.setFixedWidth(80)
        btn_dir.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
        """)
        btn_dir.clicked.connect(self.select_dir)
        dir_row.addWidget(btn_dir)
        content_layout.addLayout(dir_row)
        
        # 信息显示行
        info_row = QHBoxLayout()
        self.link_col_label = QLabel("链接列: (选择Excel后自动识别)")
        self.link_col_label.setFont(QFont("Microsoft YaHei", 9))
        self.link_col_label.setStyleSheet("color: #666;")
        info_row.addWidget(self.link_col_label)
        
        self.code_col_label = QLabel("款号列: (选择Excel后自动识别)")
        self.code_col_label.setFont(QFont("Microsoft YaHei", 9))
        self.code_col_label.setStyleSheet("color: #666;")
        info_row.addWidget(self.code_col_label)
        
        self.rows_label = QLabel("行数: -")
        self.rows_label.setFont(QFont("Microsoft YaHei", 9))
        self.rows_label.setStyleSheet("color: #666;")
        info_row.addWidget(self.rows_label)
        info_row.addStretch()
        content_layout.addLayout(info_row)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("background-color: #E0E0E0;")
        line.setFixedHeight(1)
        content_layout.addWidget(line)
        
        # 登录模式选择
        login_section = QHBoxLayout()
        
        login_mode_label = QLabel("登录模式:")
        login_mode_label.setFont(QFont("Microsoft YaHei", 10))
        login_section.addWidget(login_mode_label)
        
        self.login_mode = QRadioButton("消费者淘宝")
        self.login_mode.setChecked(True)
        self.login_mode.setFont(QFont("Microsoft YaHei", 9))
        self.login_mode.setStyleSheet("color: #333;")
        login_section.addWidget(self.login_mode)
        
        self.login_mode_seller = QRadioButton("卖家中心(子账号)")
        self.login_mode_seller.setFont(QFont("Microsoft YaHei", 9))
        self.login_mode_seller.setStyleSheet("color: #333;")
        login_section.addWidget(self.login_mode_seller)
        
        self.login_mode_qianniu = QRadioButton("千牛工作台")
        self.login_mode_qianniu.setFont(QFont("Microsoft YaHei", 9))
        self.login_mode_qianniu.setStyleSheet("color: #333;")
        login_section.addWidget(self.login_mode_qianniu)
        
        login_section.addStretch()
        
        self.login_tip = QLabel("")
        self.login_tip.setFont(QFont("Microsoft YaHei", 8))
        self.login_tip.setStyleSheet("color: #FF9800;")
        login_section.addWidget(self.login_tip)
        content_layout.addLayout(login_section)
        
        # 操作按钮区：两行布局，避免非最大化时按钮被挤出界面
        btn_row = QGridLayout()
        btn_row.setHorizontalSpacing(10)
        btn_row.setVerticalSpacing(8)
        
        self.login_btn = QPushButton("🔐 打开登录页面")
        self.login_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.login_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.login_btn.clicked.connect(self.do_login)
        self.login_btn.setMinimumWidth(150)
        btn_row.addWidget(self.login_btn, 0, 0)
        
        self.confirm_login_btn = QPushButton("✓ 我已在浏览器中完成登录")
        self.confirm_login_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.confirm_login_btn.setStyleSheet("""
            QPushButton {
                background-color: #43A047;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
            }
        """)
        self.confirm_login_btn.clicked.connect(self.confirm_login)
        self.confirm_login_btn.hide()
        self.confirm_login_btn.setMinimumWidth(180)
        btn_row.addWidget(self.confirm_login_btn, 0, 1)
        
        self.start_btn = QPushButton("⬇️ 开始下载")
        self.start_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #43A047;
                color: white;
                border: none;
                padding: 10px 25px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.start_btn.clicked.connect(self.start_download)
        self.start_btn.setMinimumWidth(132)
        btn_row.addWidget(self.start_btn, 1, 0)
        
        self.stop_btn = QPushButton("⏹ 停止")
        self.stop_btn.setFont(QFont("Microsoft YaHei", 10))
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                padding: 10px 16px;
                border-radius: 4px;
            }
        """)
        self.stop_btn.clicked.connect(self.stop_download)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setMinimumWidth(92)
        btn_row.addWidget(self.stop_btn, 1, 1)
        
        self.status_label = QLabel("未登录")
        self.status_label.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.status_label.setStyleSheet("color: #D32F2F;")
        btn_row.addWidget(self.status_label, 2, 0, 1, 3)
        btn_row.setColumnStretch(2, 1)

        content_layout.addLayout(btn_row)
        
        # SKU 透明抠图（与防封分开，下载前必选方式）
        matting_row = QHBoxLayout()
        matting_label = QLabel("SKU抠透明:")
        matting_label.setFont(QFont("Microsoft YaHei", 9))
        matting_label.setStyleSheet("color: #333;")
        matting_row.addWidget(matting_label)
        self.tb_matting_group = QButtonGroup(self)
        self.tb_matting_none = QRadioButton("不抠图")
        self.tb_matting_each = QRadioButton("每款下载时抠")
        self.tb_matting_batch = QRadioButton("全部下完再抠")
        for rb in (self.tb_matting_none, self.tb_matting_each, self.tb_matting_batch):
            rb.setFont(QFont("Microsoft YaHei", 9))
            rb.setStyleSheet("color: #333;")
        _mm = self.config.get("taobao_matting_mode", "none")
        if _mm == "each":
            self.tb_matting_each.setChecked(True)
        elif _mm == "batch":
            self.tb_matting_batch.setChecked(True)
        else:
            self.tb_matting_none.setChecked(True)
        self.tb_matting_group.addButton(self.tb_matting_none, 0)
        self.tb_matting_group.addButton(self.tb_matting_each, 1)
        self.tb_matting_group.addButton(self.tb_matting_batch, 2)
        matting_row.addWidget(self.tb_matting_none)
        matting_row.addWidget(self.tb_matting_each)
        matting_row.addWidget(self.tb_matting_batch)
        matting_engine_label = QLabel("⚙ 引擎:")
        matting_engine_label.setFont(QFont("Microsoft YaHei", 9))
        matting_engine_label.setStyleSheet("color: #1565C0; font-weight: 600;")
        matting_row.addWidget(matting_engine_label)
        self.tb_matting_engine = QComboBox()
        self.tb_matting_engine.addItems(["仅本地 U2NET.ONNX"])
        self.tb_matting_engine.setCurrentIndex(0)
        self.tb_matting_engine.setEnabled(False)
        self.tb_matting_engine.setToolTip("下载页 SKU 抠图仅使用本地 U2NET.ONNX 模型")
        self.tb_matting_engine.setStyleSheet("""
            QComboBox {
                border: 1px solid #BBDEFB;
                border-radius: 6px;
                background: #E3F2FD;
                color: #1565C0;
                padding: 4px 8px;
                min-height: 24px;
            }
            QComboBox:focus {
                border: 2px solid #1565C0;
            }
        """)
        matting_row.addWidget(self.tb_matting_engine)
        self.tb_matting_parallel = QCheckBox("🚀 下载时并发抠图")
        self.tb_matting_parallel.setFont(QFont("Microsoft YaHei", 9))
        self.tb_matting_parallel.setChecked(bool(self.config.get("taobao_matting_parallel", True)))
        self.tb_matting_parallel.setToolTip("开启后下载和抠图并行，通常更快")
        self.tb_matting_parallel.setStyleSheet("color: #1565C0;")
        matting_row.addWidget(self.tb_matting_parallel)
        self.tb_matting_workers = QSpinBox()
        self.tb_matting_workers.setRange(1, 6)
        self.tb_matting_workers.setValue(int(self.config.get("taobao_matting_workers", 2)))
        self.tb_matting_workers.setPrefix("并发")
        self.tb_matting_workers.setSuffix("路")
        self.tb_matting_workers.setFixedWidth(90)
        self.tb_matting_workers.setStyleSheet("""
            QSpinBox {
                border: 1px solid #BBDEFB;
                border-radius: 6px;
                background: #E3F2FD;
                color: #1565C0;
                padding: 3px 6px;
                min-height: 24px;
            }
            QSpinBox:focus {
                border: 2px solid #1565C0;
            }
        """)
        matting_row.addWidget(self.tb_matting_workers)
        matting_row.addStretch()
        content_layout.addLayout(matting_row)

        matting_row2 = QHBoxLayout()
        matting_tip = QLabel("（需已安装 rembg 且根目录存在 u2net.onnx；「全部下完再抠」仅在本次未被中途停止时执行）")
        matting_tip.setFont(QFont("Microsoft YaHei", 8))
        matting_tip.setStyleSheet("color: #888;")
        matting_row2.addWidget(matting_tip)
        self.tb_detail_one_pass = QCheckBox("详情区只滑1遍（更快，极个别款可能少图）")
        self.tb_detail_one_pass.setFont(QFont("Microsoft YaHei", 9))
        self.tb_detail_one_pass.setChecked(self.config.get("taobao_detail_one_pass", False))
        self.tb_detail_one_pass.setStyleSheet("color: #333;")
        matting_row2.addWidget(self.tb_detail_one_pass)
        self.tb_fast_page_scroll = QCheckBox("整页快速滚动(不用平滑)")
        self.tb_fast_page_scroll.setFont(QFont("Microsoft YaHei", 9))
        self.tb_fast_page_scroll.setChecked(self.config.get("taobao_fast_page_scroll", False))
        self.tb_fast_page_scroll.setToolTip("与防封里的「模拟人类滚动」二选一效果：勾选后整页一次拉到底，可省时间")
        self.tb_fast_page_scroll.setStyleSheet("color: #333;")
        matting_row2.addWidget(self.tb_fast_page_scroll)
        matting_row2.addStretch()
        content_layout.addLayout(matting_row2)
        
        # 选项设置
        options_row = QHBoxLayout()
        
        self.resume_check = QCheckBox("启用断点续传 (跳过已下载的商品)")
        self.resume_check.setChecked(self.config.get("resume_enabled", True))
        self.resume_check.setFont(QFont("Microsoft YaHei", 9))
        options_row.addWidget(self.resume_check)
        
        options_row.addStretch()
        
        anti_ban_btn = QPushButton("🛡️ 防封设置")
        anti_ban_btn.setFont(QFont("Microsoft YaHei", 9))
        anti_ban_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
            }
        """)
        anti_ban_btn.clicked.connect(self.show_anti_ban_settings)
        options_row.addWidget(anti_ban_btn)
        content_layout.addLayout(options_row)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                text-align: center;
                background-color: #F5F5F5;
            }
            QProgressBar::chunk {
                background-color: #43A047;
            }
        """)
        content_layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("就绪")
        self.progress_label.setFont(QFont("Microsoft YaHei", 9))
        self.progress_label.setStyleSheet("color: #666;")
        content_layout.addWidget(self.progress_label)
        
        layout.addWidget(content_frame)
        
        # 日志区域
        log_group = QGroupBox("📝 操作日志")
        log_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        log_group.setStyleSheet("QGroupBox { border: 1px solid #E0E0E0; border-radius: 4px; margin-top: 10px; padding-top: 10px; }")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        self.log_text.setStyleSheet("background-color: #F8F9FA; color: #333; border: none; padding: 5px;")
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        layout.addWidget(log_group)

        body_scroll.setWidget(body_widget)
        root_layout.addWidget(body_scroll, 1)
        
        if self.excel_path.text() and os.path.exists(self.excel_path.text()):
            self.analyze_excel(self.excel_path.text())
    
    def on_progress(self, data):
        pct, text = data
        self.progress_bar.setValue(int(pct))
        self.progress_label.setText(text)
    
    def on_log(self, msg):
        self.log_text.append(msg)

    def on_status_update(self, text, color):
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color: {color}; font-weight: bold;")
    
    def on_done(self):
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_label.setText("完成")
        win = self.window()
        if hasattr(win, "_active_task_owner") and getattr(win, "_active_task_owner", None) == "download":
            win._active_task_owner = None
        
        # 下载完成后创建汇总文件夹
        self._create_summary_folders()
        
        # 获取保存目录
        save_dir = self.save_dir.text().strip()
        
        # 点确定后跳转到保存目录
        QMessageBox.information(self, "完成", f"所有图片下载完成！\n\n保存位置：{save_dir}")
        
        # 打开保存目录
        if save_dir and os.path.exists(save_dir):
            os.startfile(save_dir)
    
    def _create_summary_folders(self):
        """创建汇总文件夹：汇总详情图和汇总视频"""
        try:
            save_dir = self.save_dir.text().strip()
            if not save_dir or not os.path.exists(save_dir):
                return
            
            # 获取下载目录下的所有款号文件夹
            product_codes = []
            for item in os.listdir(save_dir):
                item_path = os.path.join(save_dir, item)
                if os.path.isdir(item_path) and item not in ["汇总的详情图", "汇总的视频"]:
                    product_codes.append(item)
            
            if not product_codes:
                return
            
            # 创建汇总详情图文件夹
            detail_summary_dir = os.path.join(save_dir, "汇总的详情图")
            os.makedirs(detail_summary_dir, exist_ok=True)
            
            # 创建汇总视频文件夹
            video_summary_dir = os.path.join(save_dir, "汇总的视频")
            os.makedirs(video_summary_dir, exist_ok=True)
            
            detail_count = 0
            video_count = 0
            
            for code in product_codes:
                code_dir = os.path.join(save_dir, code)
                
                # 复制详情图片
                detail_src = os.path.join(code_dir, "详情图片")
                if os.path.exists(detail_src):
                    detail_dest = os.path.join(detail_summary_dir, code)
                    if not os.path.exists(detail_dest):
                        import shutil
                        shutil.copytree(detail_src, detail_dest)
                        detail_count += 1
                
                # 复制视频
                video_src = os.path.join(code_dir, "视频")
                if os.path.exists(video_src):
                    import shutil
                    for f in os.listdir(video_src):
                        if f.endswith(".mp4"):
                            src_file = os.path.join(video_src, f)
                            dst_file = os.path.join(video_summary_dir, f)
                            # 如果文件名冲突，加序号
                            if os.path.exists(dst_file):
                                base = f.replace(".mp4", "")
                                ext = ".mp4"
                                idx = 1
                                while os.path.exists(dst_file):
                                    dst_file = os.path.join(video_summary_dir, f"{base}_{idx}{ext}")
                                    idx += 1
                            shutil.copy2(src_file, dst_file)
                            video_count += 1
            
            self.log_text.append(f"已创建汇总文件夹:")
            self.log_text.append(f"  - 汇总的详情图: {detail_count} 个款")
            self.log_text.append(f"  - 汇总的视频: {video_count} 个视频")
            
        except Exception as e:
            self.log_text.append(f"创建汇总文件夹失败: {e}")
    
    def select_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)")
        if path:
            self.excel_path.setText(path)
            self.analyze_excel(path)

    def on_excel_dropped(self, path):
        """处理拖拽 Excel 文件到输入框。"""
        if not path or not os.path.exists(path):
            self.log_text.append("拖拽失败: 文件不存在")
            return
        self.log_text.append(f"已拖拽Excel: {os.path.basename(path)}")
        self.analyze_excel(path)
    
    def select_dir(self):
        path = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if path:
            self.save_dir.setText(path)
    
    def analyze_excel(self, path):
        try:
            import pandas as pd
            df = pd.read_excel(path)
            self.rows_label.setText(f"行数: {len(df)}")
            
            link_col = None
            code_col = None
            
            # 精准匹配：必须同时包含"商品"和"链接"
            for c in df.columns:
                col_str = str(c)
                if "商品" in col_str and "链接" in col_str:
                    link_col = c
                    break
            
            # 如果没找到精确匹配，尝试其他方式
            if not link_col:
                for c in df.columns:
                    if "链接" in str(c):
                        link_col = c
                        break
            
            if not link_col:
                for c in df.columns:
                    v = str(df[c].iloc[0]) if len(df) > 0 else ""
                    if "taobao.com" in v:
                        link_col = c
                        break
            
            # 精准匹配：必须包含"唯品款号"
            for c in df.columns:
                if "唯品款号" in str(c):
                    code_col = c
                    break
            
            # 如果没找到，提示用户
            if not link_col:
                self.link_col_label.setText("链接列: ❌ 未找到「商品链接」列")
                self.link_col_label.setStyleSheet("color: #D32F2F; font-weight: bold;")
            else:
                self.link_col_label.setText(f"链接列: {link_col}")
                self.link_col_label.setStyleSheet("color: #2E7D32;")
                
            if not code_col:
                self.code_col_label.setText("款号列: ❌ 未找到「唯品款号」列")
                self.code_col_label.setStyleSheet("color: #D32F2F; font-weight: bold;")
            else:
                self.code_col_label.setText(f"款号列: {code_col}")
                self.code_col_label.setStyleSheet("color: #2E7D32;")
            
            if link_col and code_col:
                self.log_text.append(f"✓ Excel分析完成: 找到「商品链接」和「唯品款号」列，共{len(df)}行")
            else:
                self.log_text.append("⚠ Excel表头缺少必要字段，请检查后重试")
        except Exception as e:
            self.log_text.append(f"读取Excel出错: {e}")
    
    def show_anti_ban_settings(self):
        dialog = AntiBanSettingsDialog(self.config, self)
        dialog.exec()
    
    def _on_captcha_need_slot(self):
        QMessageBox.information(
            self, "人机验证",
            "请在已打开的浏览器窗口中完成验证或滑块，完成后点击确定继续下载。"
        )
        ev = getattr(self, "_captcha_event_holder", None)
        if ev:
            ev.set()
    
    def _make_captcha_wait(self):
        def _wait():
            ev = threading.Event()
            self._captcha_event_holder = ev
            self.captcha_need_signal.emit()
            ev.wait(timeout=600)
        return _wait
    
    def _on_login_launch_done(self, failed: bool):
        if failed and hasattr(self, "login_btn"):
            self.login_btn.setEnabled(True)
    
    # ==================== DownloadPage 业务方法 ====================
    def do_login(self):
        self.login_btn.setEnabled(False)
        self.status_signal.emit("启动中...", "#FF9800")
        if self.login_mode_seller.isChecked():
            _login_mode = "seller"
        elif self.login_mode_qianniu.isChecked():
            _login_mode = "qianniu"
        else:
            _login_mode = "consumer"
        
        def work():
            failed = False
            try:
                anti = load_anti_ban_config()
                proxies = load_proxies()
                proxy = None
                if anti.get("use_proxy"):
                    if proxies:
                        proxy = random.choice(proxies)
                        self.log_signal.emit(f"[防封] 使用代理: {proxy}")
                    else:
                        self.log_signal.emit("[防封] 已勾选代理但未找到 proxies.txt 中的有效行，将直连")
                use_random_ua = anti.get("random_ua", True)
                ua = get_chrome_ua(use_random_ua)
                if not use_random_ua:
                    self.log_signal.emit("[防封] 使用固定 Chrome UA")
                
                self.dl_pw = PW(
                    log=lambda m: self.log_signal.emit(m),
                    captcha_wait=self._make_captcha_wait(),
                )
                self.dl_pw.start()
                
                self.dl_pw.launch(headless=False, proxy=proxy, custom_ua=ua)
                self.dl_pw.open_login(mode=_login_mode)
                
                self.log_signal.emit(">>> 请在弹出的浏览器中登录淘宝 <<<")
                self.status_signal.emit("等待登录...", "#FF9800")
            except Exception as e:
                failed = True
                self.log_signal.emit(f"出错: {e}")
                self.status_signal.emit("未登录", "#D32F2F")
            finally:
                self.login_launch_done.emit(failed)
        
        threading.Thread(target=work, daemon=True).start()

    def confirm_login(self):
        self.status_label.setText("登录成功")
        self.status_label.setStyleSheet("color: #00AA00; font-weight: bold;")
        self.login_btn.setEnabled(True)
        self.confirm_login_btn.hide()
        self.log_signal.emit("登录成功！")

    def start_download(self):
        win = self.window()
        active_owner = getattr(win, "_active_task_owner", None)
        if active_owner and active_owner != "download":
            QMessageBox.warning(self, "提示", "当前已有其它大任务在运行，请等待完成后再开始下载。")
            return
        excel = self.excel_path.text().strip()
        save_dir = self.save_dir.text().strip()
        
        if not excel or not os.path.exists(excel):
            QMessageBox.warning(self, "错误", "请选择Excel文件")
            return
        if not save_dir:
            QMessageBox.warning(self, "错误", "请选择保存目录")
            return
        if not self.dl_pw:
            QMessageBox.warning(self, "提示", "请先登录淘宝")
            return
        
        link_col = self.link_col_label.text().replace("链接列: ", "").replace(" (选择Excel后自动识别)", "")
        code_col = self.code_col_label.text().replace("款号列: ", "").replace(" (选择Excel后自动识别)", "")
        
        if "未找到" in link_col or "未找到" in code_col:
            QMessageBox.warning(self, "错误", "列检测失败，请重新选择Excel")
            return
        
        self.config["last_excel"] = excel
        self.config["last_dir"] = save_dir
        self.config["resume_enabled"] = self.resume_check.isChecked()
        if self.tb_matting_each.isChecked():
            matting_choice = "each"
        elif self.tb_matting_batch.isChecked():
            matting_choice = "batch"
        else:
            matting_choice = "none"
        self.config["taobao_matting_mode"] = matting_choice
        # 抠图统一改为本地 U2NET，清理旧配置项
        self.config.pop("cutout_engine", None)
        self.config.pop("cutout_cloud_api_key", None)
        self.config.pop("cutout_cloud_secret_key", None)
        self.config["taobao_matting_parallel"] = bool(self.tb_matting_parallel.isChecked())
        self.config["taobao_matting_workers"] = int(self.tb_matting_workers.value())
        self.config["taobao_detail_one_pass"] = self.tb_detail_one_pass.isChecked()
        self.config["taobao_fast_page_scroll"] = self.tb_fast_page_scroll.isChecked()
        save_config(self.config)
        
        # 主线程预读界面，避免在下载线程中直接访问 QCheckBox 等（易闪退）
        _dl_fast = self.tb_fast_page_scroll.isChecked()
        _dl_resume = self.resume_check.isChecked()
        _dl_detail1 = self.tb_detail_one_pass.isChecked()
        _dl_mtp = bool(self.tb_matting_parallel.isChecked())
        _dl_mtw = int(self.tb_matting_workers.value())
        
        self.dl_running = True
        if hasattr(win, "_active_task_owner"):
            win._active_task_owner = "download"
        else:
            win._active_task_owner = "download"
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        
        def work():
            import logging
            import pandas as pd
            try:
                anti_ban = load_anti_ban_config()
                human_scroll = anti_ban.get("human_scroll", True) and (not _dl_fast)
                auto_retry = anti_ban.get("auto_retry", True)
                max_retries = int(anti_ban.get("max_retries", 3))
                batch_size = max(1, int(anti_ban.get("batch_size", 5)))
                batch_rest_min = float(anti_ban.get("batch_rest_min", 30))
                batch_rest_max = float(anti_ban.get("batch_rest_max", 60))
                retries = max(1, min(max_retries, 10)) if auto_retry else 1
            
                df = pd.read_excel(excel)
                total = len(df)
                ok = fail = skip = 0
                batch_ok_count = 0
            
                history = DownloadHistory(save_dir) if _dl_resume else None
                if history and history.history:
                    self.log_signal.emit(f"发现历史记录: {len(history.history)} 个已下载商品")
            
                tasks = []
                for idx, row in df.iterrows():
                    if not self.dl_running:
                        break
                    link = str(row.get(link_col, "")).strip()
                    code = str(row.get(code_col, "")).strip()
                
                    if not link or link == "nan" or "taobao.com" not in link:
                        skip += 1
                        continue
                    if not code or code == "nan":
                        code = f"item_{idx + 1}"
                    code = clean(code)
                
                    if history and history.is_downloaded(code):
                        self.log_signal.emit(f"[{idx + 1}/{total}] {code} - 已下载，跳过")
                        skip += 1
                        continue
                
                    self.log_signal.emit(f"[{idx + 1}/{total}] {code}")
                    tasks.append((idx, row, link_col, code_col, code))
            
                n_tasks = len(tasks)
                for ti, task in enumerate(tasks):
                    if not self.dl_running:
                        break
                    idx, row, lk, cd, code = task
                    link = str(row.get(lk, "")).strip()
                
                    def attempt_download(idx=idx, code=code, link=link):
                        last_err = None
                        for attempt in range(retries):
                            if not self.dl_running:
                                return ("skip", idx, code, "已停止")
                            try:
                                passes = 1 if _dl_detail1 else 2
                                r = self.dl_pw.download_product(
                                    link, os.path.join(save_dir, code), code,
                                    human_scroll=human_scroll,
                                    matting_mode=matting_choice,
                                    matting_parallel=_dl_mtp,
                                    matting_workers=_dl_mtw,
                                    detail_pump_passes=passes,
                                )
                                if r is None:
                                    return ("need_login", idx, code, None)
                                n = sum(r.values())
                                if n > 0:
                                    if history:
                                        history.mark_downloaded(code, r)
                                    return ("ok", idx, code, r)
                                if attempt < retries - 1:
                                    self.log_signal.emit(f"  无资源，{min(2 ** attempt, 30)} 秒后重试 ({attempt + 2}/{retries})…")
                                    time.sleep(min(2 ** attempt, 30))
                                    continue
                                return ("skip", idx, code, "无资源")
                            except Exception as e:
                                last_err = e
                                err_s = str(e)
                                if attempt < retries - 1:
                                    self.log_signal.emit(f"  出错 (尝试 {attempt + 1}/{retries}): {err_s}，{min(2 ** attempt, 30)} 秒后重试…")
                                    time.sleep(min(2 ** attempt, 30))
                                    continue
                                return ("fail", idx, code, err_s)
                        return ("fail", idx, code, str(last_err) if last_err else "未知错误")
                
                    status, idx, code, detail = attempt_download()
                
                    if status == "ok":
                        self.log_signal.emit(
                            f"  完成: 主{detail['main']} SKU{detail['sku']} 详情{detail['detail']} 视频{detail['video']}"
                        )
                        ok += 1
                        batch_ok_count += 1
                        if (
                            anti_ban.get("enabled", True)
                            and batch_ok_count > 0
                            and batch_ok_count % batch_size == 0
                            and ti + 1 < n_tasks
                        ):
                            rest = random.uniform(batch_rest_min, batch_rest_max)
                            self.log_signal.emit(f"  已连续成功 {batch_size} 个，批量休息 {rest:.0f} 秒…")
                            time.sleep(rest)
                    elif status == "fail":
                        self.log_signal.emit(f"  出错: {detail}")
                        fail += 1
                    elif status == "need_login":
                        self.log_signal.emit("  登录已失效，请重新登录后再继续下载")
                        break
                    else:
                        self.log_signal.emit(f"[{idx + 1}/{total}] {code} - {detail}")
                        skip += 1
                
                    done = ok + fail + skip
                    pct = done / total * 100 if total > 0 else 0
                    self.progress_signal.emit((pct, f"{done}/{total} ({pct:.0f}%)"))
                
                    if anti_ban.get("enabled", True):
                        delay = random.uniform(anti_ban.get("min_delay", 3), anti_ban.get("max_delay", 8))
                        time.sleep(delay)
            
                self.log_signal.emit("=" * 50)
                self.log_signal.emit(f"下载完成! 成功:{ok} 失败:{fail} 跳过:{skip}")
                if matting_choice == "batch" and self.dl_running:
                    self.log_signal.emit(">>> 开始对保存目录下各款的 SKU 图统一抠透明 …")
                    try:
                        st = batch_cutout_skus_under_root(save_dir, lambda s: self.log_signal.emit(s))
                        self.log_signal.emit(
                            f"统一抠图结束: 新生成 {st.get('ok', 0)}，跳过 {st.get('skip', 0)}，失败 {st.get('fail', 0)}"
                        )
                    except Exception as e:
                        self.log_signal.emit(f"统一抠图出错: {e}")
            except Exception as e:
                logging.getLogger("toolbox.download").exception("淘宝下载任务")
                self.log_signal.emit(f"下载任务异常（已记日志）: {e}")
            finally:
                self.done_signal.emit()
        
        threading.Thread(target=work, daemon=True).start()

    def stop_download(self):
        self.dl_running = False
        if self.dl_pw:
            try:
                self.dl_pw.close()
            except Exception:
                pass
        self.log_text.append("正在停止...")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        win = self.window()
        if hasattr(win, "_active_task_owner") and getattr(win, "_active_task_owner", None) == "download":
            win._active_task_owner = None


class PDFPage(QWidget):
    # 定义信号用于线程安全地更新UI
    done_signal = Signal(tuple)  # PDF转图片完成
    done_signal2 = Signal(tuple)  # 批量生成吊牌完成（旧，保留兼容）
    pdf_log_signal = Signal(str)
    pdf_tag_done_signal = Signal(object)  # {"mode":"batch"|"single", "ok":...}
    i2p_done_signal = Signal(object)  # {"ok":bool,"msg":str,"path":""}
    permit_done_signal = Signal(object)  # {"ok":bool,"msg":str,"path":""}
    
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self._last_permit_path = ""
        self.init_ui()
        self.done_signal.connect(self.on_pdf2img_done)
        self.done_signal2.connect(self.on_pdf_batch_done)
        self.pdf_log_signal.connect(lambda s: self.pdf_log.append(s))
        self.pdf_tag_done_signal.connect(self._on_pdf_tag_done)
        self.i2p_done_signal.connect(self._on_i2p_done)
        self.permit_done_signal.connect(self._on_permit_done)
    
    def on_pdf2img_done(self, data):
        output_dir = data[0]
        self.p2i_btn.setEnabled(True)
        self.p2i_btn.setText("▶️ PDF转图片")
        if output_dir == "error":
            QMessageBox.warning(self, "错误", f"转换失败: {data[1]}")
        else:
            QMessageBox.information(self, "✅ 转换完成", f"PDF已转换为图片\n保存位置: {output_dir}")
            try:
                os.startfile(output_dir)
            except Exception:
                pass
    
    def on_pdf_batch_done(self, data):
        success, skip, output_dir = data
        self.tag_btn.setEnabled(True)
        QMessageBox.information(self, "✅ 批量生成完成", f"成功:{success} 跳过:{skip}\n保存位置: {output_dir}")
        try:
            os.startfile(output_dir)
        except Exception:
            pass
    
    def _on_pdf_tag_done(self, data):
        self.tag_btn.setEnabled(True)
        self.tag_btn.setText("▶️ 开始生成")
        mode = data.get("mode", "batch")
        if mode == "single":
            if data.get("ok"):
                QMessageBox.information(self, "✅ 完成", data.get("msg", ""))
                p = data.get("path", "")
                if p and os.path.isdir(os.path.dirname(p)):
                    try:
                        os.startfile(os.path.dirname(p))
                    except Exception:
                        pass
            else:
                QMessageBox.warning(self, "失败", data.get("msg", "未知错误"))
        else:
            if data.get("err"):
                self.pdf_log_signal.emit("批量结束：失败或表头不对，请查看上方日志。")
                return
            try:
                if data.get("out"):
                    os.startfile(data["out"])
            except Exception:
                pass
            self.pdf_log_signal.emit(
                f"批量完成：成功 {data.get('ok', 0)} 跳过 {data.get('skip', 0)} 目录 {data.get('out', '')}"
            )
    
    def select_input(self, line_edit):
        """选择目录"""
        path = QFileDialog.getExistingDirectory(self, "选择PDF文件夹")
        if path:
            line_edit.setText(path)
    
    def select_dir(self, line_edit):
        """选择目录"""
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if path:
            line_edit.setText(path)
    
    def _pick_pdf_folder(self):
        path = QFileDialog.getExistingDirectory(self, "选择PDF文件夹")
        if path:
            self.pdf_folder.setText(path)
            self._update_pdf_match_count()
    
    def _update_pdf_match_count(self):
        """根据当前文件夹与表格统计可生成吊牌的行数（不弹窗）。"""
        lab = getattr(self, "pdf_match_label", None)
        if lab is None:
            return
        pdf_folder = self.pdf_folder.text().strip()
        excel_path = self.pdf_excel.text().strip()
        if not pdf_folder or not os.path.isdir(pdf_folder):
            lab.setText("已匹配 — 款（请先选择 PDF 文件夹）")
            return
        if not excel_path or not os.path.isfile(excel_path):
            lab.setText("已匹配 — 款（请先选择 Excel 表格）")
            return
        try:
            import pandas as pd
            df = pd.read_excel(excel_path)
            columns = df.columns.tolist()
            code_col = next((c for c in columns if "唯品款号" in str(c).strip()), None)
            if not code_col:
                lab.setText("已匹配 0 款（表中无「唯品款号」列）")
                return
            pdf_files = {f[:-4].lower(): f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")}
            matched = 0
            for _, row in df.iterrows():
                code = str(row[code_col]).strip()
                if not code or code.lower() in ("nan", "none"):
                    continue
                if PDFPage.build_tag_jobs(code, pdf_files):
                    matched += 1
            lab.setText(f"已匹配 {matched} 款")
        except Exception as e:
            lab.setText(f"已匹配 — 款（统计失败，见日志）")
            self.pdf_log_signal.emit(f"匹配统计: {e}")
    
    def _pick_images_multi(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择图片（可多选）",
            "",
            "图片 (*.jpg *.jpeg *.png *.webp *.bmp)",
        )
        if paths:
            self.i2p_files.setText(";".join(paths))
            if not self.i2p_output.text().strip():
                self.i2p_output.setText(os.path.dirname(paths[0]))
    
    def _on_i2p_done(self, data):
        self.i2p_btn.setEnabled(True)
        self.i2p_btn.setText("▶️ 图片转 PDF")
        if data.get("ok"):
            self.pdf_log_signal.emit(data.get("msg", "图片转 PDF 完成"))
            p = data.get("path", "")
            if p and os.path.isfile(p):
                try:
                    os.startfile(os.path.dirname(p))
                except Exception:
                    pass
            elif p and os.path.isdir(p):
                try:
                    os.startfile(p)
                except Exception:
                    pass
        else:
            QMessageBox.warning(self, "图片转 PDF", data.get("msg", "失败"))
    
    def start_img2pdf(self):
        raw = self.i2p_files.text().strip()
        out_dir = self.i2p_output.text().strip()
        paths = [p.strip() for p in raw.replace("|", ";").split(";") if p.strip()]
        paths = [p for p in paths if os.path.isfile(p)]
        if not paths:
            QMessageBox.warning(self, "提示", "请选择至少一张图片")
            return
        if not out_dir:
            QMessageBox.warning(self, "提示", "请选择输出目录")
            return
        os.makedirs(out_dir, exist_ok=True)
        merge = self.i2p_merge_radio.isChecked()
        self.i2p_btn.setEnabled(False)
        self.i2p_btn.setText("处理中…")
        
        def work():
            data = {"ok": False, "msg": "未知", "path": ""}
            try:
                if merge:
                    base = os.path.splitext(os.path.basename(paths[0]))[0]
                    out_pdf = os.path.join(out_dir, f"{base}_合并.pdf")
                    pdf_helpers.images_to_pdf_files(paths, out_pdf)
                    data = {"ok": True, "msg": f"已生成: {out_pdf}", "path": out_pdf}
                else:
                    for p in paths:
                        stem = os.path.splitext(os.path.basename(p))[0]
                        one_out = os.path.join(out_dir, f"{stem}.pdf")
                        pdf_helpers.images_to_pdf_files([p], one_out)
                    data = {
                        "ok": True,
                        "msg": f"已生成 {len(paths)} 个 PDF 到 {out_dir}",
                        "path": out_dir,
                    }
            except Exception as e:
                data = {"ok": False, "msg": str(e), "path": ""}
            finally:
                self.i2p_done_signal.emit(data)
        
        threading.Thread(target=work, daemon=True).start()
    
    def select_excel_pdf(self, line_edit):
        """选择Excel文件"""
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel文件 (*.xlsx *.xls)")
        if path:
            line_edit.setText(path)
            self.analyze_pdf_excel(path)
    
    def analyze_pdf_excel(self, path):
        """分析 Excel：需含「唯品款号」「吊牌价」列（与批量逻辑一致）。"""
        try:
            import pandas as pd
            df = pd.read_excel(path)
            self.pdf_log_signal.emit(f"读取Excel: {len(df)} 行")
            code_col = price_col = None
            for c in df.columns:
                if "唯品款号" in str(c):
                    code_col = c
                    break
            for c in df.columns:
                col_str = str(c)
                if "吊牌" in col_str and "价" in col_str:
                    price_col = c
                    break
            if not price_col:
                for c in df.columns:
                    if "价" in str(c):
                        price_col = c
                        break
            if code_col and price_col:
                self.pdf_log_signal.emit("✅ 表头校验通过：唯品款号 + 吊牌价（已用页面「已匹配」统计代替弹窗）")
            else:
                missing = []
                if not code_col:
                    missing.append("唯品款号")
                if not price_col:
                    missing.append("吊牌价")
                self.pdf_log_signal.emit(f"⚠ Excel 缺少列: {', '.join(missing)}")
            self._update_pdf_match_count()
        except Exception as e:
            self.pdf_log_signal.emit(f"读取 Excel 失败: {e}")
            self._update_pdf_match_count()
    
    def init_ui(self):
        root = QVBoxLayout()
        root.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        inner = QWidget()
        layout = QVBoxLayout(inner)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("PDF工具箱")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1565C0;")
        layout.addWidget(title)
        
        desc = QLabel("PDF与图片互转，批量生成吊牌图")
        desc.setFont(QFont("Microsoft YaHei", 10))
        desc.setStyleSheet("color: #666;")
        layout.addWidget(desc)
        
        # 吊牌图生成（合并批量/单款）
        tag_group = QGroupBox("🏷️ 生成吊牌图")
        tag_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        tag_layout = QVBoxLayout()
        tag_layout.setSpacing(10)
        
        # 模式选择
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("模式选择:"))
        self.tag_mode_batch = QRadioButton("📦 批量模式（依赖 Excel）")
        self.tag_mode_single = QRadioButton("📌 单款模式（手动输入）")
        self.tag_mode_batch.setChecked(True)
        self.tag_mode_batch.toggled.connect(self._toggle_tag_mode)
        mode_row.addWidget(self.tag_mode_batch)
        mode_row.addWidget(self.tag_mode_single)
        mode_row.addStretch()
        tag_layout.addLayout(mode_row)
        
        # 批量模式容器
        self.batch_container = QWidget()
        batch_lay = QVBoxLayout()
        batch_lay.setContentsMargins(0, 0, 0, 0)
        batch_lay.setSpacing(10)
        
        tag_info = QLabel("📌 说明：文件夹内同名的PDF合格证和水洗唛会自动匹配，生成吊牌图")
        tag_info.setFont(QFont("Microsoft YaHei", 9))
        tag_info.setStyleSheet("color: #666;")
        batch_lay.addWidget(tag_info)
        
        pdf_row = QHBoxLayout()
        pdf_row.addWidget(QLabel("PDF文件夹:"))
        self.pdf_folder = QLineEdit()
        enable_path_drop(self.pdf_folder, mode="dir", on_accept=lambda *_: self._update_pdf_match_count())
        self.pdf_folder.setPlaceholderText("存放合格证和水洗唛PDF的文件夹...")
        self.pdf_folder.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        self.pdf_folder.editingFinished.connect(self._update_pdf_match_count)
        pdf_row.addWidget(self.pdf_folder)
        btn_pdf = QPushButton("浏览")
        btn_pdf.setMinimumWidth(96)
        btn_pdf.setMinimumHeight(34)
        btn_pdf.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_pdf.clicked.connect(self._pick_pdf_folder)
        pdf_row.addWidget(btn_pdf)
        batch_lay.addLayout(pdf_row)
        
        excel_row = QHBoxLayout()
        excel_row.addWidget(QLabel("款号价格表:"))
        self.pdf_excel = QLineEdit()
        enable_path_drop(self.pdf_excel, mode="file", extensions=(".xlsx", ".xls"), on_accept=self.analyze_pdf_excel)
        self.pdf_excel.setPlaceholderText("含「唯品款号」「吊牌价」列的Excel（可选）...")
        self.pdf_excel.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        excel_row.addWidget(self.pdf_excel)
        btn_excel = QPushButton("浏览")
        btn_excel.setMinimumWidth(96)
        btn_excel.setMinimumHeight(34)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_excel.clicked.connect(lambda: self.select_excel_pdf(self.pdf_excel))
        excel_row.addWidget(btn_excel)
        batch_lay.addLayout(excel_row)
        
        self.pdf_match_label = QLabel("已匹配 — 款（请选择 PDF 文件夹与表格后自动统计）")
        self.pdf_match_label.setFont(QFont("Microsoft YaHei", 9))
        self.pdf_match_label.setStyleSheet("color: #2E7D32; font-weight: bold;")
        batch_lay.addWidget(self.pdf_match_label)
        
        pdf_excel_tip = QLabel("📌 Excel表头必须有：①「唯品款号」列 ②「吊牌价」列（精确匹配，自动抓取价格）")
        pdf_excel_tip.setFont(QFont("Microsoft YaHei", 8))
        pdf_excel_tip.setStyleSheet("color: #D32F2F; padding-left: 20px;")
        batch_lay.addWidget(pdf_excel_tip)
        
        self.batch_container.setLayout(batch_lay)
        tag_layout.addWidget(self.batch_container)
        
        # 单款模式容器
        self.single_container = QWidget()
        single_lay = QVBoxLayout()
        single_lay.setContentsMargins(0, 0, 0, 0)
        single_lay.setSpacing(10)
        
        single_input_style = """
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus {
                border: 2px solid #1565C0;
                background-color: white;
            }
        """
        single_browse_style = """
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
        """
        sr1 = QHBoxLayout()
        sr1.addWidget(QLabel("选择PDF:"))
        self.tag_single_pdf_edit = QLineEdit()
        self.tag_single_pdf_edit.setStyleSheet(single_input_style)
        self.tag_single_pdf_edit.setPlaceholderText("选择款号.pdf（合格证），自动匹配款号-1.pdf（水洗唛）")
        enable_path_drop(self.tag_single_pdf_edit, mode="file", extensions=(".pdf",))
        sr1.addWidget(self.tag_single_pdf_edit)
        b1 = QPushButton("浏览")
        b1.setStyleSheet(single_browse_style)
        b1.setMinimumWidth(96)
        b1.setMinimumHeight(34)
        b1.clicked.connect(lambda: self._pick_pdf_file(self.tag_single_pdf_edit))
        sr1.addWidget(b1)
        single_lay.addLayout(sr1)
        sr2 = QHBoxLayout()
        sr2.addWidget(QLabel("款号:"))
        self.tag_single_code_edit = QLineEdit()
        self.tag_single_code_edit.setStyleSheet(single_input_style)
        self.tag_single_code_edit.setPlaceholderText("留空则从PDF文件名自动提取")
        sr2.addWidget(self.tag_single_code_edit)
        sr2.addWidget(QLabel("吊牌价:"))
        self.tag_single_price_edit = QLineEdit()
        self.tag_single_price_edit.setStyleSheet(single_input_style)
        self.tag_single_price_edit.setPlaceholderText("默认 399")
        self.tag_single_price_edit.setText("399")
        sr2.addWidget(self.tag_single_price_edit)
        single_lay.addLayout(sr2)
        
        self.single_container.setLayout(single_lay)
        self.single_container.setVisible(False)
        tag_layout.addWidget(self.single_container)
        
        # 输出目录（共享）
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("输出目录:"))
        self.pdf_output = QLineEdit()
        enable_path_drop(self.pdf_output, mode="dir")
        self.pdf_output.setPlaceholderText("生成的吊牌图保存位置...")
        self.pdf_output.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        out_row.addWidget(self.pdf_output)
        btn_out = QPushButton("浏览")
        btn_out.setMinimumWidth(96)
        btn_out.setMinimumHeight(34)
        btn_out.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_out.clicked.connect(lambda: self.select_dir(self.pdf_output))
        out_row.addWidget(btn_out)
        tag_layout.addLayout(out_row)
        
        # 统一开始按钮
        self.tag_btn = QPushButton("▶️ 开始生成")
        self.tag_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.tag_btn.setMinimumHeight(42)
        self.tag_btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed)
        self.tag_btn.setStyleSheet("""
            QPushButton {
                background-color: #43A047;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
        """)
        self.tag_btn.clicked.connect(self.start_generate_tag)
        tag_layout.addWidget(self.tag_btn)
        
        tag_group.setLayout(tag_layout)
        layout.addWidget(tag_group)
        
        # 日志区域
        pdf_log_group = QGroupBox("📝 处理日志")
        pdf_log_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        pdf_log_layout = QVBoxLayout()
        
        self.pdf_log = QTextEdit()
        self.pdf_log.setReadOnly(True)
        self.pdf_log.setMaximumHeight(100)
        self.pdf_log.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                background-color: #1E1E1E;
                color: #00FF00;
                font-family: Consolas;
                font-size: 9px;
            }
        """)
        pdf_log_layout.addWidget(self.pdf_log)
        
        pdf_log_group.setLayout(pdf_log_layout)
        
        tag_group.setLayout(tag_layout)
        layout.addWidget(tag_group)
        
        # 日志区域放到底部
        layout.addWidget(pdf_log_group)
        
        # PDF权限清理
        permit_group = QGroupBox("🔓 PDF权限清理与编辑")
        permit_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        permit_layout = QVBoxLayout()
        permit_layout.setSpacing(10)
        
        permit_info = QLabel("📌 自动去掉PDF的编辑限制（禁止修改/加密等），另存为可直接编辑的PDF副本")
        permit_info.setFont(QFont("Microsoft YaHei", 9))
        permit_info.setStyleSheet("color: #666;")
        permit_layout.addWidget(permit_info)
        
        _inp_style = """
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus {
                border: 2px solid #1565C0;
                background-color: white;
            }
        """
        _browse_style = """
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
        """
        
        # PDF文件选择
        pdf_row = QHBoxLayout()
        pdf_row.addWidget(QLabel("PDF文件:"))
        self.permit_file = QLineEdit()
        self.permit_file.setStyleSheet(_inp_style)
        self.permit_file.setPlaceholderText("选择要清理权限的PDF文件...")
        enable_path_drop(self.permit_file, mode="file", extensions=(".pdf",))
        pdf_row.addWidget(self.permit_file)
        btn_permit_file = QPushButton("浏览")
        btn_permit_file.setMinimumWidth(96)
        btn_permit_file.setMinimumHeight(34)
        btn_permit_file.setStyleSheet(_browse_style)
        btn_permit_file.clicked.connect(self._pick_permit_file)
        pdf_row.addWidget(btn_permit_file)
        permit_layout.addLayout(pdf_row)
        
        # 输出目录选择
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("输出目录:"))
        self.permit_out = QLineEdit()
        self.permit_out.setStyleSheet(_inp_style)
        self.permit_out.setPlaceholderText("默认与原PDF同目录...")
        enable_path_drop(self.permit_out, mode="dir")
        out_row.addWidget(self.permit_out)
        btn_permit_out = QPushButton("浏览")
        btn_permit_out.setMinimumWidth(96)
        btn_permit_out.setMinimumHeight(34)
        btn_permit_out.setStyleSheet(_browse_style)
        btn_permit_out.clicked.connect(lambda: self.select_dir(self.permit_out))
        out_row.addWidget(btn_permit_out)
        permit_layout.addLayout(out_row)
        
        # 保存模式选择
        mode_row = QHBoxLayout()
        self.permit_overwrite = QRadioButton("覆盖原文件（自动备份为 .bak.pdf）")
        self.permit_overwrite.setFont(QFont("Microsoft YaHei", 9))
        self.permit_overwrite.setStyleSheet("color: #333;")
        self.permit_saveas = QRadioButton("另存为（文件名后加 _可编辑）")
        self.permit_saveas.setFont(QFont("Microsoft YaHei", 9))
        self.permit_saveas.setStyleSheet("color: #333;")
        self.permit_saveas.setChecked(True)
        mode_row.addWidget(self.permit_overwrite)
        mode_row.addWidget(self.permit_saveas)
        mode_row.addStretch()
        permit_layout.addLayout(mode_row)
        
        # 按钮行
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        
        self.permit_btn = QPushButton("▶️ 去掉编辑限制")
        self.permit_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.permit_btn.setMinimumHeight(42)
        self.permit_btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed)
        self.permit_btn.setStyleSheet("""
            QPushButton {
                background-color: #43A047;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #388E3C;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.permit_btn.clicked.connect(self.start_remove_protection)
        btn_row.addWidget(self.permit_btn)
        
        self.permit_open_btn = QPushButton("▶️ 用系统默认程序打开")
        self.permit_open_btn.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.permit_open_btn.setMinimumHeight(42)
        self.permit_open_btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed)
        self.permit_open_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.permit_open_btn.clicked.connect(self.open_permit_file_external)
        self.permit_open_btn.setEnabled(False)
        btn_row.addWidget(self.permit_open_btn)
        
        permit_layout.addLayout(btn_row)
        
        permit_tip = QLabel("⚠️ 仅支持「双击即可直接打开、不要求输入密码」的PDF。若设置了打开密码请先用其它软件处理。")
        permit_tip.setFont(QFont("Microsoft YaHei", 8))
        permit_tip.setStyleSheet("color: #D32F2F; padding-left: 4px;")
        permit_layout.addWidget(permit_tip)
        
        permit_group.setLayout(permit_layout)
        layout.addWidget(permit_group)
        
        # PDF转图片
        conv_group = QGroupBox("🔄 PDF/图片互转")
        conv_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        conv_layout = QVBoxLayout()
        conv_layout.setSpacing(10)
        
        # PDF转图片
        p2i_row = QHBoxLayout()
        p2i_row.addWidget(QLabel("PDF文件:"))
        self.p2i_file = QLineEdit()
        enable_path_drop(self.p2i_file, mode="file", extensions=(".pdf",))
        self.p2i_file.setPlaceholderText("选择要转换的PDF文件...")
        self.p2i_file.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        p2i_row.addWidget(self.p2i_file)
        btn_p2i = QPushButton("浏览")
        btn_p2i.setMinimumWidth(96)
        btn_p2i.setMinimumHeight(34)
        btn_p2i.setStyleSheet("""
            QPushButton {
                background-color: #C8E6C9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_p2i.clicked.connect(self.select_pdf)
        p2i_row.addWidget(btn_p2i)
        conv_layout.addLayout(p2i_row)
        
        p2i_out_row = QHBoxLayout()
        p2i_out_row.addWidget(QLabel("输出目录:"))
        self.p2i_output = QLineEdit()
        enable_path_drop(self.p2i_output, mode="dir")
        self.p2i_output.setPlaceholderText("选择输出目录...")
        self.p2i_output.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        p2i_out_row.addWidget(self.p2i_output)
        btn_p2i_out = QPushButton("浏览")
        btn_p2i_out.setMinimumWidth(96)
        btn_p2i_out.setMinimumHeight(34)
        btn_p2i_out.setStyleSheet("""
            QPushButton {
                background-color: #C8E6C9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_p2i_out.clicked.connect(lambda: self.select_dir(self.p2i_output))
        p2i_out_row.addWidget(btn_p2i_out)
        conv_layout.addLayout(p2i_out_row)
        
        self.p2i_btn = QPushButton("▶️ PDF转图片")
        self.p2i_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.p2i_btn.setMinimumHeight(42)
        self.p2i_btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed)
        self.p2i_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
        """)
        self.p2i_btn.clicked.connect(self.start_pdf2img)
        conv_layout.addWidget(self.p2i_btn)
        
        i2p_row = QHBoxLayout()
        i2p_row.addWidget(QLabel("图片文件:"))
        self.i2p_files = QLineEdit()
        self.i2p_files.setPlaceholderText("可浏览多选，或拖拽多张图到此处（分号分隔）…")
        enable_path_drop(self.i2p_files, mode="file", extensions=(".jpg", ".jpeg", ".png", ".webp", ".bmp"), multi=True)
        self.i2p_files.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        i2p_row.addWidget(self.i2p_files)
        bi2p = QPushButton("浏览")
        bi2p.setMinimumWidth(96)
        bi2p.setMinimumHeight(34)
        bi2p.clicked.connect(self._pick_images_multi)
        i2p_row.addWidget(bi2p)
        conv_layout.addLayout(i2p_row)
        
        i2p_mode_row = QHBoxLayout()
        self.i2p_merge_radio = QRadioButton("多图合并为 1 个 PDF")
        self.i2p_each_radio = QRadioButton("每张图各生成 1 个 PDF")
        self.i2p_merge_radio.setChecked(True)
        i2p_mode_row.addWidget(self.i2p_merge_radio)
        i2p_mode_row.addWidget(self.i2p_each_radio)
        conv_layout.addLayout(i2p_mode_row)
        
        i2p_out_row = QHBoxLayout()
        i2p_out_row.addWidget(QLabel("输出目录:"))
        self.i2p_output = QLineEdit()
        enable_path_drop(self.i2p_output, mode="dir")
        self.i2p_output.setPlaceholderText("图片转 PDF 保存目录…")
        self.i2p_output.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        i2p_out_row.addWidget(self.i2p_output)
        bo2 = QPushButton("浏览")
        bo2.setMinimumWidth(96)
        bo2.setMinimumHeight(34)
        bo2.clicked.connect(lambda: self.select_dir(self.i2p_output))
        i2p_out_row.addWidget(bo2)
        conv_layout.addLayout(i2p_out_row)
        
        self.i2p_btn = QPushButton("▶️ 图片转 PDF")
        self.i2p_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.i2p_btn.setMinimumHeight(42)
        self.i2p_btn.setSizePolicy(QSizePolicy.Policy.MinimumExpanding, QSizePolicy.Policy.Fixed)
        self.i2p_btn.setStyleSheet("""
            QPushButton {
                background-color: #00838F;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
        """)
        self.i2p_btn.clicked.connect(self.start_img2pdf)
        conv_layout.addWidget(self.i2p_btn)
        
        conv_group.setLayout(conv_layout)
        layout.addWidget(conv_group)
        
        self._update_pdf_match_count()
        for w in (
            self.pdf_folder,
            self.pdf_excel,
            self.pdf_output,
            self.tag_single_pdf_edit,
            self.tag_single_code_edit,
            self.tag_single_price_edit,
            self.permit_file,
            self.permit_out,
            self.p2i_file,
            self.p2i_output,
            self.i2p_files,
            self.i2p_output,
        ):
            w.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            w.setMinimumWidth(140)
        inner.setMinimumWidth(520)
        scroll.setWidget(inner)
        root.addWidget(scroll, 1)
        self.setLayout(root)
    
    def start_pdf2img(self):
        pdf_path = self.p2i_file.text().strip()
        output_dir = self.p2i_output.text().strip()
        
        if not pdf_path or not os.path.exists(pdf_path):
            QMessageBox.warning(self, "提示", "请选择PDF文件")
            return
        
        if not output_dir:
            QMessageBox.warning(self, "提示", "请选择输出目录")
            return
        
        os.makedirs(output_dir, exist_ok=True)
        self.p2i_btn.setEnabled(False)
        self.p2i_btn.setText("🔄 处理中...")
        
        def work():
            out_t = None
            err = None
            try:
                import fitz
                doc = fitz.open(pdf_path)
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    img_path = os.path.join(output_dir, f"page_{page_num + 1}.png")
                    pix.save(img_path)
                doc.close()
                out_t = (output_dir,)
            except Exception as e:
                err = e
            finally:
                if err is not None:
                    self.done_signal.emit(("error", str(err)))
                elif out_t is not None:
                    self.done_signal.emit(out_t)
                else:
                    self.done_signal.emit(("error", "未生成图片输出"))
        
        threading.Thread(target=work, daemon=True).start()
    
    def select_pdf(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择PDF", "", "PDF文件 (*.pdf)")
        if path:
            self.p2i_file.setText(path)
            self.p2i_output.setText(os.path.dirname(path))
    
    def _pick_pdf_file(self, line_edit):
        path, _ = QFileDialog.getOpenFileName(self, "选择PDF", "", "PDF文件 (*.pdf)")
        if path:
            line_edit.setText(path)
    
    def _toggle_tag_mode(self, checked):
        is_batch = self.tag_mode_batch.isChecked()
        self.batch_container.setVisible(is_batch)
        self.single_container.setVisible(not is_batch)
    
    def start_generate_tag(self):
        if self.tag_mode_batch.isChecked():
            self._start_batch_mode()
        else:
            self._start_single_mode()
    
    def _start_batch_mode(self):
        pdf_folder = self.pdf_folder.text().strip()
        excel_path = self.pdf_excel.text().strip()
        output_dir = self.pdf_output.text().strip()
        
        if not pdf_folder or not output_dir:
            QMessageBox.warning(self, "提示", "请选择PDF文件夹和输出目录")
            return
        
        if not excel_path:
            QMessageBox.warning(self, "提示", "请选择Excel文件")
            return
        
        self.tag_btn.setEnabled(False)
        self.tag_btn.setText("🔄 处理中...")
        
        def work():
            payload = {"ok": 0, "skip": 0, "err": True, "out": output_dir}
            try:
                import pandas as pd
                df = pd.read_excel(excel_path)
                columns = df.columns.tolist()
                code_col = next((c for c in columns if "唯品款号" in str(c).strip()), None)
                price_col = next((c for c in columns if "吊牌价" in str(c).strip()), None)
                if not price_col:
                    for c in columns:
                        col_str = str(c)
                        if "吊牌" in col_str and "价" in col_str:
                            price_col = c
                            break
                if not price_col:
                    for c in columns:
                        if "价" in str(c):
                            price_col = c
                            break
                if not code_col or not price_col:
                    self.pdf_log_signal.emit("错误: 未找到「唯品款号」或「吊牌价」列")
                    return
                
                pdf_files = {f[:-4].lower(): f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")}
                success = skip = 0
                for _, row in df.iterrows():
                    code = str(row[code_col]).strip()
                    price = row[price_col]
                    if not code or code in ("nan", "NaN", ""):
                        continue
                    jobs = PDFPage.build_tag_jobs(code, pdf_files)
                    if not jobs:
                        self.pdf_log_signal.emit(f"跳过 {code}: 未找到可配对的 PDF")
                        skip += 1
                        continue
                    try:
                        out_folder = os.path.join(output_dir, code)
                        os.makedirs(out_folder, exist_ok=True)
                        price_s = str(price) if price is not None else "0"
                        for main_pdf, var_pdf, out_base in jobs:
                            out_path = os.path.join(out_folder, f"{out_base}.jpg")
                            pdf_helpers.compose_tag826_jpg(
                                os.path.join(pdf_folder, main_pdf),
                                os.path.join(pdf_folder, var_pdf),
                                price_s,
                                out_path,
                            )
                        self.pdf_log_signal.emit(f"✓ {code}（{len(jobs)} 张）")
                        success += 1
                    except Exception as e:
                        self.pdf_log_signal.emit(f"错误 {code}: {e}")
                        skip += 1
                self.pdf_log_signal.emit(f"\n完成! 成功:{success} 跳过:{skip}")
                payload = {"ok": success, "skip": skip, "err": False, "out": output_dir}
            except Exception as e:
                self.pdf_log_signal.emit(f"错误: {e}")
            finally:
                payload["mode"] = "batch"
                self.pdf_tag_done_signal.emit(payload)
        
        threading.Thread(target=work, daemon=True).start()
    
    @staticmethod
    def stem_matches_row_code(stem_lower: str, code: str) -> bool:
        return pdf_helpers.stem_matches_row_code(stem_lower, code)
    
    @staticmethod
    def list_cert_stems_for_row(stems_lower: set[str]) -> list[str]:
        return pdf_helpers.list_cert_stems_for_row(stems_lower)
    
    @staticmethod
    def _strip_code_prefix(orig_stem: str, code: str) -> str:
        return pdf_helpers.strip_code_prefix(orig_stem, code)
    
    @staticmethod
    def build_tag_jobs(code: str, pdf_lower_to_orig: dict) -> list[tuple[str, str, str]]:
        return pdf_helpers.build_tag_jobs(code, pdf_lower_to_orig)
    
    def _start_single_mode(self):
        pdf_path = self.tag_single_pdf_edit.text().strip()
        code = self.tag_single_code_edit.text().strip()
        price = self.tag_single_price_edit.text().strip() or "399"
        out_root = self.pdf_output.text().strip()
        if not pdf_path or not out_root:
            QMessageBox.warning(self, "提示", "请选择PDF文件和输出目录")
            return
        if not os.path.isfile(pdf_path):
            QMessageBox.warning(self, "提示", "PDF 文件不存在")
            return
        
        # 自动推导合格证和水洗唛
        folder = os.path.dirname(pdf_path)
        base = os.path.basename(pdf_path)
        stem, ext = os.path.splitext(base)
        if stem.endswith("-1"):
            main_stem = stem[:-2]
            main_p = os.path.join(folder, main_stem + ext)
            var_p = pdf_path
        else:
            main_p = pdf_path
            var_p = os.path.join(folder, stem + "-1" + ext)
        
        if not code:
            code = stem.replace("-1", "")
        
        if not os.path.isfile(main_p) or not os.path.isfile(var_p):
            QMessageBox.warning(self, "提示", f"缺失PDF文件，请确保 {os.path.basename(main_p)} 和 {os.path.basename(var_p)} 均存在")
            return
        
        self.tag_btn.setEnabled(False)
        self.tag_btn.setText("🔄 处理中...")
        out_folder = os.path.join(out_root, code)
        out_path = os.path.join(out_folder, "826.jpg")
        
        def work():
            pl = {"ok": False, "msg": "未知", "path": "", "mode": "single"}
            try:
                pdf_helpers.compose_tag826_jpg(main_p, var_p, price, out_path)
                pl = {"ok": True, "msg": f"已生成:\n{out_path}", "path": out_path, "mode": "single"}
            except Exception as e:
                pl = {"ok": False, "msg": str(e), "path": "", "mode": "single"}
            finally:
                self.pdf_tag_done_signal.emit(pl)
        
        threading.Thread(target=work, daemon=True).start()

    # ==================== PDF 权限清理方法 ====================

    def _pick_permit_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择PDF文件", "", "PDF文件 (*.pdf)")
        if path:
            self.permit_file.setText(path)
            if not self.permit_out.text().strip():
                self.permit_out.setText(os.path.dirname(path))

    def _on_permit_done(self, data):
        self.permit_btn.setEnabled(True)
        self.permit_btn.setText("▶️ 去掉编辑限制")
        if data.get("ok"):
            out_path = data.get("path", "")
            self._last_permit_path = out_path
            self.permit_open_btn.setEnabled(True)
            self.pdf_log_signal.emit(data.get("msg", ""))
            QMessageBox.information(self, "✅ 完成", data.get("msg", ""))
            if out_path and os.path.isfile(out_path):
                try:
                    os.startfile(os.path.dirname(out_path))
                except Exception:
                    pass
        else:
            self._last_permit_path = ""
            self.permit_open_btn.setEnabled(False)
            QMessageBox.warning(self, "失败", data.get("msg", "处理出错"))

    def start_remove_protection(self):
        pdf_path = self.permit_file.text().strip()
        out_dir = self.permit_out.text().strip()

        if not pdf_path or not os.path.isfile(pdf_path):
            QMessageBox.warning(self, "提示", "请选择有效的PDF文件")
            return

        overwrite = self.permit_overwrite.isChecked()
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        src_dir = os.path.dirname(pdf_path)

        if not out_dir:
            out_dir = src_dir

        os.makedirs(out_dir, exist_ok=True)

        if overwrite:
            bak_path = os.path.join(src_dir, f"{base}.bak.pdf")
            out_path = pdf_path
            try:
                import shutil
                shutil.copy2(pdf_path, bak_path)
                self.pdf_log_signal.emit(f"已备份原文件: {os.path.basename(bak_path)}")
            except Exception as e:
                QMessageBox.warning(self, "备份失败", f"无法备份原文件:\n{e}")
                return
        else:
            out_path = os.path.join(out_dir, f"{base}_可编辑.pdf")

        self.permit_btn.setEnabled(False)
        self.permit_btn.setText("🔄 处理中...")
        self._last_permit_path = ""
        self.permit_open_btn.setEnabled(False)

        def work():
            pl = {"ok": False, "msg": "未知", "path": ""}
            try:
                doc, meta = pdf_edit_core.open_pdf_editable_path(pdf_path)
                pdf_edit_core.save_document(doc, out_path)
                doc.close()
                msg = f"已生成可编辑PDF: {out_path}"
                if meta.get("security_stripped"):
                    msg += "\n（已自动去掉编辑限制）"
                else:
                    msg += "\n（该PDF原本无限制，已保存副本）"
                pl = {"ok": True, "msg": msg, "path": out_path}
            except ValueError as e:
                pl = {"ok": False, "msg": str(e), "path": ""}
            except Exception as e:
                pl = {"ok": False, "msg": f"处理失败: {e}", "path": ""}
            finally:
                self.permit_done_signal.emit(pl)

        threading.Thread(target=work, daemon=True).start()

    def open_permit_file_external(self):
        path = getattr(self, "_last_permit_path", "")
        if not path or not os.path.isfile(path):
            path = self.permit_file.text().strip()
            if not path or not os.path.isfile(path):
                QMessageBox.warning(self, "提示", "没有可打开的文件，请先去掉编辑限制")
                return
        try:
            os.startfile(path)
            self.pdf_log_signal.emit(f"已打开: {path}")
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"无法打开文件:\n{e}")


class CutoutPage(QWidget):
    progress_signal = Signal(int)
    log_signal = Signal(str)
    done_signal = Signal(tuple)
    
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.current_files = []
        self.current_idx = 0
        self._should_stop = False
        
        self.progress_signal.connect(self.on_progress)
        self.log_signal.connect(self.on_log)
        self.done_signal.connect(self.on_done)
        
        self.init_ui()
    
    def init_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        page_scroll = QScrollArea()
        page_scroll.setWidgetResizable(True)
        page_scroll.setFrameShape(QFrame.NoFrame)
        page_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        outer.addWidget(page_scroll, 1)
        content = QWidget()
        root = QVBoxLayout(content)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)
        
        title = QLabel("批量抠图工具")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1565C0;")
        root.addWidget(title)
        
        desc = QLabel("使用AI智能抠图，去除背景保留主体，透明PNG输出")
        desc.setFont(QFont("Microsoft YaHei", 10))
        desc.setStyleSheet("color: #666;")
        root.addWidget(desc)
        
        status_layout = QHBoxLayout()
        self.rembg_status_label = QLabel("")
        self.u2net_status_label = QLabel("")
        status_layout.addWidget(self.rembg_status_label)
        status_layout.addWidget(self.u2net_status_label)
        status_layout.addStretch()
        root.addLayout(status_layout)
        self._refresh_cutout_runtime_status()
        
        params_group = QGroupBox("⚙️ 参数设置")
        params_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        params_wrap = QVBoxLayout()
        params_layout = QHBoxLayout()
        params_layout.addWidget(QLabel("输出宽度:"))
        self.cutout_width = QSpinBox()
        self.cutout_width.setRange(100, 2000)
        self.cutout_width.setValue(self.config.get("cutout_width", 800))
        self.cutout_width.setSuffix(" px")
        self.cutout_width.setFixedWidth(120)
        params_layout.addWidget(self.cutout_width)
        params_layout.addWidget(QLabel("  输出体积上限:"))
        self.cutout_maxkb = QSpinBox()
        self.cutout_maxkb.setRange(10, 2000)
        self.cutout_maxkb.setValue(self.config.get("cutout_maxkb", 600))
        self.cutout_maxkb.setSuffix(" KB")
        self.cutout_maxkb.setFixedWidth(120)
        params_layout.addWidget(self.cutout_maxkb)
        params_layout.addWidget(QLabel("  边缘扩展:"))
        self.cutout_dilate = QSpinBox()
        self.cutout_dilate.setRange(0, 50)
        self.cutout_dilate.setValue(self.config.get("cutout_dilate", 1))
        self.cutout_dilate.setSuffix(" px")
        self.cutout_dilate.setToolTip("先扩张再收缩遮罩，将被误裁的部分（袖子等）重新接回。1-3轻微补偿，4+强力连接（0=关闭）")
        self.cutout_dilate.setFixedWidth(80)
        params_layout.addWidget(self.cutout_dilate)
        params_layout.addStretch()
        params_wrap.addLayout(params_layout)
        vol_hint = QLabel(
            "说明：保存的文件体积会「小于」上面填的 KB，并尽量接近该值，让透明图尽量清晰。"
        )
        vol_hint.setWordWrap(True)
        vol_hint.setFont(QFont("Microsoft YaHei", 8))
        vol_hint.setStyleSheet("color: #666;")
        params_wrap.addWidget(vol_hint)
        params_group.setLayout(params_wrap)
        
        io_group = QGroupBox("📁 输入输出设置")
        io_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        io_layout = QVBoxLayout()
        io_layout.setSpacing(10)
        io_note = QLabel("💡 支持两种模式：\n  • 选择文件夹：处理文件夹内所有图片\n  • 选择文件：处理选中的图片文件")
        io_note.setFont(QFont("Microsoft YaHei", 9))
        io_note.setStyleSheet("color: #666;")
        io_layout.addWidget(io_note)
        input_row = QHBoxLayout()
        input_row.addWidget(QLabel("输入:"))
        self.cutout_input = QLineEdit()
        enable_path_drop(
            self.cutout_input,
            mode="file_or_dir",
            extensions=(".jpg", ".jpeg", ".png", ".bmp", ".webp"),
            multi=True,
        )
        self.cutout_input.setText(self.config.get("cutout_dir", ""))
        self.cutout_input.setPlaceholderText("选择图片文件夹或文件...")
        self.cutout_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.cutout_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        input_row.addWidget(self.cutout_input)
        btn_input = QPushButton("浏览")
        btn_input.setMinimumWidth(88)
        btn_input.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_input.clicked.connect(self.select_input)
        input_row.addWidget(btn_input)
        io_layout.addLayout(input_row)
        output_row = QHBoxLayout()
        output_row.addWidget(QLabel("输出:"))
        self.cutout_output = QLineEdit()
        enable_path_drop(self.cutout_output, mode="dir")
        self.cutout_output.setText(self.config.get("cutout_output", ""))
        self.cutout_output.setPlaceholderText("选择抠图后的保存位置...")
        self.cutout_output.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.cutout_output.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
        """)
        output_row.addWidget(self.cutout_output)
        btn_output = QPushButton("浏览")
        btn_output.setMinimumWidth(88)
        btn_output.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 8px 12px;
                border-radius: 4px;
            }
        """)
        btn_output.clicked.connect(lambda: self.select_dir(self.cutout_output))
        output_row.addWidget(btn_output)
        io_layout.addLayout(output_row)
        io_group.setLayout(io_layout)
        
        self.cutout_btn = QPushButton("▶️ 开始批量抠图")
        self.cutout_btn.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.cutout_btn.setMinimumHeight(40)
        self.cutout_btn.setStyleSheet("""
            QPushButton {
                background-color: #43A047;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 4px;
            }
        """)
        self.cutout_btn.clicked.connect(self.start_cutout)

        filter_group = QGroupBox("📂 文件夹过滤")
        filter_group.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        filter_layout = QVBoxLayout()
        filter_layout.setSpacing(8)
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("文件夹名:"))
        self.cutout_folder_filter = QLineEdit(self.config.get("cutout_folder_filter", "SKU"))
        self.cutout_folder_filter.setPlaceholderText("留空则处理所有图片")
        self.cutout_folder_filter.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 8px 12px;
                background-color: #FAFAFA;
            }
            QLineEdit:focus {
                border: 2px solid #1565C0;
                background-color: white;
            }
        """)
        filter_row.addWidget(self.cutout_folder_filter)
        filter_layout.addLayout(filter_row)
        filter_note = QLabel("只处理指定名称的子文件夹内的图片（如 SKU），留空则处理所有图片")
        filter_note.setFont(QFont("Microsoft YaHei", 8))
        filter_note.setStyleSheet("color: #666;")
        filter_layout.addWidget(filter_note)
        filter_group.setLayout(filter_layout)
        
        left_col = QVBoxLayout()
        left_col.addWidget(params_group)
        left_col.addWidget(io_group)
        left_col.addWidget(filter_group)
        cut_hint = QLabel("输出说明：抠图结果与输入文件同名，扩展名为 .png（透明底）。")
        cut_hint.setFont(QFont("Microsoft YaHei", 8))
        cut_hint.setStyleSheet("color: #666;")
        cut_hint.setWordWrap(True)
        left_col.addWidget(cut_hint)
        btn_row = QHBoxLayout()
        btn_row.addWidget(self.cutout_btn)
        self.cutout_stop_btn = QPushButton("⏹ 停止")
        self.cutout_stop_btn.setFont(QFont("Microsoft YaHei", 10))
        self.cutout_stop_btn.setMinimumHeight(40)
        self.cutout_stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #D32F2F;
                color: white;
                border: none;
                padding: 12px 20px;
                border-radius: 4px;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
        """)
        self.cutout_stop_btn.setEnabled(False)
        self.cutout_stop_btn.clicked.connect(self.stop_cutout)
        btn_row.addWidget(self.cutout_stop_btn)
        left_col.addLayout(btn_row)
        left_w = QWidget()
        left_w.setLayout(left_col)
        root.addWidget(left_w, 0)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("处理进度 %p%")
        self.progress_bar.setStyleSheet("""
            QProgressBar { border: 1px solid #E0E0E0; border-radius: 4px; text-align: center; height: 22px; }
            QProgressBar::chunk { background-color: #4CAF50; }
        """)
        root.addWidget(self.progress_bar)
        
        log_group = QGroupBox("📝 处理日志")
        log_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(64)
        self.log_text.setMaximumHeight(88)
        self.log_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                background-color: #1E1E1E;
                color: #00FF00;
                font-family: Consolas;
                font-size: 9px;
            }
        """)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        root.addWidget(log_group, 0)
        page_scroll.setWidget(content)
    
    def log_message(self, msg):
        # 页面初始化早期可能先触发依赖检查，此时日志框尚未创建
        if getattr(self, "log_text", None) is not None:
            self.log_text.append(msg)
    
    # 信号处理函数（在主线程中执行，安全）
    def on_progress(self, value):
        if getattr(self, "progress_bar", None) is not None:
            self.progress_bar.setValue(int(value))
    
    def on_log(self, msg):
        if getattr(self, "log_text", None) is not None:
            self.log_text.append(msg)
    
    def stop_cutout(self):
        self._should_stop = True
        self.cutout_stop_btn.setEnabled(False)
        self.log_message("正在停止抠图...")

    def on_done(self, data):
        success, total, output_dir = data
        # 恢复按钮
        if self.cutout_btn:
            self.cutout_btn.setEnabled(True)
            self.cutout_btn.setText("▶️ 开始批量抠图")
        if getattr(self, "cutout_stop_btn", None):
            self.cutout_stop_btn.setEnabled(False)
        if getattr(self, "progress_bar", None) is not None:
            self.progress_bar.setVisible(False)
            self.progress_bar.setValue(0)
        stopped = self._should_stop
        self._should_stop = False
        # 在主线程中显示消息框（安全）
        if stopped:
            QMessageBox.information(self, "已停止", f"抠图已停止，已完成 {success}/{total} 张图片\n保存位置: {output_dir}")
        else:
            QMessageBox.information(self, "✅ 抠图完成", f"已处理 {success}/{total} 张图片\n保存位置: {output_dir}")
        try:
            os.startfile(output_dir)
        except Exception as e:
            print(f"打开目录失败: {e}")
        win = self.window()
        if hasattr(win, "_active_task_owner") and getattr(win, "_active_task_owner", None) == "cutout":
            win._active_task_owner = None
    
    def select_dir(self, line_edit):
        path = QFileDialog.getExistingDirectory(self, "选择目录")
        if path:
            line_edit.setText(path)
    
    def select_input(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("选择输入类型")
        msg.setText("请选择输入类型：")
        msg.setIcon(QMessageBox.Question)
        
        folder_btn = msg.addButton("📁 文件夹（批量处理）", QMessageBox.ActionRole)
        file_btn = msg.addButton("🖼️ 图片文件（多选）", QMessageBox.ActionRole)
        cancel_btn = msg.addButton("取消", QMessageBox.RejectRole)
        
        msg.exec()
        
        if msg.clickedButton() == folder_btn:
            path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
            if path:
                self.cutout_input.setText(path)
                self.current_files = []
                self.current_idx = 0
        elif msg.clickedButton() == file_btn:
            files, _ = QFileDialog.getOpenFileNames(self, "选择图片文件", "", "图片文件 (*.jpg *.jpeg *.png *.bmp *.webp)")
            if files:
                self.cutout_input.setText(";".join(files))
                self.current_files = files
                self.current_idx = 0
    
    def _resolve_cutout_files(self, custom_out_dir=None):
        input_path = self.cutout_input.text().strip()
        exts = (".jpg", ".jpeg", ".png", ".bmp", ".webp")
        folder_filter = self.cutout_folder_filter.text().strip()
        self._file_output_map = {}
        
        if os.path.isdir(input_path) and folder_filter:
            files = []
            for root, dirs, _ in os.walk(input_path):
                if os.path.basename(root) == folder_filter:
                    sku_parent = os.path.dirname(root)
                    parent_name = os.path.basename(sku_parent)
                    out_folder_name = f"{parent_name}_抠图"
                    if custom_out_dir:
                        out_dir = os.path.join(custom_out_dir, out_folder_name)
                    else:
                        out_dir = os.path.join(os.path.dirname(sku_parent), out_folder_name)
                    for f in os.listdir(root):
                        if f.lower().endswith(exts):
                            fpath = os.path.join(root, f)
                            files.append(fpath)
                            stem = os.path.splitext(f)[0]
                            self._file_output_map[fpath] = os.path.join(out_dir, f"{stem}.png")
            files.sort()
        elif os.path.isdir(input_path):
            files = sorted(
                [os.path.join(input_path, f) for f in os.listdir(input_path) if f.lower().endswith(exts)]
            )
        else:
            files = [p.strip() for p in input_path.split(";") if p.strip()]
            files = [p for p in files if os.path.isfile(p) and p.lower().endswith(exts)]
            if folder_filter:
                files = [f for f in files if os.path.basename(os.path.dirname(f)) == folder_filter]
        self.current_files = files
        self.current_idx = 0
        mode_info = "递归" if (os.path.isdir(input_path) and folder_filter) else "普通"
        self.log_message(f"已加载 {len(files)} 张图片（{mode_info}模式，过滤：{folder_filter or '无'}）")
        return files

    def _find_u2net_model_path(self):
        candidates = [
            os.path.join(APP_ROOT, "u2net.onnx"),
            os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "u2net.onnx"),
            os.path.join(os.path.expanduser("~"), ".u2net", "u2net.onnx"),
        ]
        for path in candidates:
            if os.path.isfile(path):
                return path
        return ""

    def _probe_rembg_runtime(self):
        """实时检测 rembg 是否能在当前程序 Python 环境里导入。"""
        try:
            from rembg import remove as _remove  # noqa: F401
            return True, ""
        except Exception as e:
            return False, str(e)

    def _refresh_cutout_runtime_status(self):
        """每次进入/开始前都实时检查，避免安装后仍显示未安装。"""
        ok, err = self._probe_rembg_runtime()
        self.rembg_ok = ok
        if ok:
            self.rembg_status_label.setText("✓ rembg 已就绪（当前程序环境）")
            self.rembg_status_label.setStyleSheet("color: #2E7D32; font-weight: bold;")
        else:
            self.rembg_status_label.setText("✗ rembg 不可用（请装到当前程序环境）")
            self.rembg_status_label.setStyleSheet("color: #D32F2F; font-weight: bold;")
            if err:
                self.log_message(f"[抠图依赖] rembg 导入失败：{err}")
            self.log_message(f"[抠图依赖] 当前程序 Python：{sys.executable}")

        if self._find_u2net_model_path():
            self.u2net_status_label.setText("✓ u2net.onnx 本地模型已就绪")
            self.u2net_status_label.setStyleSheet("color: #2E7D32; font-weight: bold;")
        else:
            self.u2net_status_label.setText("✗ 未找到 u2net.onnx（请放到项目根目录）")
            self.u2net_status_label.setStyleSheet("color: #D32F2F; font-weight: bold;")

    def _run_rembg_cutout(self, img):
        try:
            from rembg import remove
        except Exception as e:
            raise RuntimeError(
                f"当前程序环境无法导入 rembg。\n程序Python：{sys.executable}\n导入错误：{e}"
            ) from e
        model_path = self._find_u2net_model_path()
        if not model_path:
            raise RuntimeError("未找到 u2net.onnx，请将模型文件放到项目根目录")
        u2net_home = os.environ.get("U2NET_HOME")
        os.environ["U2NET_HOME"] = os.path.dirname(model_path)
        try:
            return remove(img), "本地模型(U2NET.ONNX)"
        finally:
            if u2net_home:
                os.environ["U2NET_HOME"] = u2net_home
            else:
                os.environ.pop("U2NET_HOME", None)

    def _expand_alpha_mask(self, img, pixels):
        if pixels <= 0:
            return img
        from PIL import Image, ImageFilter
        r, g, b, a = img.split()
        size = pixels * 2 + 1
        a = a.filter(ImageFilter.MaxFilter(size))
        if pixels > 2:
            erode_r = max(1, pixels // 3)
            a = a.filter(ImageFilter.MinFilter(erode_r * 2 + 1))
        return Image.merge("RGBA", (r, g, b, a))

    def start_cutout(self):
        self._refresh_cutout_runtime_status()
        win = self.window()
        active_owner = getattr(win, "_active_task_owner", None)
        if active_owner and active_owner != "cutout":
            QMessageBox.warning(self, "提示", "当前已有其它大任务在运行，请等待完成后再开始抠图。")
            return
        input_path = self.cutout_input.text().strip()
        output_dir = self.cutout_output.text().strip()
        folder_filter = self.cutout_folder_filter.text().strip()
        is_recursive = os.path.isdir(input_path) and bool(folder_filter)

        ok, err = self._probe_rembg_runtime()
        self.rembg_ok = ok
        if not ok:
            err_text = f"\n导入错误：{err}" if err else ""
            QMessageBox.warning(
                self,
                "错误",
                "当前程序环境里 rembg 不可用。\n"
                f"请在命令行执行：\n\"{sys.executable}\" -m pip install rembg onnxruntime\n"
                f"程序Python：{sys.executable}{err_text}",
            )
            return
        if not self._find_u2net_model_path():
            QMessageBox.warning(self, "错误", "未找到 u2net.onnx，请将模型文件放到项目根目录")
            return
        
        if not input_path:
            QMessageBox.warning(self, "提示", "请选择输入")
            return
        if not is_recursive and not output_dir:
            QMessageBox.warning(self, "提示", "请选择输出目录")
            return
        
        if not is_recursive:
            os.makedirs(output_dir, exist_ok=True)
        
        self._should_stop = False
        self.cutout_btn.setEnabled(False)
        self.cutout_stop_btn.setEnabled(True)
        self.cutout_btn.setText("🔄 处理中...")
        if hasattr(win, "_active_task_owner"):
            win._active_task_owner = "cutout"
        else:
            win._active_task_owner = "cutout"
        
        # 获取图片列表
        files = self._resolve_cutout_files(output_dir if is_recursive else None)
        if not files:
            QMessageBox.warning(self, "提示", "没有找到可处理的图片，请检查输入")
            self.cutout_btn.setEnabled(True)
            self.cutout_btn.setText("▶️ 开始批量抠图")
            if hasattr(win, "_active_task_owner") and getattr(win, "_active_task_owner", None) == "cutout":
                win._active_task_owner = None
            return
        
        total = len(files)
        _co_w = int(self.cutout_width.value())
        _co_mkb = int(self.cutout_maxkb.value())
        self.log_message(f"开始处理 {total} 张图片...")
        self.progress_bar.setMaximum(max(total, 1))
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        
        def work():
            import logging
            logger = logging.getLogger(__name__)
            target_width, max_kb = _co_w, _co_mkb
            success = 0
            failed = 0
            try:
                from PIL import Image
                
                mode_label = "递归多文件夹" if is_recursive else "普通"
                logger.info(f"[{mode_label}] 开始处理 {total} 张图片，宽度:{target_width}px，最大{max_kb}KB")
                self.log_signal.emit(f"[{mode_label}] 开始处理 {total} 张图片（目标宽度:{target_width}px，最大{max_kb}KB）...")
                
                dilate_px = int(self.cutout_dilate.value())
                for i, f in enumerate(files):
                    if self._should_stop:
                        self.log_signal.emit("用户已停止抠图")
                        break
                    try:
                        logger.debug(f"处理第 {i+1} 张: {f}")
                        img = Image.open(f)
                        output, used_engine = self._run_rembg_cutout(img)
                        
                        # 2. 扩展遮罩边缘，改善浅色衣物被误抠
                        from PIL import ImageFilter
                        output = self._expand_alpha_mask(output, dilate_px)
                        output = output.filter(ImageFilter.SMOOTH_MORE)
                        
                        # 3. 调整大小（按目标宽度）
                        if output.width != target_width:
                            ratio = target_width / output.width
                            new_height = int(output.height * ratio)
                            output = output.resize((target_width, new_height), Image.LANCZOS)
                        
                        # 4. 获取输出路径：递归模式用映射，普通模式用输出目录
                        fmap = getattr(self, '_file_output_map', {})
                        if f in fmap:
                            out_path = fmap[f]
                            os.makedirs(os.path.dirname(out_path), exist_ok=True)
                        else:
                            name = os.path.splitext(os.path.basename(f))[0]
                            out_path = os.path.join(output_dir, f"{name}.png")
                        
                        # 5. 压缩到目标体积
                        output, actual_size = compress_image_to_size_v2(output, max_kb)
                        # 保存为PNG保持透明，参数与压缩测算一致
                        output.save(out_path, "PNG", optimize=True, compress_level=9)
                        
                        success += 1
                        # 使用信号更新进度（线程安全）
                        self.progress_signal.emit(i + 1)
                        self.log_signal.emit(f"完成: {os.path.basename(f)}（{used_engine}）")
                        logger.debug(f"完成: {out_path} ({actual_size/1024:.1f}KB)")
                    except Exception as e:
                        failed += 1
                        logger.error(f"处理失败: {os.path.basename(f)} - {e}")
                        self.log_signal.emit(f"处理失败: {os.path.basename(f)} - {e}")
                
                logger.info(f"完成! 成功:{success} 失败:{failed}")
                self.log_signal.emit(f"✓ 完成！成功处理 {success}/{total} 张图片")
            except Exception as e:
                logger.error(f"抠图异常: {e}")
                self.log_signal.emit(f"错误: {e}")
                success = 0
            finally:
                report_dir = output_dir if output_dir else input_path
                self.done_signal.emit((success, total, report_dir))
        
        threading.Thread(target=work, daemon=True).start()
    
    def save_settings(self):
        self.config["cutout_width"] = self.cutout_width.value()
        self.config["cutout_maxkb"] = self.cutout_maxkb.value()
        self.config["cutout_dilate"] = self.cutout_dilate.value()
        self.config["cutout_dir"] = self.cutout_input.text()
        self.config["cutout_output"] = self.cutout_output.text()
        self.config["cutout_folder_filter"] = self.cutout_folder_filter.text()
        # 抠图统一仅本地 U2NET，清理历史配置
        self.config.pop("cutout_engine", None)
        self.config.pop("cutout_cloud_api_key", None)
        self.config.pop("cutout_cloud_secret_key", None)


class SizeTablePage(QWidget):
    # 定义信号用于线程安全地更新UI
    done_signal = Signal(tuple)
    log_signal = Signal(str)
    progress_signal = Signal(int, int)  # current, total
    
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        # hybrid：自动/仅本地/仅云端；cloud_only：分发包无本地 Paddle，只走云端（与抠图无关）
        self._ocr_tier = get_ocr_feature_tier()
        legacy_cloud_only = bool(self.config.get("size_ocr_cloud_only", False))
        self.size_ocr_mode = str(self.config.get("size_ocr_mode", "cloud" if legacy_cloud_only else "auto"))
        if self._ocr_tier == "cloud_only":
            self.size_ocr_mode = "cloud"
        if self.size_ocr_mode not in ("auto", "local", "cloud"):
            self.size_ocr_mode = "auto" if self._ocr_tier != "cloud_only" else "cloud"
        self.results = []
        self.current_idx = 0
        self.confirmed_results = []  # 已确认的结果
        
        # 参考数据
        self.category_type_map = {}  # 唯品类目 -> (一级分类, 上装/裤装, 模板类型)
        self.number_map = {}  # (一级分类, 模板类型, 尺码) -> 号型
        self.field_aliases = {}  # 字段别名
        self.exclude_keywords = []
        self.required_fields = {
            "上装": ["肩宽", "胸围", "衣长", "袖长", "腰围"],
            "裤装": ["腰围", "臀围", "裤长", "裤脚围"],
        }
        
        # 尺码勾选相关
        self.size_vars = {}
        self.size_order = ["S", "M", "L", "XL", "2XL", "3XL"]
        
        # 当前识别的OCR数据
        self.ocr_recognized = {}
        
        # 从Excel加载的款号尺码号型映射 {唯品款号: {尺码: 号型}}（尺码键已统一为 2XL 等标准写法）
        self.excel_size_number_map = {}
        # Excel 与尺码图 OCR 不一致记录，供导出提醒用户手改
        self._mismatch_rows = []
        self._mismatch_seen = set()
        # 外部 JSON 尺码别名，与「尺码映射/尺码别名.json」同步
        self._external_size_aliases = {}
        
        # 连接信号
        self.done_signal.connect(self.on_done)
        self.log_signal.connect(self.on_log)
        self.progress_signal.connect(self.on_progress)
        
        self.init_ui()
        
        # 加载参考数据
        self._apply_ocr_feature_tier()
        self.load_ref_data()
        self.init_field_aliases()

    def _apply_ocr_feature_tier(self):
        if getattr(self, "_ocr_tier", "hybrid") != "cloud_only" or not hasattr(
            self, "ocr_mode_combo"
        ):
            return
        self.ocr_mode_combo.blockSignals(True)
        self.ocr_mode_combo.clear()
        self.ocr_mode_combo.addItem("仅云端（本包不含本地OCR，仅PaddleX等）")
        self.ocr_mode_combo.setCurrentIndex(0)
        self.ocr_mode_combo.setEnabled(False)
        self.ocr_mode_combo.setToolTip("本分发包为仅云端档：请在设置里配好 PaddleX/云端 Token。抠图仍为本地 u2net。")
        self.ocr_mode_combo.blockSignals(False)
        self.size_ocr_mode = "cloud"
        self.config["size_ocr_mode"] = "cloud"
        self.check_ocr_status()
    
    def on_done(self, data):
        success, output_dir = data
        self.st_start_btn.setEnabled(True)
        hint = ""
        if self._mismatch_rows:
            hint = f"\n\n已记录 {len(self._mismatch_rows)} 条「表/图尺码不一致」，在「导出全部」时会自动一并导出手改表。"
        QMessageBox.information(
            self, "✅ 尺码表完成", f"已处理 {success} 个文件\n保存位置: {output_dir}{hint}"
        )
        try:
            os.startfile(output_dir)
        except Exception:
            pass
    
    def on_log(self, message):
        """日志信号处理"""
        self.log_text.append(message)
    
    def on_progress(self, current, total):
        """进度信号处理"""
        self.st_status.setText(f"正在处理: {current}/{total}")
    
    def select_input(self, line_edit):
        """选择目录或文件"""
        path = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if path:
            line_edit.setText(path)
    
    def init_ui(self):
        """初始化UI - 滚动区域覆盖整个页面"""
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        self.setLayout(layout)
        
        # 顶部：标题
        layout.addWidget(self._create_header_section())
        
        # 整个内容区域放入滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                width: 10px;
            }
            QScrollBar::handle:vertical {
                background-color: #BDBDBD;
                border-radius: 5px;
            }
        """)
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout()
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(12)
        
        # 新增：配置+结果 并排显示
        config_results_row = QHBoxLayout()
        config_results_row.setSpacing(15)
        config_results_row.addWidget(self._create_config_compact(), 3)  # 左侧：配置（更宽）
        config_results_row.addWidget(self._create_results_compact(), 2)  # 右侧：结果列表
        scroll_layout.addLayout(config_results_row)
        
        # 预览与编辑区域
        scroll_layout.addWidget(self._create_preview_section())
        
        # 日志区域
        scroll_layout.addWidget(self._create_log_section())
        
        scroll_layout.addStretch()
        
        scroll_content.setLayout(scroll_layout)
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area, 1)
        
        # 绑定信号
        self._bind_signals()
    
    def _create_header_section(self):
        """创建顶部标题区域"""
        header_frame = QFrame()
        header_layout = QVBoxLayout()
        header_layout.setSpacing(8)
        
        title = QLabel("批量尺码表录入")
        title.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title.setStyleSheet("color: #1565C0;")
        header_layout.addWidget(title)
        
        desc = QLabel("批量从图片中识别尺码信息，支持人工确认和表格编辑")
        desc.setFont(QFont("Microsoft YaHei", 10))
        desc.setStyleSheet("color: #666;")
        header_layout.addWidget(desc)
        
        header_frame.setLayout(header_layout)
        return header_frame
    
    def _create_config_compact(self):
        """创建紧凑的配置面板（不占用太多垂直空间）"""
        config_frame = QFrame()
        config_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 6px;
                padding: 8px;
            }
        """)
        
        main_layout = QHBoxLayout()
        main_layout.setSpacing(15)
        
        # 左侧：输入配置
        left_layout = QVBoxLayout()
        left_layout.setSpacing(6)
        
        # Excel文件行
        excel_row = QHBoxLayout()
        excel_row.addWidget(QLabel("Excel文件:"))
        self.st_excel = QLineEdit()
        enable_path_drop(self.st_excel, mode="file", extensions=(".xlsx", ".xls"), on_accept=self.analyze_size_excel)
        self.st_excel.setPlaceholderText("必填：用于读取款号和唯品类目（灰字说明逻辑）")
        self.st_excel.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #FAFAFA;
            }
        """)
        excel_row.addWidget(self.st_excel)
        
        btn_excel = QPushButton("浏览")
        btn_excel.setFixedWidth(60)
        btn_excel.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 6px 10px;
                border-radius: 4px;
            }
        """)
        btn_excel.clicked.connect(lambda: self.select_excel_size(self.st_excel))
        excel_row.addWidget(btn_excel)
        left_layout.addLayout(excel_row)

        # 图片文件夹行
        folder_row = QHBoxLayout()
        folder_row.addWidget(QLabel("图片文件夹:"))
        self.st_folder = QLineEdit()
        enable_path_drop(self.st_folder, mode="dir")
        self.st_folder.setPlaceholderText("必填：包含款号图片的文件夹（灰字说明逻辑）")
        self.st_folder.setStyleSheet("""
            QLineEdit {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #FAFAFA;
            }
        """)
        folder_row.addWidget(self.st_folder)
        
        btn_folder = QPushButton("浏览")
        btn_folder.setFixedWidth(60)
        btn_folder.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 6px 10px;
                border-radius: 4px;
            }
        """)
        btn_folder.clicked.connect(lambda: self.select_input(self.st_folder))
        folder_row.addWidget(btn_folder)
        left_layout.addLayout(folder_row)

        # 导出目录（放在输入路径下面，颜色区分）
        export_row = QHBoxLayout()
        export_label = QLabel("导出目录:")
        export_label.setStyleSheet("color: #EF6C00; font-weight: bold;")
        export_row.addWidget(export_label)
        self.st_export_dir = QLineEdit()
        enable_path_drop(self.st_export_dir, mode="dir")
        self.st_export_dir.setPlaceholderText("选填：为空时自动用Excel目录；无Excel则用图片目录")
        self.st_export_dir.setText(str(self.config.get("size_export_dir", "") or ""))
        self.st_export_dir.setStyleSheet("""
            QLineEdit {
                border: 1px solid #FFB74D;
                border-radius: 4px;
                padding: 6px 10px;
                background-color: #FFF8E1;
            }
        """)
        export_row.addWidget(self.st_export_dir)
        btn_export_dir = QPushButton("浏览")
        btn_export_dir.setFixedWidth(60)
        btn_export_dir.setStyleSheet("""
            QPushButton {
                background-color: #FFF3E0;
                color: #EF6C00;
                border: 1px solid #EF6C00;
                padding: 6px 10px;
                border-radius: 4px;
            }
        """)
        btn_export_dir.clicked.connect(lambda: self.select_dir(self.st_export_dir))
        export_row.addWidget(btn_export_dir)
        left_layout.addLayout(export_row)
        
        # 提示
        st_excel_tip = QLabel("📌 Excel表头必须有：①「唯品款号」列 ②「唯品类目」列")
        st_excel_tip.setFont(QFont("Microsoft YaHei", 8))
        st_excel_tip.setStyleSheet("color: #D32F2F;")
        left_layout.addWidget(st_excel_tip)
        
        # OCR状态行
        ocr_status_row = QHBoxLayout()
        status_icon = QLabel("🔧 OCR:")
        status_icon.setFont(QFont("Microsoft YaHei", 9))
        ocr_status_row.addWidget(status_icon)
        
        self.ocr_status = QLabel("检测中...")
        self.ocr_status.setFont(QFont("Microsoft YaHei", 9))
        ocr_status_row.addWidget(self.ocr_status)
        
        tutorial_btn = QPushButton("📖 教程")
        tutorial_btn.setFont(QFont("Microsoft YaHei", 8))
        tutorial_btn.setFixedWidth(50)
        tutorial_btn.setStyleSheet("""
            QPushButton {
                background-color: #E3F2FD;
                color: #1565C0;
                border: 1px solid #1565C0;
                padding: 2px 6px;
                border-radius: 3px;
            }
        """)
        tutorial_btn.clicked.connect(self.show_ocr_tutorial)
        ocr_status_row.addWidget(tutorial_btn)
        
        set_btn = QPushButton("⚙")
        set_btn.setFont(QFont("Microsoft YaHei", 9))
        set_btn.setFixedWidth(25)
        set_btn.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border: none;
                padding: 2px 6px;
                border-radius: 3px;
            }
        """)
        set_btn.clicked.connect(self.show_ocr_settings)
        ocr_status_row.addWidget(set_btn)
        
        ocr_status_row.addStretch()
        left_layout.addLayout(ocr_status_row)

        # OCR模式选择
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("识别模式:"))
        self.ocr_mode_combo = QComboBox()
        self.ocr_mode_combo.addItems(["自动（本地优先）", "仅本地模型", "仅云端（PaddleX）"])
        mode_index = {"auto": 0, "local": 1, "cloud": 2}.get(self.size_ocr_mode, 0)
        self.ocr_mode_combo.setCurrentIndex(mode_index)
        self.ocr_mode_combo.currentIndexChanged.connect(self.on_ocr_mode_changed)
        mode_row.addWidget(self.ocr_mode_combo, 1)
        left_layout.addLayout(mode_row)
        
        # 操作按钮行
        btn_row = QHBoxLayout()
        self.st_start_btn = QPushButton("🔍 开始批量识别")
        self.st_start_btn.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.st_start_btn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
        """)
        btn_row.addWidget(self.st_start_btn)
        
        self.st_export_all_btn = QPushButton("📦 导出全部")
        self.st_export_all_btn.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.st_export_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #00796B;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
        """)
        btn_row.addWidget(self.st_export_all_btn)
        
        btn_row.addStretch()
        self.st_status = QLabel("就绪")
        self.st_status.setFont(QFont("Microsoft YaHei", 9))
        self.st_status.setStyleSheet("color: #666;")
        btn_row.addWidget(self.st_status)
        
        left_layout.addLayout(btn_row)
        
        main_layout.addLayout(left_layout, 1)
        
        config_frame.setLayout(main_layout)
        
        # 初始化OCR状态
        self.check_ocr_status()
        
        return config_frame
    
    def _create_preview_section(self):
        """创建预览与编辑区域"""
        preview_group = QGroupBox("🖼️ 预览与编辑")
        preview_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        preview_layout = QHBoxLayout()
        preview_layout.setSpacing(10)
        
        # 左侧：图片预览（更大预览区域）
        img_frame = QFrame()
        img_frame.setStyleSheet("""
            QFrame {
                background-color: #F5F5F5;
                border: 1px solid #E0E0E0;
                border-radius: 4px;
            }
        """)
        img_layout = QVBoxLayout()
        img_layout.setContentsMargins(5, 5, 5, 5)
        
        img_title = QLabel("📷 当前图片")
        img_title.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
        img_title.setStyleSheet("color: #1565C0;")
        img_layout.addWidget(img_title)
        
        self.preview_label = QLabel("无预览图片")
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setStyleSheet("""
            QLabel {
                background-color: #FAFAFA;
                border: 1px dashed #CCC;
                color: #999;
                font-size: 14px;
            }
        """)
        # 设置最小宽度确保完整显示
        self.preview_label.setMinimumWidth(280)
        self.preview_label.setMinimumHeight(300)
        self.preview_label.setScaledContents(False)
        img_layout.addWidget(self.preview_label)
        
        # 导航按钮
        nav_row = QHBoxLayout()
        self.prev_btn = QPushButton("◀")
        self.prev_btn.setFixedWidth(40)
        self.prev_btn.setFont(QFont("Microsoft YaHei", 9))
        self.prev_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                padding: 5px 10px;
                border-radius: 4px;
            }
        """)
        nav_row.addWidget(self.prev_btn)
        
        self.nav_label = QLabel("0/0")
        self.nav_label.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
        self.nav_label.setStyleSheet("background-color: #E3F2FD; color: #1565C0; padding: 5px 12px; border-radius: 4px;")
        self.nav_label.setAlignment(Qt.AlignCenter)
        nav_row.addWidget(self.nav_label)
        
        self.next_btn = QPushButton("▶")
        self.next_btn.setFixedWidth(40)
        self.next_btn.setFont(QFont("Microsoft YaHei", 9))
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                padding: 5px 10px;
                border-radius: 4px;
            }
        """)
        nav_row.addWidget(self.next_btn)
        img_layout.addLayout(nav_row)
        
        img_frame.setLayout(img_layout)
        preview_layout.addWidget(img_frame, 1)
        
        # 右侧：表格编辑
        table_edit_frame = QFrame()
        table_edit_layout = QVBoxLayout()
        table_edit_layout.setSpacing(5)
        
        # 款号+类型选择
        info_row = QHBoxLayout()
        info_row.addWidget(QLabel("款号:"))
        self.current_code = QLabel("-")
        self.current_code.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        self.current_code.setStyleSheet("color: #1565C0;")
        info_row.addWidget(self.current_code)
        
        info_row.addWidget(QLabel("  类型:"))
        self.type_group = QButtonGroup(self)
        self.type_top = QRadioButton("上装")
        self.type_top.setChecked(True)
        self.type_group.addButton(self.type_top, 1)
        info_row.addWidget(self.type_top)
        
        self.type_pants = QRadioButton("裤装")
        self.type_group.addButton(self.type_pants, 2)
        info_row.addWidget(self.type_pants)
        
        info_row.addStretch()
        table_edit_layout.addLayout(info_row)
        
        # 尺码勾选
        size_sel_row = QHBoxLayout()
        size_sel_row.addWidget(QLabel("选择尺码:"))
        size_sel_row.addSpacing(10)
        
        self.size_vars = {}
        self.size_order = ["S", "M", "L", "XL", "2XL", "3XL"]
        for size in self.size_order:
            var = QCheckBox(size)
            var.setChecked(True)
            var.stateChanged.connect(self.on_size_change)
            # 更明显的选中效果
            var.setStyleSheet("""
                QCheckBox {
                    background-color: #E3F2FD;
                    color: #1565C0;
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-weight: bold;
                    border: 1px solid #1565C0;
                    font-size: 8px;
                }
                QCheckBox:checked {
                    background-color: #1565C0;
                    color: white;
                }
                QCheckBox::indicator {
                    width: 0px;
                    height: 0px;
                }
            """)
            self.size_vars[size] = var
            size_sel_row.addWidget(var)
        
        select_all_btn = QPushButton("全选")
        select_all_btn.setFixedWidth(42)
        select_all_btn.setFont(QFont("Microsoft YaHei", 7))
        select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #E8F5E9;
                color: #2E7D32;
                border: 1px solid #2E7D32;
                padding: 1px 4px;
                border-radius: 3px;
            }
        """)
        select_all_btn.clicked.connect(self.select_all_sizes)
        size_sel_row.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("全不选")
        deselect_all_btn.setFixedWidth(46)
        deselect_all_btn.setFont(QFont("Microsoft YaHei", 7))
        deselect_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #FFEBEE;
                color: #C62828;
                border: 1px solid #C62828;
                padding: 1px 4px;
                border-radius: 3px;
            }
        """)
        deselect_all_btn.clicked.connect(self.deselect_all_sizes)
        size_sel_row.addWidget(deselect_all_btn)
        
        size_sel_row.addStretch()
        table_edit_layout.addLayout(size_sel_row)
        
        # 温馨提示
        tip_row = QHBoxLayout()
        tip_row.addWidget(QLabel("温馨提示:"))
        self.warm_tip = QLineEdit()
        self.warm_tip.setText("产品尺码均为手工测量，会存在1-3CM误差，属于正常范围。")
        self.warm_tip.setFont(QFont("Microsoft YaHei", 8))
        self.warm_tip.setStyleSheet("border: 1px solid #E0E0E0; padding: 4px 8px; border-radius: 4px;")
        tip_row.addWidget(self.warm_tip)
        table_edit_layout.addLayout(tip_row)
        
        # 尺码表格
        self.size_table = QTableWidget()
        self.size_table.setColumnCount(11)
        if self.type_top.isChecked():
            self.size_table.setHorizontalHeaderLabels(["尺码", "号型", "肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"])
        else:
            self.size_table.setHorizontalHeaderLabels(["尺码", "号型", "腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"])
        
        self.size_table.installEventFilter(self)
        self.size_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E0E0E0;
                gridline-color: #E0E0E0;
            }
            QHeaderView::section {
                background-color: #E3F2FD;
                padding: 6px;
                border: 1px solid #E0E0E0;
                font-weight: bold;
                font-size: 11px;
                color: #1565C0;
            }
            QTableWidget::item {
                text-align: center;
                font-size: 10px;
            }
        """)
        
        self.size_table.horizontalHeader().setDefaultSectionSize(52)
        self.size_table.setColumnWidth(0, 42)  # 尺码列
        self.size_table.setColumnWidth(1, 92)  # 号型列（更宽，避免内容被截断）
        for col in range(2, 11):
            self.size_table.setColumnWidth(col, 50)  # 平均缩小其他列，给号型让空间
        
        self.size_table.verticalHeader().setDefaultSectionSize(28)
        self.size_table.setMinimumHeight(220)
        self.size_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.size_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        table_edit_layout.addWidget(self.size_table, 1)
        
        # 编辑按钮
        edit_btns = QHBoxLayout()
        self.st_reload_btn = QPushButton("🔄 重新识别")
        self.st_reload_btn.setFont(QFont("Microsoft YaHei", 9))
        self.st_reload_btn.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
            }
        """)
        edit_btns.addWidget(self.st_reload_btn)
        
        self.save_btn = QPushButton("💾 保存")
        self.save_btn.setFont(QFont("Microsoft YaHei", 9))
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
            }
        """)
        edit_btns.addWidget(self.save_btn)
        
        edit_btns.addStretch()
        
        self.st_clear_btn = QPushButton("🗑 清空结果")
        self.st_clear_btn.setFont(QFont("Microsoft YaHei", 9))
        self.st_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
            }
        """)
        edit_btns.addWidget(self.st_clear_btn)
        
        edit_btns.addStretch()
        table_edit_layout.addLayout(edit_btns)
        
        table_edit_frame.setLayout(table_edit_layout)
        preview_layout.addWidget(table_edit_frame, 1)
        
        preview_group.setLayout(preview_layout)
        return preview_group
    
    def _create_results_section(self):
        """创建已识别列表区域"""
        list_group = QGroupBox("📋 已识别款号列表")
        list_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        list_layout = QVBoxLayout()
        
        self.result_list = QTableWidget()
        self.result_list.setColumnCount(4)
        self.result_list.setHorizontalHeaderLabels(["款号", "唯品类目", "尺码表类型", "状态"])
        self.result_list.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E0E0E0;
                gridline-color: #E0E0E0;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 6px 8px;
                border: 1px solid #E0E0E0;
                font-weight: bold;
                font-size: 11px;
                min-height: 22px;
            }
            QTableWidget::item {
                font-size: 9px;
            }
        """)
        header = self.result_list.horizontalHeader()
        # 四列均分宽度，避免前两列过宽、后两列过窄
        for col in range(4):
            header.setSectionResizeMode(col, QHeaderView.Stretch)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setMinimumHeight(44)
        self.result_list.verticalHeader().setVisible(False)
        self.result_list.verticalHeader().setDefaultSectionSize(30)
        self.result_list.setMinimumHeight(180)
        
        self.result_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.result_list.doubleClicked.connect(self.on_result_select)
        
        list_layout.addWidget(self.result_list)
        list_group.setLayout(list_layout)
        return list_group
    
    def _create_log_section(self):
        """创建日志区域"""
        log_group = QGroupBox("📝 OCR识别日志")
        log_group.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(120)
        self.log_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #E0E0E0;
                background-color: #1E1E1E;
                color: #00FF00;
                font-family: Consolas, Microsoft YaHei;
                font-size: 9px;
            }
        """)
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        return log_group
    
    def _create_results_compact(self):
        """创建紧凑的结果列表（显示在配置右侧）"""
        results_frame = QFrame()
        results_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 6px;
                padding: 8px;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(6)
        layout.setContentsMargins(6, 6, 6, 6)
        
        # 标题
        title_label = QLabel("📋 已识别款号")
        title_label.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
        title_label.setStyleSheet("color: #1565C0;")
        layout.addWidget(title_label)
        
        # 结果列表
        self.result_list = QTableWidget()
        self.result_list.setColumnCount(4)
        self.result_list.setHorizontalHeaderLabels(["款号", "唯品类目", "类型", "状态"])
        self.result_list.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E0E0E0;
                gridline-color: #E0E0E0;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 6px 8px;
                border: 1px solid #E0E0E0;
                font-weight: bold;
                font-size: 11px;
                min-height: 22px;
            }
            QTableWidget::item {
                font-size: 9px;
            }
        """)
        header = self.result_list.horizontalHeader()
        # 四列均分宽度，避免前两列过宽、后两列过窄
        for col in range(4):
            header.setSectionResizeMode(col, QHeaderView.Stretch)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setMinimumHeight(44)
        self.result_list.verticalHeader().setVisible(False)
        self.result_list.verticalHeader().setDefaultSectionSize(30)
        self.result_list.setMinimumHeight(180)
        self.result_list.setMaximumHeight(240)
        
        self.result_list.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.result_list.doubleClicked.connect(self.on_result_select)
        
        layout.addWidget(self.result_list)
        results_frame.setLayout(layout)
        return results_frame
    
    def _bind_signals(self):
        """绑定信号"""
        self.st_start_btn.clicked.connect(self.start_batch_recognition)
        self.st_reload_btn.clicked.connect(self.reload_current)
        self.st_clear_btn.clicked.connect(self.clear_results)
        self.st_export_all_btn.clicked.connect(self.export_all)
        self.save_btn.clicked.connect(self.save_current)
        self.prev_btn.clicked.connect(self.prev_image)
        self.next_btn.clicked.connect(self.next_image)
        self.type_group.buttonClicked.connect(self.on_type_change)
    
    def get_ocr_engine(self):
        from toolbox.core.ocr import OCREngine
        from toolbox.core.utils import BASE_DIR
        return OCREngine(BASE_DIR)
    
    def check_ocr_status(self):
        try:
            ocr = self.get_ocr_engine()
            mode = self._get_ocr_mode()
            baidu_ok = ocr.is_baidu_available()
            paddlex_ok = ocr.is_paddlex_available() if hasattr(ocr, "is_paddlex_available") else False
            if mode == "cloud":
                if baidu_ok or paddlex_ok:
                    self.ocr_status.setText("🟡 仅云端模式（PaddleX优先）")
                    self.ocr_status.setStyleSheet("color: #FF9800; font-weight: bold;")
                else:
                    self.ocr_status.setText("🔴 云端OCR未配置")
                    self.ocr_status.setStyleSheet("color: #F44336; font-weight: bold;")
            elif mode == "local":
                if ocr.is_paddle_ocr_available():
                    self.ocr_status.setText("🟢 仅本地模式已就绪")
                    self.ocr_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
                else:
                    self.ocr_status.setText("🔴 本地OCR未找到")
                    self.ocr_status.setStyleSheet("color: #F44336; font-weight: bold;")
            elif ocr.is_paddle_ocr_available():
                self.ocr_status.setText("🟢 PaddleOCR已就绪")
                self.ocr_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
            elif baidu_ok or paddlex_ok:
                if paddlex_ok and baidu_ok:
                    self.ocr_status.setText("🟡 云端OCR已就绪（百度/PaddleX）")
                elif paddlex_ok:
                    self.ocr_status.setText("🟡 PaddleX OCR已就绪")
                else:
                    self.ocr_status.setText("🟡 百度OCR已就绪")
                self.ocr_status.setStyleSheet("color: #FF9800; font-weight: bold;")
            else:
                self.ocr_status.setText("🔴 OCR未配置")
                self.ocr_status.setStyleSheet("color: #F44336; font-weight: bold;")
        except Exception:
            self.ocr_status.setText("🔴 OCR未配置")
            self.ocr_status.setStyleSheet("color: #F44336; font-weight: bold;")

    def on_ocr_mode_changed(self, _):
        if getattr(self, "_ocr_tier", "hybrid") == "cloud_only":
            return
        mode = self._get_ocr_mode()
        self.size_ocr_mode = mode
        self.config["size_ocr_mode"] = mode
        self.check_ocr_status()
        if mode == "cloud":
            self.log_signal.emit("OCR模式已切换为：仅云端（PaddleX优先）")
        elif mode == "local":
            self.log_signal.emit("OCR模式已切换为：仅本地模型")
        else:
            self.log_signal.emit("OCR模式已切换为：自动（本地优先）")
    
    def _get_ocr_mode(self):
        if getattr(self, "_ocr_tier", "hybrid") == "cloud_only":
            return "cloud"
        idx = self.ocr_mode_combo.currentIndex() if hasattr(self, "ocr_mode_combo") else 0
        return {0: "auto", 1: "local", 2: "cloud"}.get(idx, "auto")
    
    def show_ocr_tutorial(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("OCR使用教程")
        msg.setIcon(QMessageBox.Information)
        msg.setText("""
<b>本地OCR配置教程：</b><br><br>
1. 下载 PaddleOCR-json_v1.4.1<br>
2. 将文件夹放入 <code>ocr_engine</code> 目录<br>
3. 确保文件夹内有 <code>PaddleOCR_json.exe</code><br><br>
<b>云端OCR配置：</b><br><br>
1. 使用内置的 PaddleX PP-OCRv5 即可直接测试<br>
2. 若需替换账号，可在设置里改 API URL 和 Token<br>
3. 仅云端模式建议选「仅云端（PaddleX）」
        """)
        msg.exec()
    
    def show_ocr_settings(self):
        if getattr(self, "_ocr_tier", "hybrid") == "cloud_only":
            dialog = QDialog(self)
            dialog.setWindowTitle("OCR 设置")
            dialog.setMinimumSize(520, 300)
            layout = QVBoxLayout()
            t = QLabel("OCR 设置（仅云端分发包）\n本包不含本地识别模型，只需配置下方 PaddleX / 或编辑 baidu_api.json。")
            t.setWordWrap(True)
            layout.addWidget(t)
            try:
                ocr = self.get_ocr_engine()
            except Exception:
                ocr = None
            if ocr is None:
                err = QLabel("OCR 模块加载失败")
                err.setStyleSheet("color: #F44336;")
                layout.addWidget(err)
            else:
                paddlex_form = QFormLayout()
                pu, pt = QLineEdit(), QLineEdit()
                pu.setPlaceholderText("PaddleX API URL")
                pt.setPlaceholderText("Access Token")
                try:
                    cfg = ocr.baidu_config or {}
                    pu.setText(str(cfg.get("paddlex_api_url", "") or ""))
                    pt.setText(str(cfg.get("paddlex_token", "") or ""))
                except Exception:
                    pass
                paddlex_form.addRow("API URL:", pu)
                paddlex_form.addRow("Access Token:", pt)
                layout.addLayout(paddlex_form)
            btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            btns.accepted.connect(dialog.accept)
            btns.rejected.connect(dialog.reject)
            layout.addWidget(btns)
            dialog.setLayout(layout)
            if dialog.exec() == QDialog.Accepted and ocr is not None:
                url, tok = pu.text().strip(), pt.text().strip()
                if bool(url) ^ bool(tok):
                    QMessageBox.warning(self, "提示", "PaddleX 需要同时填写 API URL 和 Access Token")
                    return
                ocr.save_cloud_config(
                    app_id="",
                    api_key=ocr.baidu_config.get("api_key", ""),
                    secret_key=ocr.baidu_config.get("secret_key", ""),
                    paddlex_api_url=url,
                    paddlex_token=tok,
                    cloud_provider="paddlex",
                )
                self.check_ocr_status()
                self.log_signal.emit("OCR 配置已保存")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("OCR 设置")
        dialog.setMinimumSize(520, 320)
        layout = QVBoxLayout()
        
        title = QLabel("OCR 设置")
        title.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        layout.addWidget(title)
        
        # PaddleOCR状态
        paddle_frame = QFrame()
        paddle_frame.setStyleSheet("background-color: #F5F5F5; padding: 10px; border-radius: 4px;")
        paddle_layout = QVBoxLayout()
        
        ocr = None
        try:
            ocr = self.get_ocr_engine()
            if ocr.is_paddle_ocr_available():
                paddle_status = QLabel("✓ 已找到 PaddleOCR_json.exe")
                paddle_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
            else:
                paddle_status = QLabel("✗ 未找到 PaddleOCR_json.exe")
                paddle_status.setStyleSheet("color: #F44336; font-weight: bold;")
        except Exception:
            paddle_status = QLabel("✗ OCR模块加载失败")
            paddle_status.setStyleSheet("color: #F44336; font-weight: bold;")
        
        paddle_layout.addWidget(paddle_status)
        
        paddle_hint = QLabel("请将 PaddleOCR-json_v1.4.1 文件夹放入 ocr_engine 文件夹")
        paddle_hint.setStyleSheet("color: #666;")
        paddle_layout.addWidget(paddle_hint)
        
        paddle_frame.setLayout(paddle_layout)
        layout.addWidget(paddle_frame)
        
        # PaddleX 设置（保留此项，去掉不常用配置）
        paddlex_title = QLabel("PaddleX OCR API 设置")
        paddlex_title.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        layout.addWidget(paddlex_title)
        paddlex_form = QFormLayout()
        paddlex_url = QLineEdit()
        paddlex_url.setPlaceholderText("请输入 PaddleX API URL（示例中提供）")
        paddlex_form.addRow("API URL:", paddlex_url)
        paddlex_token = QLineEdit()
        paddlex_token.setPlaceholderText("请输入 Access Token")
        paddlex_form.addRow("Access Token:", paddlex_token)

        # 回填现有配置，便于直接修改
        if ocr is not None:
            try:
                cfg = ocr.baidu_config or {}
                paddlex_url.setText(str(cfg.get("paddlex_api_url", "") or ""))
                paddlex_token.setText(str(cfg.get("paddlex_token", "") or ""))
            except Exception:
                pass
        
        layout.addLayout(paddlex_form)
        
        # 按钮
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dialog.accept)
        btns.rejected.connect(dialog.reject)
        layout.addWidget(btns)
        
        dialog.setLayout(layout)
        if dialog.exec() == QDialog.Accepted:
            paddlex_url_text = paddlex_url.text().strip()
            paddlex_token_text = paddlex_token.text().strip()
            if bool(paddlex_url_text) ^ bool(paddlex_token_text):
                QMessageBox.warning(self, "提示", "PaddleX 需要同时填写 API URL 和 Access Token")
                return
            try:
                ocr_save = self.get_ocr_engine()
                if hasattr(ocr_save, "save_cloud_config"):
                    ocr_save.save_cloud_config(
                        app_id="",
                        api_key=ocr_save.baidu_config.get("api_key", ""),
                        secret_key=ocr_save.baidu_config.get("secret_key", ""),
                        paddlex_api_url=paddlex_url_text,
                        paddlex_token=paddlex_token_text,
                        cloud_provider="paddlex",
                    )
                else:
                    ocr_save.save_baidu_config("", "", "")
                self.check_ocr_status()
                self.log_signal.emit("PaddleX OCR配置已保存")
                QMessageBox.information(self, "完成", "OCR配置已保存")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"OCR配置保存失败: {e}")
    
    def select_dir(self, line_edit):
        path = QFileDialog.getExistingDirectory(self, "选择目录")
        if path:
            line_edit.setText(path)
    
    def select_excel_size(self, line_edit):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel文件 (*.xlsx *.xls)")
        if path:
            line_edit.setText(path)
            self.analyze_size_excel(path)
    
    def analyze_size_excel(self, path):
        try:
            import pandas as pd
            df = pd.read_excel(path)
            
            code_col = None
            category_col = None
            
            for c in df.columns:
                if "唯品款号" in str(c):
                    code_col = c
                    break
            
            for c in df.columns:
                if "品类" in str(c):
                    category_col = c
                    break
            
            if code_col and category_col:
                self.log_signal.emit(f"✓ Excel导入成功: 找到「唯品款号」和「唯品类目」列，共 {len(df)} 行数据")
            else:
                missing = []
                if not code_col:
                    missing.append("唯品款号")
                if not category_col:
                    missing.append("唯品类目")
                self.log_signal.emit(f"⚠ Excel表头缺少字段: {', '.join(missing)}")
        except Exception as e:
            self.log_signal.emit(f"⚠ 读取Excel失败: {e}")
    
    # ============ 尺码表核心业务方法 ============
    
    def load_ref_data(self):
        """加载参考数据：尺码表类型映射和号型映射"""
        try:
            from toolbox.core.utils import REFERENCE_DIR
            import pandas as pd
            import openpyxl
            
            # 加载尺码表类型映射
            type_file = os.path.join(REFERENCE_DIR, "尺码表类型.xlsx")
            if os.path.exists(type_file):
                df = pd.read_excel(type_file)
                for _, row in df.iterrows():
                    category = str(row.get("唯品类目", "")).strip()
                    first_class = str(row.get("一级分类", "")).strip()
                    template_type = str(row.get("尺码表模板类型", "")).strip()
                    if category and template_type:
                        if "上装" in template_type:
                            prod_type = "上装"
                        elif "裤装" in template_type:
                            prod_type = "裤装"
                        else:
                            prod_type = "上装"
                        self.category_type_map[category] = (first_class, prod_type, template_type)
            
            # 加载号型映射
            number_file = os.path.join(REFERENCE_DIR, "号型.xlsx")
            if os.path.exists(number_file):
                wb = openpyxl.load_workbook(number_file)
                ws = wb.active
                header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                group_info = []
                for i in range(0, len(header_row), 4):
                    if i + 1 < len(header_row):
                        template_type = str(header_row[i + 1]).strip() if header_row[i + 1] else ""
                        if template_type and template_type not in ["None", ""]:
                            group_info.append((i, template_type))
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for col_offset, template_type in group_info:
                        first_class = str(row[col_offset]).strip() if row[col_offset] else ""
                        size_val = str(row[col_offset + 1]).strip() if row[col_offset + 1] else ""
                        num_val = str(row[col_offset + 2]).strip() if row[col_offset + 2] else ""
                        if first_class and size_val and num_val and first_class not in ["一级分类", "None", ""]:
                            key = (first_class, template_type, size_val.upper())
                            self.number_map[key] = num_val
            
            self.log_signal.emit(f"✓ 已加载 {len(self.category_type_map)} 个类目映射")
            self.log_signal.emit(f"✓ 已加载 {len(self.number_map)} 个号型映射")
        except Exception as e:
            self.log_signal.emit(f"⚠ 加载参考数据失败: {e}")
    
    def load_excel_size_mapping(self, excel_file):
        """从Excel商品表加载每个款号的尺码和尺码明细，建立按款号的号型映射
        每个款号可能有不同的尺码-号型对应关系
        """
        if not excel_file or not os.path.exists(excel_file):
            return
        # 重新读取「尺码映射/尺码别名.json」，便于不重启就改尺码对应关系
        self._external_size_aliases = load_external_size_aliases()
        
        try:
            import pandas as pd
            
            df = pd.read_excel(excel_file)
            
            # 精确匹配列名
            code_col = None
            size_col = None
            detail_col = None
            
            for c in df.columns:
                col_str = str(c).strip()
                # 精确匹配：唯品款号
                if col_str == "唯品款号":
                    code_col = c
                # 精确匹配：尺码（不含明细）
                elif col_str == "尺码":
                    size_col = c
                # 精确匹配：尺码明细
                elif col_str == "尺码明细":
                    detail_col = c
            
            if not code_col or not size_col or not detail_col:
                self.log_signal.emit(f"⚠ Excel表缺少关键列，请检查是否有：唯品款号、尺码、尺码明细")
                return
            
            # 解析每个款号的尺码-号型映射
            for _, row in df.iterrows():
                code = str(row[code_col]).strip() if pd.notna(row[code_col]) else ""
                if not code or code == "nan":
                    continue
                
                size_str = str(row[size_col]).strip() if pd.notna(row[size_col]) else ""
                detail_str = str(row[detail_col]).strip() if pd.notna(row[detail_col]) else ""
                
                if not size_str or not detail_str:
                    # 尺码或尺码明细为空时，整行跳过（不建立任何映射）
                    continue
                
                size_list = [s.strip() for s in size_str.split(",") if s.strip()]
                detail_list = [d.strip() for d in detail_str.split(",") if d.strip()]
                
                # 建立该款号的映射：尺码字母 -> 号型（可能为空）
                # 如果某个尺码没有对应号型，存为空字符串
                size_map = {}
                for i, size_val in enumerate(size_list):
                    size_code = size_val
                    if "(" in size_val and ")" in size_val:
                        size_code = size_val[size_val.find("(")+1:size_val.find(")")]
                    elif size_val.isdigit() and len(size_val) == 2:
                        size_code = self._convert_size_code(size_val)
                    # 统一成界面用尺码（XXL→2XL），避免 Excel 写 XXL 与界面 2XL 对不上
                    canon = canonicalize_size(size_code, self._external_size_aliases)
                    map_key = canon if canon else str(size_code).strip().upper()
                    
                    # 只有当尺码明细列表中有对应位置的数据时才存
                    if i < len(detail_list):
                        size_map[map_key] = detail_list[i]
                    else:
                        size_map[map_key] = ""  # 没有对应号型，存空字符串
                
                if size_map:
                    self.excel_size_number_map[code] = size_map
            
            self.log_signal.emit(f"✓ 已加载 {len(self.excel_size_number_map)} 个款号的尺码号型映射")
            
        except Exception as e:
            self.log_signal.emit(f"⚠ 加载Excel尺码映射失败: {e}")
    
    def _convert_size_code(self, code):
        """将数字尺码码转换为字母尺码，如 03->S, 04->M"""
        size_map = {"03": "S", "04": "M", "05": "L", "06": "XL", 
                    "07": "2XL", "08": "3XL", "09": "4XL", "10": "5XL"}
        return size_map.get(code, code)
    
    def init_field_aliases(self):
        """初始化字段别名"""
        import json
        ensure_default_mapping_files()
        try:
            from toolbox.core.utils import BASE_DIR
            mapping_file = os.path.join(BASE_DIR, "尺码字段映射.json")
            if os.path.exists(mapping_file):
                with open(mapping_file, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    field_mapping = config.get("字段映射", {})
                    self.field_aliases = {}
                    for field_name, field_info in field_mapping.items():
                        standard_name = field_info.get("标准名", field_name)
                        aliases = field_info.get("别名", [])
                        for alias in aliases:
                            self.field_aliases[alias] = standard_name
                    self.exclude_keywords = config.get("排除关键词", [])
                    self.required_fields = config.get("必填字段", {
                        "上装": ["肩宽", "胸围", "衣长", "袖长", "腰围"],
                        "裤装": ["腰围", "臀围", "裤长", "裤脚围"],
                    })
            else:
                self._init_default_aliases()
        except Exception:
            self._init_default_aliases()
        # 外部文件夹中的字段别名（可覆盖同名键）
        self._external_size_aliases = load_external_size_aliases()
        for alias, std in load_external_field_aliases().items():
            self.field_aliases[alias] = std
        try:
            self.log_signal.emit(
                f"✓ 外部映射目录: {MAPPINGS_DIR}（尺码别名在加载 Excel 或识别图片时会自动重读；字段别名改完需重启程序）"
            )
        except Exception:
            pass
    
    def _init_default_aliases(self):
        """初始化默认字段别名"""
        self.field_aliases = {
            "衣长": "衣长", "身长": "衣长", "后中长": "衣长",
            "胸围": "胸围", "胸": "胸围",
            "肩宽": "肩宽", "肩": "肩宽",
            "袖长": "袖长", "袖": "袖长", "肩袖长": "袖长",
            "腰围": "腰围", "腰": "腰围",
            "领围": "领围", "领": "领围",
            "下摆围": "下摆围", "下摆": "下摆围",
            "袖笼围": "袖笼围", "袖笼": "袖笼围",
            "充绒量": "充绒量", "充绒": "充绒量",
            "臀围": "臀围", "臀": "臀围",
            "大腿围": "大腿围", "大腿": "大腿围",
            "裤长": "裤长", "裤": "裤长",
            "前浪": "前浪", "后浪": "后浪",
            "膝围": "膝围", "膝": "膝围",
            "裤脚围": "裤脚围", "裤脚": "裤脚围", "脚口": "裤脚围", "裤口": "裤脚围",
        }
        self.exclude_keywords = ["尺码信息", "尺码表", "尺寸表", "单位", "CM", "单位:cm"]
        self.required_fields = {
            "上装": ["肩宽", "胸围", "衣长", "袖长", "腰围"],
            "裤装": ["腰围", "臀围", "裤长", "裤脚围"],
        }
    
    def get_type_from_category(self, category):
        """根据唯品类目判断上装/裤装"""
        if not category:
            return "上装"
        if category in self.category_type_map:
            return self.category_type_map[category][1]
        return None
    
    def get_default_number(self, size, prod_type, code=None):
        """获取号型 - 只用Excel表格中的数据，没有就留空"""
        # 从Excel表格中该款号的映射获取（尺码已统一为 2XL 等与界面一致）
        if code and code in self.excel_size_number_map:
            size_map = self.excel_size_number_map[code]
            key = canonicalize_size(size, self._external_size_aliases) or str(size).strip().upper()
            if key in size_map:
                return size_map[key]
            if size.upper() in size_map:
                return size_map[size.upper()]
        # 没有则返回空字符串
        return ""
    
    def start_batch_recognition(self):
        """开始批量识别"""
        folder = self.st_folder.text().strip()
        if not folder or not os.path.exists(folder):
            QMessageBox.warning(self, "错误", "请选择有效的图片文件夹")
            return
        
        excel_file = self.st_excel.text().strip()
        
        self.log_signal.emit("开始批量识别...")
        self._mismatch_rows = []
        self._mismatch_seen = set()
        
        import pandas as pd
        img_exts = (".jpg", ".jpeg", ".png", ".bmp", ".webp")
        images = []
        
        # 从Excel获取款号列表和唯品类目
        excel_codes = set()
        code_category_map = {}
        if excel_file and os.path.exists(excel_file):
            try:
                df = pd.read_excel(excel_file)
                code_col = None
                category_col = None
                for c in df.columns:
                    if "唯品款号" in str(c) and not code_col:
                        code_col = c
                    if "唯品类目" in str(c) and not category_col:
                        category_col = c
                
                if code_col:
                    for _, row in df.iterrows():
                        code = str(row[code_col]).strip() if pd.notna(row[code_col]) else ""
                        if code and code != "nan":
                            excel_codes.add(code)
                            if category_col:
                                category = str(row[category_col]).strip() if pd.notna(row[category_col]) else ""
                                if category and category != "nan":
                                    code_category_map[code] = category
                    
                    self.log_signal.emit(f"从Excel读取到 {len(excel_codes)} 个款号")
                    
                    # 加载款号尺码号型映射（优先表格映射，兜底默认映射）
                    self.load_excel_size_mapping(excel_file)
            except Exception as e:
                self.log_signal.emit(f"读取Excel失败: {e}")
        
        try:
            folder_items = os.listdir(folder)
            has_subfolders = any(os.path.isdir(os.path.join(folder, item)) for item in folder_items)
            has_images = any(
                os.path.isfile(os.path.join(folder, item)) and item.lower().endswith(img_exts)
                for item in folder_items
            )
            
            self.log_signal.emit(f"文件夹结构检查: 子文件夹={has_subfolders}, 图片文件={has_images}")
            
            if has_subfolders:
                # 优先使用子文件夹模式：文件夹名=款号
                self.log_signal.emit("使用文件夹名称匹配模式（文件夹名=唯品款号）")
                for item in folder_items:
                    item_path = os.path.join(folder, item)
                    if os.path.isdir(item_path):
                        code = item.strip()
                        if excel_codes and code not in excel_codes:
                            self.log_signal.emit(f"  跳过（不在Excel中）: {code}")
                            continue
                        category = code_category_map.get(code, "")
                        prod_type = self.get_type_from_category(category)
                        if prod_type is None:
                            self.log_signal.emit(f"  跳过（无类别信息）: {code}")
                            continue
                        try:
                            # 在子文件夹中找到第一张图片
                            img_file_found = False
                            for img_file in os.listdir(item_path):
                                if img_file.lower().endswith(img_exts):
                                    img_path = os.path.join(item_path, img_file)
                                    images.append({"code": code, "path": img_path, "type": prod_type, "category": category})
                                    self.log_signal.emit(f"  匹配: {code} -> {img_file}")
                                    img_file_found = True
                                    break
                            if not img_file_found:
                                self.log_signal.emit(f"  警告: {code} 文件夹中没有图片文件")
                        except Exception as e:
                            self.log_signal.emit(f"  错误: {code} 文件夹读取失败: {e}")
                            continue
            elif has_images:
                # 图片平铺模式：图片文件名需要包含款号
                self.log_signal.emit("使用图片文件名匹配模式（图片名需包含款号）")
                for item in folder_items:
                    item_path = os.path.join(folder, item)
                    if os.path.isfile(item_path) and item.lower().endswith(img_exts):
                        # 尝试从图片文件名提取款号（去掉扩展名）
                        code = item.rsplit(".", 1)[0].strip()
                        if excel_codes and code not in excel_codes:
                            # 如果完整文件名不匹配，尝试部分匹配
                            found_match = False
                            for excel_code in excel_codes:
                                if excel_code in code or code in excel_code:
                                    code = excel_code
                                    found_match = True
                                    break
                            if not found_match:
                                self.log_signal.emit(f"  跳过（不匹配款号）: {item}")
                                continue
                        category = code_category_map.get(code, "")
                        prod_type = self.get_type_from_category(category)
                        if prod_type is None:
                            self.log_signal.emit(f"  跳过（无类别信息）: {code}")
                            continue
                        images.append({"code": code, "path": item_path, "type": prod_type, "category": category})
                        self.log_signal.emit(f"  匹配: {item} -> {code}")
            else:
                self.log_signal.emit("错误: 文件夹中既没有子文件夹也没有图片文件")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"扫描文件夹失败:\n{e}")
            return
        
        self.results = images
        self.current_idx = 0
        self.st_status.setText(f"找到 {len(images)} 个款号")
        self.log_signal.emit(f"找到 {len(images)} 个款号")
        
        if images:
            self.show_current()
        else:
            self.log_signal.emit("未找到匹配的图片")
    
    def show_current(self):
        """显示当前识别结果"""
        if not self.results or self.current_idx >= len(self.results):
            return
        
        item = self.results[self.current_idx]
        
        # 更新款号显示
        self.current_code.setText(item["code"])
        
        # 更新导航
        self.nav_label.setText(f"{self.current_idx + 1}/{len(self.results)}")
        
        # 根据款号类型自动切换上装/裤装
        prod_type = item.get("type", "上装")
        if prod_type == "上装":
            self.type_top.setChecked(True)
        else:
            self.type_pants.setChecked(True)
        
        # 显示图片预览 - 使用更大的尺寸，保持纵横比
        try:
            from PIL import Image
            img = Image.open(item["path"]).convert("RGB")
            
            # 获取Label的实际尺寸
            label_width = self.preview_label.width()
            label_height = self.preview_label.height()
            
            if label_width <= 0:
                label_width = 280
                label_height = 450
            
            # 计算缩放 - 保持纵横比，让图片完全显示
            img_width, img_height = img.size
            if img_width <= 0 or img_height <= 0:
                raise RuntimeError("图片尺寸异常")
            scale = min(label_width / img_width, label_height / img_height)
            new_width = max(1, int(img_width * scale))
            new_height = max(1, int(img_height * scale))
            
            # 使用LANCZOS高质量缩放
            img_resized = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            
            # 采用稳定的 RGB 字节转 QImage，避免部分环境 ImageQt 引起绘制设备异常
            data = img_resized.tobytes("raw", "RGB")
            qimg = QImage(data, new_width, new_height, new_width * 3, QImage.Format_RGB888).copy()
            pixmap = QPixmap.fromImage(qimg)
            if pixmap.isNull():
                raise RuntimeError("预览图创建失败")
            
            self.preview_label.setPixmap(pixmap)
            self.preview_label.setText("")
            
        except Exception as e:
            self.preview_label.setText(f"图片加载失败: {e}")
            self.log_signal.emit(f"图片加载失败: {e}")
        
        # 检查该款是否已保存
        code = item["code"]
        saved_data = None
        self.log_signal.emit(f"检查款号: '{code}'，已保存数量: {len(self.confirmed_results)}")
        self.log_signal.emit(f"已保存的所有款号: {[c.get('款号') for c in self.confirmed_results]}")
        for confirmed in self.confirmed_results:
            confirmed_code = confirmed.get("款号")
            self.log_signal.emit(f"  对比: '{code}' == '{confirmed_code}' ? {code == confirmed_code}")
            if confirmed_code == code:
                saved_data = confirmed
                break
        
        if saved_data:
            # 已保存：用保存的数据填充表格（不OCR，不rebuild）
            self.log_signal.emit(f"✓ 加载已保存的数据: {code}")
            self.log_signal.emit(f"  保存的数据: {saved_data}")
            self.ocr_recognized = {}
            self.init_table_with_saved_data(saved_data)
            # 不再调用rebuild_table或ocr_image
        else:
            # 未保存：重新OCR识别
            self.log_signal.emit(f"未找到保存数据，重新OCR: {code}")
            self.ocr_recognized = {}
            self.init_table()
            self.ocr_image(item["path"])
        
        # 更新结果列表
        self.update_result_list()
    
    def init_table_with_saved_data(self, saved_data):
        """用保存的数据初始化表格 - 同时考虑尺码勾选状态"""
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        
        if prod_type == "上装":
            headers = ["尺码", "号型", "肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
            required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
        else:
            headers = ["尺码", "号型", "腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
            required_fields = ["腰围", "臀围", "裤长", "裤脚围"]
        
        self.size_table.setHorizontalHeaderLabels(headers)
        self.size_table.setRowCount(0)
        
        size_order = ["S", "M", "L", "XL", "2XL", "3XL"]
        code = saved_data.get("款号", "")
        row_idx = 0
        
        for size in size_order:
            # 检查用户是否勾选了这个尺码
            var = self.size_vars.get(size)
            if var and not var.isChecked():
                continue
            
            # 检查这个尺码是否有保存的数据
            has_data = False
            for key, value in saved_data.items():
                if key.startswith(f"{size}_"):
                    has_data = True
                    break
            
            self.size_table.insertRow(row_idx)
            
            # 尺码列
            size_item = QTableWidgetItem(size)
            size_item.setTextAlignment(Qt.AlignCenter)
            self.size_table.setItem(row_idx, 0, size_item)
            
            if has_data:
                # 号型列
                number = saved_data.get(f"{size}_号型", "")
                number_item = QTableWidgetItem(number)
                number_item.setTextAlignment(Qt.AlignCenter)
                self.size_table.setItem(row_idx, 1, number_item)
                
                # 其他字段列
                for col in range(2, len(headers)):
                    field = headers[col]
                    value = saved_data.get(f"{size}_{field}", "")
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.size_table.setItem(row_idx, col, item)
            else:
                # 新增尺码：填充号型，其他字段留空或填/
                number = self.get_default_number(size, prod_type, code)
                number_item = QTableWidgetItem(number)
                number_item.setTextAlignment(Qt.AlignCenter)
                self.size_table.setItem(row_idx, 1, number_item)
                
                for col in range(2, len(headers)):
                    field = headers[col]
                    if field in required_fields:
                        item = QTableWidgetItem("/")
                    else:
                        item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.size_table.setItem(row_idx, col, item)
            
            row_idx += 1
        
        # 设置列宽
        self.size_table.setColumnWidth(0, 42)
        self.size_table.setColumnWidth(1, 92)
    
    def init_table(self):
        """初始化尺码表格 - 新行数据留空让用户自行填写"""
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        
        # 获取当前款号
        code = None
        if self.results and self.current_idx < len(self.results):
            code = self.results[self.current_idx].get("code")
        
        if prod_type == "上装":
            headers = ["尺码", "号型", "肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
        else:
            headers = ["尺码", "号型", "腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
        
        self.size_table.setHorizontalHeaderLabels(headers)
        self.size_table.setRowCount(0)
        
        # 只添加勾选的尺码 - 新行数据留空
        row_idx = 0
        for size in self.size_order:
            var = self.size_vars.get(size)
            if var and var.isChecked():
                self.size_table.insertRow(row_idx)
                
                # 尺码列
                size_item = QTableWidgetItem(size)
                size_item.setTextAlignment(Qt.AlignCenter)
                self.size_table.setItem(row_idx, 0, size_item)
                
                # 号型列 - 自动获取（优先表格映射，兜底默认映射）
                number = self.get_default_number(size, prod_type, code)
                number_item = QTableWidgetItem(number)
                number_item.setTextAlignment(Qt.AlignCenter)
                self.size_table.setItem(row_idx, 1, number_item)
                
                # 其他列 - 留空让用户自行填写
                for j in range(2, len(headers)):
                    item = QTableWidgetItem("")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.size_table.setItem(row_idx, j, item)
                
                row_idx += 1
        
        # 设置列宽
        self.size_table.setColumnWidth(0, 42)  # 尺码
        self.size_table.setColumnWidth(1, 92)  # 号型
    
    def rebuild_table(self):
        """根据当前勾选状态重建表格（保留用户编辑的数据）"""
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        
        # 获取必填字段
        if prod_type == "上装":
            required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
            headers = ["尺码", "号型", "肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
        else:
            required_fields = ["腰围", "臀围", "裤长", "裤脚围"]  # 4个必填
            headers = ["尺码", "号型", "腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
        
        # 更新表头
        self.size_table.setHorizontalHeaderLabels(headers)
        
        # 获取OCR识别数据
        ocr_data = self.ocr_recognized if hasattr(self, 'ocr_recognized') else {}
        self.log_signal.emit(f"rebuild_table: ocr_data = {ocr_data}")
        
        # 保存当前表格中的用户编辑数据
        existing_data = {}
        for row in range(self.size_table.rowCount()):
            size_item = self.size_table.item(row, 0)
            if size_item and size_item.text():
                size = size_item.text()
                row_data = {}
                for col in range(1, self.size_table.columnCount()):  # 从1开始，跳过尺码列
                    item = self.size_table.item(row, col)
                    header = self.size_table.horizontalHeaderItem(col)
                    if header:
                        row_data[header.text()] = item.text() if item else ""
                existing_data[size] = row_data
        
        # 清空表格
        self.size_table.setRowCount(0)
        
        # 按顺序添加勾选的尺码
        row_idx = 0
        for size in self.size_order:
            var = self.size_vars.get(size)
            if var and var.isChecked():
                # 获取当前款号，用于号型映射
                code = self.results[self.current_idx]["code"] if self.results and self.current_idx < len(self.results) else None
                
                # 优先使用用户编辑的数据
                if size in existing_data:
                    user_data = existing_data[size]
                    number = user_data.get("号型", self.get_default_number(size, prod_type, code))
                    values = [size, number]
                    
                    # 填充其他字段
                    for field in headers[2:]:
                        values.append(user_data.get(field, ""))
                else:
                    # 使用OCR识别数据
                    size_ocr_data = ocr_data.get(size, {})
                    number = self.get_default_number(size, prod_type, code)
                    values = [size, number]
                    
                    # 填充字段
                    for field in headers[2:]:
                        if size_ocr_data.get(field):
                            values.append(size_ocr_data[field])
                        elif field in required_fields:
                            values.append("/")
                        else:
                            values.append("")
                    
                    self.log_signal.emit(f"  尺码{size} OCR数据: {size_ocr_data}")
                
                # 添加行
                self.size_table.insertRow(row_idx)
                for col_idx, val in enumerate(values):
                    item = QTableWidgetItem(val)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.size_table.setItem(row_idx, col_idx, item)
                
                row_idx += 1
        
        # 更新列宽设置
        self.size_table.setColumnWidth(0, 42)  # 尺码
        self.size_table.setColumnWidth(1, 92)  # 号型
    
    def on_size_change(self):
        """尺码勾选变化时重建表格 - 始终考虑勾选状态"""
        code = self.results[self.current_idx]["code"] if self.results and self.current_idx < len(self.results) else None
        
        # 检查是否已保存
        saved_data = None
        if code:
            for confirmed in self.confirmed_results:
                if confirmed.get("款号") == code:
                    saved_data = confirmed
                    break
        
        if saved_data:
            # 已保存：用保存的数据（init_table_with_saved_data已支持勾选状态）
            self.init_table_with_saved_data(saved_data)
        else:
            # 未保存：正常rebuild
            self.rebuild_table()
        
        self.log_signal.emit("尺码选择已更新")
    
    def select_all_sizes(self):
        """全选尺码"""
        for var in self.size_vars.values():
            var.setChecked(True)
        self.rebuild_table()
        self.log_signal.emit("已全选尺码")
    
    def deselect_all_sizes(self):
        """全不选尺码"""
        for var in self.size_vars.values():
            var.setChecked(False)
        self.rebuild_table()
        self.log_signal.emit("已取消全选")
    
    def _record_excel_image_size_diff(self):
        """对比当前款 Excel「尺码/尺码明细」与 OCR 识别到的尺码行，记入待导出列表。"""
        if not self.results or self.current_idx >= len(self.results):
            return
        code = str(self.results[self.current_idx].get("code", "")).strip()
        if not code:
            return
        excel_map = self.excel_size_number_map.get(code) or {}
        excel_canon = set()
        for k in excel_map:
            c = canonicalize_size(str(k), self._external_size_aliases) or str(k).strip().upper()
            if c:
                excel_canon.add(c)
        ocr_canon = set()
        for k in (self.ocr_recognized or {}):
            c = canonicalize_size(str(k), self._external_size_aliases) or str(k).strip().upper()
            if c:
                ocr_canon.add(c)
        for s in sorted(excel_canon - ocr_canon):
            key = (code, "excel_only", s)
            if key in self._mismatch_seen:
                continue
            self._mismatch_seen.add(key)
            self._mismatch_rows.append(
                {
                    "唯品款号": code,
                    "问题类型": "表里有、图上未识别到",
                    "尺码": s,
                    "说明": "Excel 中「尺码/尺码明细」含该尺码，但当前尺码图 OCR 未得到该行，请核对图片或手改表/图。",
                }
            )
        for s in sorted(ocr_canon - excel_canon):
            key = (code, "ocr_only", s)
            if key in self._mismatch_seen:
                continue
            self._mismatch_seen.add(key)
            self._mismatch_rows.append(
                {
                    "唯品款号": code,
                    "问题类型": "图上有、表里未配置",
                    "尺码": s,
                    "说明": "尺码图 OCR 识别到该尺码行，但 Excel 该款「尺码/尺码明细」未包含，请补表或核对是否误识别。",
                }
            )

    def export_size_mismatch_report(self):
        """导出尺码不一致记录为 Excel，便于用户手动更正。"""
        if not self._mismatch_rows:
            QMessageBox.information(self, "提示", "当前没有已记录的尺码不一致项（需先对款号做过识别且 Excel 含尺码映射列）。")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "保存不一致表", "尺码不一致需手改.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            import pandas as pd

            pd.DataFrame(self._mismatch_rows).to_excel(path, index=False)
            QMessageBox.information(self, "完成", f"已导出 {len(self._mismatch_rows)} 条记录：\n{path}")
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    def _auto_export_size_mismatch_report(self, save_dir):
        """导出全部时自动导出不一致表：有数据才导出。"""
        if not self._mismatch_rows:
            return ""
        try:
            import pandas as pd

            path = os.path.join(save_dir, "尺码不一致需手改.xlsx")
            pd.DataFrame(self._mismatch_rows).to_excel(path, index=False)
            self.log_signal.emit(f"导出不一致表: {path}")
            return path
        except Exception as e:
            self.log_signal.emit(f"导出不一致表失败: {e}")
            return ""

    def ocr_image(self, image_path):
        """对图片进行OCR识别"""
        self.log_signal.emit(f"正在识别: {image_path}")
        
        try:
            ocr = self.get_ocr_engine()

            mode = self._get_ocr_mode()
            if mode == "cloud":
                self.log_signal.emit("使用云端OCR（已禁用本地模型）...")
            else:
                self.log_signal.emit("使用自动OCR（优先本地，失败再云端）...")
            ocr_result, engine_name = ocr.ocr_image_items(image_path, engine=mode)
            if ocr_result:
                self.log_signal.emit(f"识别成功（{engine_name}），{len(ocr_result)} 个元素")
                self.log_signal.emit(f"调试OCR结果: {ocr_result[:5]}...")
                self.parse_ocr_result(ocr_result)
                return

            self.log_signal.emit("OCR识别失败，请检查OCR设置")
        except Exception as e:
            self.log_signal.emit(f"OCR识别出错: {e}")
    
    def parse_ocr_result(self, ocr_items):
        """解析OCR结果并填充表格 - 完整复刻源码逻辑"""
        import re
        
        self._external_size_aliases = load_external_size_aliases()
        self.log_signal.emit(f"解析 {len(ocr_items)} 个OCR元素...")
        
        # 获取当前类型的所有字段
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        if prod_type == "上装":
            all_fields = ["肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
        else:
            all_fields = ["腰围", "臀围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "大腿围", "充绒量"]
        
        # 构建字段别名映射
        alias_to_field = {}
        for alias, standard in self.field_aliases.items():
            if standard in all_fields:  # 只处理当前类型的字段
                alias_to_field[alias] = standard
        
        # 按别名长度排序（优先匹配更长的）
        sorted_aliases = sorted(alias_to_field.keys(), key=len, reverse=True)
        
        # 分类OCR元素
        size_items = []  # 尺码
        field_items = []  # 字段名
        value_items = []  # 数值
        
        for item in ocr_items:
            text = item.get("text", "").strip()
            x = item.get("x", 0)
            y = item.get("y", 0)
            
            if not text:
                continue
            
            # 跳过排除关键词
            if any(kw in text for kw in self.exclude_keywords):
                continue
            
            # 尺码匹配（XXL→2XL、XXXL→3XL 等，含外部 JSON）
            normalized_size = canonicalize_size(text, self._external_size_aliases)
            if normalized_size:
                size_items.append({"text": normalized_size, "x": x, "y": y})
                continue
            
            # 字段匹配
            matched_field = None
            for alias in sorted_aliases:
                if alias in text or text == alias:
                    matched_field = alias_to_field[alias]
                    break
            
            if matched_field:
                field_items.append({"text": text, "field": matched_field, "x": x, "y": y})
                continue
            
            # 数值提取
            numbers = re.findall(r"\d+\.?\d*", text)
            if numbers and len(text) < 15:
                try:
                    val = float(numbers[0])
                    if 10 < val < 200:  # 合理的尺码范围
                        value_items.append({"text": numbers[0], "value": val, "x": x, "y": y})
                except (ValueError, TypeError, OverflowError):
                    pass
        
        # 改为“内容优先”策略：不再按固定区域硬裁剪，避免尺码超出区域被漏识别
        self.log_signal.emit(
            f"  内容优先解析: 尺码候选{len(size_items)}个, 数值候选{len(value_items)}个, 字段{len(field_items)}个"
        )
        
        self.log_signal.emit(f"  识别到尺码: {[s['text'] for s in size_items]}")
        
        # 保存OCR识别的尺码数据（用于后续填充）
        # 结构: {size: {field: value}}
        self.ocr_recognized = {}
        
        # 按y坐标分行
        rows = {}
        for si in size_items:
            y_key = si["y"] // 30 * 30
            if y_key not in rows:
                rows[y_key] = {"Size": si["text"], "size_x": si["x"], "values": {}}
            else:
                # 同一行出现多个尺码词时，优先取更靠左的（更可能是真正“尺码列”）
                if si["x"] < rows[y_key].get("size_x", 10**9):
                    rows[y_key]["Size"] = si["text"]
                    rows[y_key]["size_x"] = si["x"]

        # 关键改造：即便某行尺码字母没识别到，也先为该 y 行创建容器，避免中间码整行丢失
        for vi in value_items:
            y_key = vi["y"] // 30 * 30
            if y_key not in rows:
                rows[y_key] = {"Size": "", "size_x": 10**9, "values": {}}
        
        # 分配数值到行和列
        for vi in value_items:
            best_row = None
            best_dist = float("inf")
            for y_key in sorted(rows.keys()):
                dist = abs(vi["y"] - y_key)
                if dist < best_dist:
                    best_dist = dist
                    best_row = y_key
            
            if best_row is not None and best_dist < 130:
                best_field = None
                best_field_dist = float("inf")
                for fi in field_items:
                    dist = abs(vi["x"] - fi["x"])
                    if dist < best_field_dist:
                        best_field_dist = dist
                        best_field = fi["field"]
                
                if best_field and best_field_dist < 200:
                    rows[best_row]["values"][best_field] = vi["text"]

        # 尺码补全：当中间码字母缺失时，按上下已识别尺码推断（例如 S _ L -> S M L）
        ordered_rows = sorted(
            [(yk, rv) for yk, rv in rows.items() if len(rv["values"]) >= 2],
            key=lambda x: x[0],
        )
        if ordered_rows:
            row_sizes = []
            for _, rv in ordered_rows:
                sz = rv.get("Size", "")
                row_sizes.append(sz if sz in self.size_order else "")

            for i, sz in enumerate(row_sizes):
                if sz:
                    continue
                prev_idx = None
                next_idx = None
                for j in range(i - 1, -1, -1):
                    if row_sizes[j]:
                        prev_idx = j
                        break
                for j in range(i + 1, len(row_sizes)):
                    if row_sizes[j]:
                        next_idx = j
                        break

                guess = ""
                if prev_idx is not None and next_idx is not None:
                    p = row_sizes[prev_idx]
                    n = row_sizes[next_idx]
                    if p in self.size_order and n in self.size_order:
                        p_i = self.size_order.index(p)
                        n_i = self.size_order.index(n)
                        if n_i - p_i == 2:
                            guess = self.size_order[p_i + 1]
                elif prev_idx is None and next_idx is not None:
                    n = row_sizes[next_idx]
                    if n in self.size_order:
                        n_i = self.size_order.index(n)
                        if n_i > 0:
                            guess = self.size_order[n_i - 1]
                elif next_idx is None and prev_idx is not None:
                    p = row_sizes[prev_idx]
                    if p in self.size_order:
                        p_i = self.size_order.index(p)
                        if p_i + 1 < len(self.size_order):
                            guess = self.size_order[p_i + 1]

                if guess:
                    yk, rv = ordered_rows[i]
                    rows[yk]["Size"] = guess
                    rows[yk]["size_x"] = min(rows[yk].get("size_x", 10**9), 0)
                    row_sizes[i] = guess
        
        # 汇总尺码数据 - 处理可能的重复尺码
        for y_key, row in rows.items():
            # 优先避免漏尺码：即使该行只有少量字段，也保留首次出现的尺码行
            if len(row["values"]) == 0 and row["Size"] in self.ocr_recognized:
                continue
            ocr_size = row["Size"]
            for single_size in ocr_size.split(","):
                single_size = single_size.strip()
                if single_size not in self.ocr_recognized:
                    self.ocr_recognized[single_size] = {}
                for field, value in row["values"].items():
                    if field in all_fields:
                        self.ocr_recognized[single_size][field] = value
        
        if self.ocr_recognized:
            # 1. 自动勾选识别到的尺码
            for size in self.size_order:
                var = self.size_vars.get(size)
                if var:
                    is_recognized = size in self.ocr_recognized
                    var.setChecked(is_recognized)
                    if is_recognized:
                        self.log_signal.emit(f"  自动勾选: {size}")
            
            # 2. 重建表格（带OCR数据）
            self.rebuild_table()
            
            # 3. 手动填充数据到表格（确保填充）
            self._fill_ocr_data_to_table()
            
            self.log_signal.emit(f"✓ OCR识别完成，已勾选: {list(self.ocr_recognized.keys())}")
        else:
            self.log_signal.emit("⚠ 未识别到尺码数据")
            # 即使没识别到数据，也要重建表格（使用默认尺码）
            self.rebuild_table()
        self._record_excel_image_size_diff()
    
    def reload_current(self):
        """重新识别当前款号"""
        if not self.results or self.current_idx >= len(self.results):
            return
        
        item = self.results[self.current_idx]
        img_path = item.get("path", "")
        
        if not img_path or not os.path.exists(img_path):
            self.log_signal.emit(f"图片不存在: {img_path}")
            return
        
        self.log_signal.emit(f"重新识别: {item.get('code', '')}")
        
        # 清空之前的数据
        self.ocr_recognized = {}
        
        self.ocr_image(img_path)
    
    def _fill_ocr_data_to_table(self):
        """手动将OCR数据填充到表格 - 识别到的填入，无数据的必填字段填/"""
        ocr_data = self.ocr_recognized if hasattr(self, 'ocr_recognized') else {}
        
        if not ocr_data:
            self.log_signal.emit("没有OCR数据可填充")
            return
        
        self.log_signal.emit(f"开始填充OCR数据到表格...")
        
        # 获取必填字段
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        if prod_type == "上装":
            required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
        else:
            required_fields = ["腰围", "臀围", "裤长", "裤脚围"]
        
        # 遍历表格每一行
        for row in range(self.size_table.rowCount()):
            size_item = self.size_table.item(row, 0)
            if not size_item:
                continue
            
            size = size_item.text().strip()
            if not size:
                continue
            
            # 获取该尺码的OCR数据
            size_data = ocr_data.get(size, {})
            self.log_signal.emit(f"  尺码 {size}: {size_data}")
            
            # 填充每一列
            for col in range(1, self.size_table.columnCount()):  # 从1开始，跳过尺码列
                header = self.size_table.horizontalHeaderItem(col)
                if not header:
                    continue
                
                field = header.text()
                
                # 如果有识别数据就填入，否则必填字段填"/"
                if field in size_data and size_data[field]:
                    value = size_data[field]
                    item = self.size_table.item(row, col)
                    if item:
                        item.setText(value)
                        self.log_signal.emit(f"    填充 {field} = {value}")
                elif field in required_fields:
                    # 必填字段无数据填"/"
                    item = self.size_table.item(row, col)
                    if item:
                        item.setText("/")
                        self.log_signal.emit(f"    填充 {field} = / (必填无数据)")
                # 非必填字段留空
    
    def clear_results(self):
        """清空当前表格数据 - 不影响其他款"""
        self.ocr_recognized = {}  # 清空当前款OCR数据
        self.size_table.setRowCount(0)  # 清空表格
        self.log_signal.emit("已清空当前表格数据")
    
    def save_current(self):
        """保存当前款号的编辑结果"""
        if not self.results or self.current_idx >= len(self.results):
            QMessageBox.warning(self, "提示", "没有识别结果")
            return
        
        if self.size_table.rowCount() == 0:
            QMessageBox.warning(self, "提示", "表格中没有数据")
            return
        
        code = self.results[self.current_idx]["code"]
        prod_type = "上装" if self.type_top.isChecked() else "裤装"
        
        # 收集表格数据
        data = {"款号": code, "类型": prod_type}
        
        for row in range(self.size_table.rowCount()):
            size_item = self.size_table.item(row, 0)
            if not size_item or not size_item.text():
                continue
            size = size_item.text()
            
            # 号型
            number_item = self.size_table.item(row, 1)
            number = number_item.text() if number_item else ""
            data[f"{size}_号型"] = number
            
            # 其他字段 - 保存所有字段值（包括空值标记/）
            for col in range(2, self.size_table.columnCount()):
                field = self.size_table.horizontalHeaderItem(col).text()
                item = self.size_table.item(row, col)
                value = item.text() if item else ""
                if value:
                    data[f"{size}_{field}"] = value
        
        # 获取唯品类目
        data["唯品类目"] = self.results[self.current_idx].get("category", "")
        
        # 更新或添加保存数据（去重）
        self.confirmed_results = [d for d in self.confirmed_results if d.get("款号") != code]
        self.confirmed_results.append(data)
        self.log_signal.emit(f"✓ 已保存: {code}，共 {len(data)-2} 个字段")
        self.log_signal.emit(f"  保存详情: { {k:v for k,v in data.items() if k not in ['款号','类型','唯品类目']} }")
        
        # 更新结果列表
        self.update_result_list()
    
    def update_result_list(self):
        """更新已识别结果列表 - 显示所有款号和状态"""
        # 已确认的款号集合
        confirmed_codes = set(d.get("款号") for d in self.confirmed_results)
        
        # 设置行数
        self.result_list.setRowCount(len(self.results))
        
        # 添加所有项
        for i, item in enumerate(self.results):
            code = item.get("code", "-")
            category = item.get("category", "-")  # 唯品类目
            prod_type = item.get("type", "-")
            size_type = "上装" if "上装" in prod_type else ("裤装" if "裤装" in prod_type else "-")
            
            # 根据状态显示
            if code in confirmed_codes:
                status = "✓ 已确认"
                status_color = "#4CAF50"
            else:
                is_current = "▶" if i == self.current_idx else "○"
                status = f"{is_current} 待处理"
                status_color = "#FF9800"
            
            self.result_list.setItem(i, 0, QTableWidgetItem(code))
            self.result_list.setItem(i, 1, QTableWidgetItem(category))
            self.result_list.setItem(i, 2, QTableWidgetItem(size_type))
            self.result_list.setItem(i, 3, QTableWidgetItem(status))
        
        # 更新底部统计
        total = len(self.results)
        confirmed = len(self.confirmed_results)
        pending = total - confirmed
        self.st_status.setText(f"总计: {total} | 已确认: {confirmed} | 待处理: {pending}")
    
    def export_current(self):
        """导出当前款号的尺码表 - 唯品会模板格式"""
        self.log_signal.emit(f"[DEBUG] export_current 开始执行，current_idx={self.current_idx}, results长度={len(self.results) if self.results else 0}")
        if not self.results or self.current_idx >= len(self.results):
            self.log_signal.emit("[DEBUG] export_current: 没有识别结果，返回")
            QMessageBox.warning(self, "提示", "没有识别结果")
            return
        
        if self.size_table.rowCount() == 0:
            self.log_signal.emit("[DEBUG] export_current: 表格中没有数据，返回")
            QMessageBox.warning(self, "提示", "表格中没有数据")
            return
        
        save_dir = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if not save_dir:
            self.log_signal.emit("[DEBUG] export_current: 用户取消了目录选择，返回")
            return
        
        try:
            import xlwt
            code = self.results[self.current_idx]["code"]
            prod_type = "上装" if self.type_top.isChecked() else "裤装"
            category = self.results[self.current_idx].get("category", "")
            
            # 检查是否已保存（确认）
            confirmed_data = None
            for confirmed in self.confirmed_results:
                if confirmed.get("款号") == code:
                    confirmed_data = confirmed
                    break
            
            if confirmed_data:
                # 使用保存后的数据
                data = confirmed_data.copy()
                self.log_signal.emit(f"  {code}: 使用保存后的数据")
            else:
                # 使用表格中的当前数据（用户可能修改过）
                data = {"款号": code, "类型": prod_type, "唯品类目": category}
                
                for row in range(self.size_table.rowCount()):
                    size_item = self.size_table.item(row, 0)
                    if not size_item or not size_item.text():
                        continue
                    size = size_item.text()
                    
                    # 号型
                    number_item = self.size_table.item(row, 1)
                    if number_item and number_item.text():
                        data[f"{size}_号型"] = number_item.text()
                    
                    # 其他字段
                    for col in range(2, self.size_table.columnCount()):
                        field = self.size_table.horizontalHeaderItem(col).text()
                        item = self.size_table.item(row, col)
                        if item and item.text():
                            data[f"{size}_{field}"] = item.text()
            
            # 检查必填字段是否为空
            missing_fields = []
            for size in data:
                if size in ["款号", "类型", "唯品类目"]:
                    continue
                if "_" not in size:
                    continue
                size_val, field = size.split("_", 1)
                if field in required_fields and not data[size]:
                    missing_fields.append(f"{size_val}的{field}")
            
            if missing_fields:
                QMessageBox.warning(self, "警告", f"以下必填字段为空：\n{', '.join(missing_fields)}\n\n请手动补齐后再导出")
                return
            
            # 创建Excel - 唯品会模板格式
            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet("尺码表")
            
            # 样式
            style = xlwt.easyxf("font: name Arial, height 200; align: horiz center, vert center;")
            
            # 获取字段
            if prod_type == "上装":
                fields = ["肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
                required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
            else:
                fields = ["腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
                required_fields = ["腰围", "臀围", "裤长", "裤脚围"]
            
            row = 0
            # 第一行：唯品会模板格式
            ws.write(row, 0, "尺码模板名称", style)
            ws.write(row, 1, code, style)
            ws.write(row, 2, "尺码模板ID", style)
            ws.write(row, 3, "", style)
            ws.write(row, 4, "绑定三级分类名称", style)
            ws.write(row, 5, category, style)
            for col in range(6, 6 + len(fields)):
                ws.write(row, col, "", style)
            row += 1
            
            # 第二行：表头
            headers = ["尺码", "号型"] + [f"{f}(cm)" if f != "充绒量" else f"{f}(g)" for f in fields]
            for i, h in enumerate(headers):
                ws.write(row, i, h, style)
            row += 1
            
            # 数据行
            size_order = ["S", "M", "L", "XL", "2XL", "3XL"]
            # 过滤掉不在size_order中的尺码
            valid_sizes = [s for s in data.keys() if "_" in s and s.split("_")[0] in size_order]
            # 提取实际存在的尺码
            sizes_present = set()
            for key in data.keys():
                if "_" in key:
                    size_part = key.split("_")[0]
                    if size_part in size_order:
                        sizes_present.add(size_part)
            sorted_sizes = sorted(list(sizes_present), key=lambda x: size_order.index(x))
            
            for size in sorted_sizes:
                # 创建该尺码的数据行
                ws.write(row, 0, size, style)
                ws.write(row, 1, data.get(f"{size}_号型", ""), style)
                for i, field in enumerate(fields):
                    value = data.get(f"{size}_{field}", "")
                    if value:
                        ws.write(row, i + 2, value, style)
                    elif field in required_fields:
                        ws.write(row, i + 2, "/", style)
                row += 1
            
            # 温馨提示
            warm_tip = self.warm_tip.text() if hasattr(self, 'warm_tip') and self.warm_tip.text() else "产品尺码均为手工测量，会存在1-3CM误差，属于正常范围。"
            ws.write(row, 0, "温馨提示", style)
            ws.write(row, 1, warm_tip, style)
            row += 2  # 款号之间空一行
            
            save_path = os.path.join(save_dir, f"尺码表_{code}.xls")
            wb.save(save_path)
            
            self.log_signal.emit(f"✓ 已导出: {save_path}")
            
            # 打开保存目录
            try:
                os.startfile(save_dir)
            except Exception:
                pass
            
            QMessageBox.information(self, "成功", f"已导出到:\n{save_path}")
        except ImportError:
            QMessageBox.critical(self, "错误", "导出功能需要xlwt库，请先安装：pip install xlwt")
            return
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {e}")
            self.log_signal.emit(f"导出出错: {e}")
        finally:
            self.log_signal.emit(f"[DEBUG] export_current 执行完成")
    
    def _parse_ocr_for_export(self, ocr_result, prod_type):
        """解析OCR结果用于导出（与parse_ocr_result类似，但不涉及UI）"""
        import re
        size_data = {}
        ext_sizes = load_external_size_aliases()
        
        if not ocr_result:
            return size_data
        
        # 获取字段列表
        if prod_type == "上装":
            all_fields = ["肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
        else:
            all_fields = ["腰围", "臀围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "大腿围", "充绒量"]
        
        # 构建字段别名映射
        alias_to_field = {}
        for alias, standard in self.field_aliases.items():
            if standard in all_fields:
                alias_to_field[alias] = standard
        
        sorted_aliases = sorted(alias_to_field.keys(), key=len, reverse=True)
        
        # 分类元素
        size_items = []
        field_items = []
        value_items = []
        
        for item in ocr_result:
            text = item.get("text", "").strip()
            x = item.get("x", 0)
            y = item.get("y", 0)
            
            if not text:
                continue
            
            # 跳过排除关键词
            if any(kw in text for kw in self.exclude_keywords):
                continue
            
            # 尺码
            normalized_size = canonicalize_size(text, ext_sizes)
            if normalized_size:
                size_items.append({"text": normalized_size, "x": x, "y": y})
                continue
            
            # 字段
            matched_field = None
            for alias in sorted_aliases:
                if alias in text or text == alias:
                    matched_field = alias_to_field[alias]
                    break
            
            if matched_field:
                field_items.append({"text": text, "field": matched_field, "x": x, "y": y})
                continue
            
            # 数值
            numbers = re.findall(r"\d+\.?\d*", text)
            if numbers and len(text) < 15:
                try:
                    val = float(numbers[0])
                    if 10 < val < 200:
                        value_items.append({"text": numbers[0], "value": val, "x": x, "y": y})
                except Exception:
                    pass
        
        # 改为“内容优先”策略：不再按固定区域硬裁剪，避免尺码超出区域被漏识别
        
        # 按y坐标分行
        rows = {}
        for si in size_items:
            y_key = si["y"] // 30 * 30
            if y_key not in rows:
                rows[y_key] = {"Size": si["text"], "size_x": si["x"], "values": {}}
            else:
                # 同一行出现多个尺码词时，优先取更靠左的（更可能是真正“尺码列”）
                if si["x"] < rows[y_key].get("size_x", 10**9):
                    rows[y_key]["Size"] = si["text"]
                    rows[y_key]["size_x"] = si["x"]

        # 同步修复导出流程：先建立数值行，避免中间尺码行丢失
        for vi in value_items:
            y_key = vi["y"] // 30 * 30
            if y_key not in rows:
                rows[y_key] = {"Size": "", "size_x": 10**9, "values": {}}
        
        # 分配数值
        for vi in value_items:
            best_row = None
            best_dist = float("inf")
            for y_key in sorted(rows.keys()):
                dist = abs(vi["y"] - y_key)
                if dist < best_dist:
                    best_dist = dist
                    best_row = y_key
            
            if best_row is not None and best_dist < 130:
                best_field = None
                best_field_dist = float("inf")
                for fi in field_items:
                    dist = abs(vi["x"] - fi["x"])
                    if dist < best_field_dist:
                        best_field_dist = dist
                        best_field = fi["field"]
                
                if best_field and best_field_dist < 200:
                    rows[best_row]["values"][best_field] = vi["text"]

        # 导出时同样做中间尺码补全
        ordered_rows = sorted(
            [(yk, rv) for yk, rv in rows.items() if len(rv["values"]) >= 1],
            key=lambda x: x[0],
        )
        if ordered_rows:
            row_sizes = []
            for _, rv in ordered_rows:
                sz = rv.get("Size", "")
                row_sizes.append(sz if sz in self.size_order else "")

            for i, sz in enumerate(row_sizes):
                if sz:
                    continue
                prev_idx = None
                next_idx = None
                for j in range(i - 1, -1, -1):
                    if row_sizes[j]:
                        prev_idx = j
                        break
                for j in range(i + 1, len(row_sizes)):
                    if row_sizes[j]:
                        next_idx = j
                        break

                guess = ""
                if prev_idx is not None and next_idx is not None:
                    p = row_sizes[prev_idx]
                    n = row_sizes[next_idx]
                    if p in self.size_order and n in self.size_order:
                        p_i = self.size_order.index(p)
                        n_i = self.size_order.index(n)
                        if n_i - p_i == 2:
                            guess = self.size_order[p_i + 1]
                elif prev_idx is None and next_idx is not None:
                    n = row_sizes[next_idx]
                    if n in self.size_order:
                        n_i = self.size_order.index(n)
                        if n_i > 0:
                            guess = self.size_order[n_i - 1]
                elif next_idx is None and prev_idx is not None:
                    p = row_sizes[prev_idx]
                    if p in self.size_order:
                        p_i = self.size_order.index(p)
                        if p_i + 1 < len(self.size_order):
                            guess = self.size_order[p_i + 1]

                if guess:
                    yk, rv = ordered_rows[i]
                    rows[yk]["Size"] = guess
                    rows[yk]["size_x"] = min(rows[yk].get("size_x", 10**9), 0)
                    row_sizes[i] = guess
        
        # 汇总
        for y_key, row in rows.items():
            # 导出时放宽到至少1个字段，减少首行尺码被漏掉
            if len(row["values"]) < 1:
                continue
            ocr_size = row["Size"]
            for single_size in ocr_size.split(","):
                single_size = single_size.strip()
                if single_size not in size_data:
                    size_data[single_size] = {}
                for field, value in row["values"].items():
                    if field in all_fields:
                        size_data[single_size][field] = value
        
        return size_data
    
    def _export_template_to_file(self, data_list, output_path, type_name, warm_tip, mismatch_codes=None):
        """导出尺码表模板到文件（多个款号在一个表格）"""
        import xlwt
        
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("尺码表")
        mismatch_codes = set(mismatch_codes or [])
        
        if type_name == "上装":
            fields = ["肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
            required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
        else:
            fields = ["腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
            required_fields = ["腰围", "臀围", "裤长", "裤脚围"]
        
        style = xlwt.easyxf("font: name Arial, height 200; align: horiz center, vert center;")
        mismatch_style = xlwt.easyxf(
            "font: name Arial, height 200, colour_index red, bold on; "
            "pattern: pattern solid, fore_colour light_yellow; "
            "align: horiz center, vert center;"
        )
        mismatch_note_style = xlwt.easyxf(
            "font: name Arial, height 200, colour_index red, bold on; "
            "pattern: pattern solid, fore_colour light_yellow; "
            "align: horiz left, vert center;"
        )
        
        row = 0
        
        for item in data_list:
            code = item.get("款号", "")
            category = item.get("唯品类目", "")
            is_mismatch_code = str(code).strip() in mismatch_codes
            
            # 整理尺码数据
            sizes = {}
            for key, value in item.items():
                if key in ["款号", "类型", "唯品类目"]:
                    continue
                if "_" in key:
                    size, field = key.split("_", 1)
                    if size not in sizes:
                        sizes[size] = {}
                    sizes[size][field] = value
            
            # 第一行：唯品会模板格式
            ws.write(row, 0, "尺码模板名称", style)
            ws.write(row, 1, code, mismatch_style if is_mismatch_code else style)
            ws.write(row, 2, "尺码模板ID", style)
            ws.write(row, 3, "", style)
            ws.write(row, 4, "绑定三级分类名称", style)
            ws.write(row, 5, category, style)
            for col in range(6, 6 + len(fields)):
                ws.write(row, col, "", style)
            if is_mismatch_code:
                note_col = 6 + len(fields)
                ws.write(row, note_col, "⚠ 尺码不一致，请重点核对", mismatch_note_style)
            row += 1
            
            # 第二行：表头
            headers = ["尺码", "号型"] + [f"{f}(cm)" if f != "充绒量" else f"{f}(g)" for f in fields]
            for i, h in enumerate(headers):
                ws.write(row, i, h, style)
            row += 1
            
            # 数据行
            size_order = ["S", "M", "L", "XL", "2XL", "3XL"]
            # 过滤掉不在size_order中的尺码
            valid_sizes = [s for s in sizes.keys() if s in size_order]
            sorted_sizes = sorted(valid_sizes, key=lambda x: size_order.index(x))
            
            for size in sorted_sizes:
                data_row = sizes[size]
                number = data_row.get("号型", "")
                ws.write(row, 0, size, style)
                ws.write(row, 1, number, style)
                for i, field in enumerate(fields):
                    value = data_row.get(field, "")
                    if value:
                        ws.write(row, i + 2, value, style)
                    elif field in required_fields:
                        ws.write(row, i + 2, "/", style)
                row += 1
            
            # 温馨提示
            ws.write(row, 0, "温馨提示", style)
            ws.write(row, 1, warm_tip, style)
            row += 2  # 款号之间空一行
        
        wb.save(output_path)
        return output_path
    
    def parse_ocr_for_export(self, ocr_result, prod_type):
        """解析OCR结果用于导出（不涉及UI操作）"""
        size_data = {}
        ext_sizes = load_external_size_aliases()
        
        if not ocr_result:
            return size_data
        
        # 获取必填字段
        if prod_type == "上装":
            required_fields = ["肩宽", "胸围", "衣长", "袖长", "腰围"]
            all_fields = ["肩宽", "胸围", "衣长", "袖长", "领围", "腰围", "下摆围", "袖笼围", "充绒量"]
        else:
            required_fields = ["腰围", "臀围", "裤长", "裤脚围"]
            all_fields = ["腰围", "臀围", "大腿围", "裤长", "前浪", "后浪", "膝围", "裤脚围", "充绒量"]
        
        # 获取尺码和字段的坐标
        sizes_coords = {}  # {size: y坐标}
        fields_coords = {}  # {field: x坐标范围}
        
        for item in ocr_result:
            text = item.get("text", "").strip()
            if not text:
                continue
            
            # 检测尺码（含 XXL、XXXL 等与界面 2XL、3XL 对应）
            cs = canonicalize_size(text, ext_sizes)
            if cs:
                bbox = item.get("bbox", [])
                if len(bbox) >= 4:
                    sizes_coords[cs] = (bbox[1] + bbox[3]) / 2  # y中心
            
            # 检测字段
            for field in all_fields:
                if field in self.field_aliases:
                    aliases = self.field_aliases[field]
                    if text in aliases or text == field:
                        bbox = item.get("bbox", [])
                        if len(bbox) >= 4:
                            x_center = (bbox[0] + bbox[2]) / 2
                            if field not in fields_coords:
                                fields_coords[field] = []
                            fields_coords[field].append(x_center)
        
        # 合并同一字段的坐标（取平均）
        for field in fields_coords:
            coords = fields_coords[field]
            if coords:
                fields_coords[field] = sum(coords) / len(coords)
        
        # 解析每个尺码的数据
        for item in ocr_result:
            text = item.get("text", "").strip()
            if not text:
                continue
            
            bbox = item.get("bbox", [])
            if len(bbox) < 4:
                continue
            
            y_center = (bbox[1] + bbox[3]) / 2
            x_center = (bbox[0] + bbox[2]) / 2
            
            # 判断这个位置属于哪个尺码
            matched_size = None
            for size, y in sizes_coords.items():
                if abs(y_center - y) < 30:  # 容差
                    matched_size = size
                    break
            
            if not matched_size:
                continue
            
            # 初始化尺码数据
            if matched_size not in size_data:
                size_data[matched_size] = {}
                for field in all_fields:
                    size_data[matched_size][field] = ""
            
            # 判断这个位置属于哪个字段
            for field, x in fields_coords.items():
                if abs(x_center - x) < 50:  # 容差
                    # 检查是否是数值
                    import re
                    if re.match(r'^[\d.]+$', text):
                        size_data[matched_size][field] = text
                    break
        
        return size_data
    
    def prev_image(self):
        """上一张"""
        if self.current_idx > 0:
            self.current_idx -= 1
            self.show_current()
    
    def next_image(self):
        """下一张"""
        if self.current_idx < len(self.results) - 1:
            self.current_idx += 1
            self.show_current()
    
    def on_type_change(self):
        """上装/裤装切换 - 如果当前款已保存则用保存数据，否则重建空表格"""
        code = self.results[self.current_idx]["code"] if self.results and self.current_idx < len(self.results) else None
        
        # 检查是否已保存
        saved_data = None
        if code:
            for confirmed in self.confirmed_results:
                if confirmed.get("款号") == code:
                    saved_data = confirmed
                    break
        
        if saved_data:
            # 已保存：用保存的数据
            self.init_table_with_saved_data(saved_data)
        else:
            # 未保存：重建空表格（不OCR）
            self.ocr_recognized = {}
            self.init_table()
        
        self.log_signal.emit(f"已切换到: {'上装' if self.type_top.isChecked() else '裤装'}")
    
    def on_result_select(self, index):
        """双击已识别列表选择款号"""
        row = index.row()
        if row < len(self.results):
            self.current_idx = row
            self.show_current()
    
    def eventFilter(self, obj, event):
        """事件过滤器 - 捕获表格键盘事件"""
        if obj == self.size_table:
            from PySide6.QtCore import QEvent
            if event.type() == QEvent.KeyPress:
                key = event.key()
                # 按键可以正常移动单元格，不需要额外处理
                pass
        return super().eventFilter(obj, event)
    
        # ============ 结束尺码表方法 ============
 
 
    def export_all(self):
        """导出全部款号 - 上装/裤装两个表格，未保存的款号使用OCR重新识别"""
        if not self.results:
            QMessageBox.warning(self, "提示", "没有识别结果，请先进行批量扫描")
            return
        
        # 导出目录：保留选择路径界面；默认定位到表格目录（没有则输入目录）
        excel_path = self.st_excel.text().strip() if hasattr(self, "st_excel") else ""
        folder_path = self.st_folder.text().strip() if hasattr(self, "st_folder") else ""
        if excel_path and os.path.isfile(excel_path):
            base_dir = os.path.dirname(excel_path)
        elif folder_path and os.path.isdir(folder_path):
            base_dir = folder_path
        else:
            base_dir = os.getcwd()
        manual_export_dir = self.st_export_dir.text().strip() if hasattr(self, "st_export_dir") else ""
        save_dir = manual_export_dir or base_dir
        os.makedirs(save_dir, exist_ok=True)
        self.log_signal.emit(f"导出目录: {save_dir}")
        
        warm_tip = self.warm_tip.text() if hasattr(self, 'warm_tip') and self.warm_tip.text() else "产品尺码均为手工测量，会存在1-3CM误差，属于正常范围。"
        
        # 创建进度对话框
        from PySide6.QtCore import Qt
        progress = QProgressDialog("正在导出...", "取消", 0, len(self.results), self)
        progress.setWindowTitle("导出进度")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        self.log_signal.emit(f"开始导出 {len(self.results)} 个款号...")
        
        upper_data = []  # 上装数据
        lower_data = []  # 裤装数据
        failed = []  # 失败的款号
        engine_counter = {"PaddleOCR": 0, "百度OCR": 0, "PaddleX OCR": 0, "other": 0}
        
        # 预创建OCR引擎以提高性能
        ocr = None
        try:
            ocr = self.get_ocr_engine()
        except Exception as e:
            self.log_signal.emit(f"OCR引擎初始化失败: {e}")
        ocr_mode = self._get_ocr_mode()
        if ocr_mode == "cloud":
            self.log_signal.emit("当前导出OCR模式：仅云端（不使用本地模型）")
        else:
            self.log_signal.emit("当前导出OCR模式：自动（优先本地，失败再云端）")
        
        def _ocr_for_export_with_retry(_ocr, _img_path, _mode):
            """导出时OCR兜底：先按当前模式，失败后自动回退。"""
            from PySide6.QtWidgets import QApplication
            plan = []
            if _mode == "cloud":
                plan = ["cloud", "cloud", "auto"]
            elif _mode == "local":
                plan = ["local", "auto"]
            else:
                plan = ["auto", "cloud"]
            last_engine = ""
            tried_chain = []
            for p in plan:
                try:
                    QApplication.processEvents()
                    items, used = _ocr.ocr_image_items(_img_path, engine=p)
                    last_engine = used or ""
                    tried_chain.append(f"{p}->{used or 'none'}")
                    if items:
                        return items, (used or p), tried_chain
                except Exception:
                    tried_chain.append(f"{p}->error")
                    pass
                QApplication.processEvents()
            return None, last_engine, tried_chain

        # 遍历所有识别结果
        for idx, item in enumerate(self.results):
            # 检查用户是否取消
            if progress.wasCanceled():
                self.log_signal.emit("导出已取消")
                return
            
            progress.setValue(idx)
            progress.setLabelText(f"正在处理 {idx+1}/{len(self.results)}: {item.get('code', '')}")
            
            code = item.get("code", "")
            prod_type = item.get("type", "上装")
            category = item.get("category", "")
            img_path = item.get("path", "")
            
            self.log_signal.emit(f"处理 {idx+1}/{len(self.results)}: {code}")
            
            # 检查是否已保存（确认）
            confirmed_data = None
            self.log_signal.emit(f"  查找已保存数据: 款号='{code}', 已保存数={len(self.confirmed_results)}")
            for confirmed in self.confirmed_results:
                confirmed_code = confirmed.get("款号")
                if confirmed_code == code:
                    confirmed_data = confirmed
                    self.log_signal.emit(f"  匹配成功: '{code}' == '{confirmed_code}'")
                    break
                else:
                    self.log_signal.emit(f"  不匹配: '{code}' != '{confirmed_code}'")
            
            if confirmed_data:
                # 使用保存后的数据
                data = confirmed_data.copy()
                self.log_signal.emit(f"  {code}: 使用保存后的数据")
                # 从保存的数据中获取类型
                prod_type = data.get("类型", prod_type)
            else:
                # 使用OCR重新识别
                data = {"款号": code, "类型": prod_type, "唯品类目": category}
                
                if img_path and os.path.exists(img_path) and ocr is not None:
                    try:
                        ocr_result, used_engine, tried_chain = _ocr_for_export_with_retry(ocr, img_path, ocr_mode)
                        
                        if ocr_result:
                            # 解析OCR结果获取尺码数据
                            size_data = self._parse_ocr_for_export(ocr_result, prod_type)
                            for size, fields_data in size_data.items():
                                # 优先用表格映射，兜底用默认映射
                                number = self.get_default_number(size, prod_type, code)
                                data[f"{size}_号型"] = number
                                for field, value in fields_data.items():
                                    data[f"{size}_{field}"] = value
                            self.log_signal.emit(f"  {code}: OCR识别到 {len(size_data)} 个尺码（{used_engine}）")
                            self.log_signal.emit(f"  {code}: 识别链路 {tried_chain}")
                            if used_engine in engine_counter:
                                engine_counter[used_engine] += 1
                            else:
                                engine_counter["other"] += 1
                            if not size_data:
                                # 云端偶发返回碎片文本时兜底：按 Excel 尺码先生成行，避免漏款
                                size_map = self.excel_size_number_map.get(code, {})
                                for k in size_map.keys():
                                    csize = canonicalize_size(k, self._external_size_aliases) or str(k).strip().upper()
                                    if csize:
                                        data[f"{csize}_号型"] = self.get_default_number(csize, prod_type, code)
                                if size_map:
                                    self.log_signal.emit(f"  {code}: OCR无结构化尺码，已按Excel映射兜底导出")
                        else:
                            # 兜底：即使 OCR 失败，也尽量按 Excel 尺码映射导出该款，避免整款丢失
                            size_map = self.excel_size_number_map.get(code, {})
                            if size_map:
                                for k in size_map.keys():
                                    csize = canonicalize_size(k, self._external_size_aliases) or str(k).strip().upper()
                                    if csize:
                                        data[f"{csize}_号型"] = self.get_default_number(csize, prod_type, code)
                                self.log_signal.emit(f"  {code}: OCR失败（链路 {tried_chain}），已按Excel映射兜底导出")
                            else:
                                self.log_signal.emit(f"  {code}: OCR识别失败（链路 {tried_chain}）且无Excel尺码映射，跳过")
                                failed.append(code)
                                continue
                    except Exception as e:
                        self.log_signal.emit(f"  {code}: OCR识别出错: {e}")
                        failed.append(code)
                        continue
                else:
                    if img_path and os.path.exists(img_path):
                        self.log_signal.emit(f"  {code}: OCR引擎不可用，跳过OCR识别")
                    else:
                        self.log_signal.emit(f"  {code}: 图片不存在，跳过")
                    failed.append(code)
                    continue
            
            # 按上装/裤装分类
            if "裤装" in prod_type:
                lower_data.append(data)
            else:
                upper_data.append(data)
        
        progress.setValue(len(self.results))
        
        try:
            import xlwt
            exported = 0
            mismatch_path = ""
            mismatch_codes = {str(r.get("唯品款号", "")).strip() for r in (self._mismatch_rows or []) if str(r.get("唯品款号", "")).strip()}
            
            # 导出上装表格
            if upper_data:
                path = os.path.join(save_dir, "尺码表_成人上装.xls")
                self._export_template_to_file(upper_data, path, "上装", warm_tip, mismatch_codes=mismatch_codes)
                exported += 1
                self.log_signal.emit(f"导出上装: {path}")
            
            # 导出裤装表格
            if lower_data:
                path = os.path.join(save_dir, "尺码表_成人裤装.xls")
                self._export_template_to_file(lower_data, path, "裤装", warm_tip, mismatch_codes=mismatch_codes)
                exported += 1
                self.log_signal.emit(f"导出裤装: {path}")

            # 选择「导出全部」时，若存在不一致项就自动一起导出
            mismatch_path = self._auto_export_size_mismatch_report(save_dir)
            if mismatch_path:
                exported += 1
            
            extra = ""
            mismatch_count = len(mismatch_codes)
            if mismatch_path:
                extra = f"\n⚠ 高亮提示：发现 {len(mismatch_codes)} 个不一致款号，已自动导出「尺码不一致需手改.xlsx」，且尺码表中对应款号已高亮。"
            if failed:
                msg = QMessageBox(self)
                msg.setWindowTitle("导出完成")
                msg.setTextFormat(Qt.RichText)
                msg.setText(
                    f"导出完成，共 {exported} 个文件<br>"
                    f"失败 {len(failed)} 个款号<br>"
                    f"保存位置: {save_dir}<br>"
                    f"{extra}<br>"
                    f"不一致款号数：<span style='color:#D32F2F;font-weight:700;'>{mismatch_count}</span>"
                )
                msg.exec()
            else:
                msg = QMessageBox(self)
                msg.setWindowTitle("导出完成")
                msg.setTextFormat(Qt.RichText)
                msg.setText(
                    f"导出完成，共 {exported} 个文件<br>"
                    f"保存位置: {save_dir}<br>"
                    f"{extra}<br>"
                    f"不一致款号数：<span style='color:#D32F2F;font-weight:700;'>{mismatch_count}</span>"
                )
                msg.exec()
            # 点完导出结果弹窗后再打开目录，避免提示还没看清就被打断
            try:
                os.startfile(save_dir)
            except Exception as e:
                self.log_signal.emit(f"打开目录失败: {e}")
            self.log_signal.emit(
                f"导出识别引擎统计: 本地Paddle={engine_counter['PaddleOCR']}，百度={engine_counter['百度OCR']}，PaddleX={engine_counter['PaddleX OCR']}，其它={engine_counter['other']}"
            )
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {e}")
            self.log_signal.emit(f"导出出错: {e}")

    def save_settings(self):
        self.config["size_ocr_mode"] = self._get_ocr_mode()
        if hasattr(self, "st_export_dir"):
            self.config["size_export_dir"] = self.st_export_dir.text().strip()
 
 
